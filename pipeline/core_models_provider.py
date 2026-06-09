"""Aggregate per-model Price&Uptime&Usage weekly workbooks into one Core Model Provider Excel."""

from __future__ import annotations

from datetime import date, timedelta
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from pipeline import config
from pipeline.core_models_usage import OUTPUT_SUBDIR as CORE_MODELS_SUBDIR, _style_metadata_sheet
from pipeline.logger import get_logger
from pipeline.model_provider_price_uptime import (
    MonitoredModel,
    iso_week_label,
    load_monitored_models,
    workbook_path_for_model,
)
from pipeline.model_provider_usage import week_dates
from pipeline.model_ranking_weekly import format_cn_date, format_data_range
from pipeline.workbook_metadata import resolve_updated_at

logger = get_logger()

PRICE_UPTIME_INPUT_SUBDIR = "Price&Uptime&Usage"
WORKBOOK_BASENAME = "Core Model Provider"
DATE_COLUMN = "日期"
METADATA_SHEET = "Metadata"
EXCEL_SHEET_TITLE_MAX = 31


def output_path_for_week(week_start: date, output_dir: Path | None = None) -> Path:
    base_dir = output_dir or (config.OUTPUT_DIR / CORE_MODELS_SUBDIR)
    return base_dir / f"{WORKBOOK_BASENAME} {iso_week_label(week_start)}.xlsx"


def _sanitize_sheet_title(name: str) -> str:
    invalid = set(r'[]:*?/\\')
    cleaned = "".join("_" if ch in invalid else ch for ch in name.strip())
    return cleaned[:EXCEL_SHEET_TITLE_MAX] or "model"


def _parse_sheet_date(name: str) -> date | None:
    try:
        return date.fromisoformat(name.strip())
    except ValueError:
        return None


def _week_sheet_names(week_start: date, sheet_names: list[str]) -> list[str]:
    allowed = {day.isoformat() for day in week_dates(week_start)}
    dated = []
    for name in sheet_names:
        parsed = _parse_sheet_date(name)
        if parsed and name in allowed:
            dated.append(name)
    return sorted(dated)


def _align_row_to_headers(row: list[Any], headers: list[str], base_headers: list[str]) -> list[Any]:
    """Map a row from one tab's headers onto the aggregated base header order."""
    index_by_header = {header: idx for idx, header in enumerate(headers) if header}
    return [row[index_by_header[header]] if header in index_by_header else None for header in base_headers]


def _read_daily_sheet(ws) -> tuple[list[str], list[list[Any]]]:
    row_iter = ws.iter_rows(values_only=True)
    try:
        header_row = next(row_iter)
    except StopIteration:
        return [], []
    headers = list(header_row)
    rows: list[list[Any]] = []
    width = len(headers)
    for values in row_iter:
        row = list(values[:width]) + [None] * max(0, width - len(values))
        if any(value is not None and str(value).strip() != "" for value in row):
            rows.append(row)
    return headers, rows


def load_model_week_rows(
    source_path: Path,
    week_start: date,
) -> tuple[list[str], list[list[Any]], list[str]]:
    """Return (headers with 日期 first, data rows, included day sheet names)."""
    wb = load_workbook(source_path, data_only=True, read_only=True)
    day_sheets = _week_sheet_names(week_start, wb.sheetnames)
    if not day_sheets:
        wb.close()
        return [DATE_COLUMN], [], []

    base_headers: list[str] | None = None
    all_rows: list[list[Any]] = []

    for sheet_name in day_sheets:
        ws = wb[sheet_name]
        headers, rows = _read_daily_sheet(ws)
        if base_headers is None:
            base_headers = headers
        elif headers != base_headers:
            logger.warning(
                "Header mismatch on %s tab %s; aligning columns to first tab",
                source_path.name,
                sheet_name,
            )
            for row in rows:
                all_rows.append([sheet_name, *_align_row_to_headers(row, headers, base_headers)])
            continue
        for row in rows:
            all_rows.append([sheet_name, *row])

    wb.close()
    headers_out = [DATE_COLUMN] + (base_headers or [])
    return headers_out, all_rows, day_sheets


def _collect_snapshot_dates_from_sources(
    model_sources: list[tuple[MonitoredModel, Path | None, list[str]]],
    week_start: date,
) -> list[date]:
    dates: list[date] = []
    for _, _, day_sheets in model_sources:
        for sheet_name in day_sheets:
            try:
                dates.append(date.fromisoformat(sheet_name))
            except ValueError:
                continue
    if dates:
        return dates
    return [week_start + timedelta(days=6)]


def build_metadata(
    week_start: date,
    model_sources: list[tuple[MonitoredModel, Path | None, list[str]]],
    updated_at: date | None = None,
    *,
    snapshot_dates: list[date] | None = None,
) -> list[tuple[str, str]]:
    resolved = resolve_updated_at(
        updated_at,
        snapshot_dates=snapshot_dates or [week_start + timedelta(days=6)],
    )
    rows: list[tuple[str, str]] = [
        ("数据范围", format_data_range(week_start)),
        ("数据更新时间", format_cn_date(resolved)),
        (
            "数据来源",
            f"汇总自 data/output/{PRICE_UPTIME_INPUT_SUBDIR}/{{model_slug}}/{{model_slug}} {{周}}.xlsx",
        ),
        ("数据周", iso_week_label(week_start)),
        (
            "汇总说明",
            "每个 tab 对应一个 core 模型；tab 内按「日期」列合并该周各日 Price&Uptime&Usage 行（每行一个 provider）。",
        ),
    ]
    for model, path, day_sheets in model_sources:
        if path is None:
            rows.append((f"模型 {model.model_slug}", "（未找到源工作簿，已跳过）"))
            continue
        days_text = "、".join(day_sheets) if day_sheets else "（该周无日期 tab）"
        rows.append((f"模型 {model.model_slug}", f"{path}；包含日期 tab：{days_text}"))
    return rows


def _style_provider_sheet(ws, headers: list[str]) -> None:
    header_fill = PatternFill("solid", fgColor="FCE4D6")
    usage_fill = PatternFill("solid", fgColor="E2F0D9")
    usage_headers = {
        "Provider URL",
        "展示状态",
        "Provider 承接用量",
        "Provider当日总量",
        "承接占比",
    }

    col_by_header = {header: index + 1 for index, header in enumerate(headers) if header}
    for header, col in col_by_header.items():
        cell = ws.cell(1, col)
        cell.font = Font(bold=True)
        cell.fill = usage_fill if header in usage_headers else header_fill
        cell.alignment = Alignment(horizontal="center")
        if header == DATE_COLUMN:
            ws.column_dimensions[get_column_letter(col)].width = 12
        elif header == "Provider":
            ws.column_dimensions[get_column_letter(col)].width = 20
        elif header == "Provider URL":
            ws.column_dimensions[get_column_letter(col)].width = 34
        else:
            ws.column_dimensions[get_column_letter(col)].width = 14

    price_headers = {"Input Price", "Output Price", "Cache Read"}
    share_col = col_by_header.get("承接占比")
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            header = headers[cell.column - 1] if cell.column <= len(headers) else None
            if header in price_headers and isinstance(cell.value, (int, float)):
                cell.number_format = "$0.####"
            elif header == "Uptime" and isinstance(cell.value, (int, float)):
                cell.number_format = '0.0"%"'
            elif header == share_col and isinstance(cell.value, (int, float)):
                cell.number_format = "0.00%"
    ws.freeze_panes = "A2"


def save_core_models_provider_workbook(
    week_start: date,
    model_tables: list[tuple[MonitoredModel, list[str], list[list[Any]]]],
    metadata_rows: list[tuple[str, str]],
    output_dir: Path | None = None,
) -> Path:
    out_path = output_path_for_week(week_start, output_dir=output_dir)
    out_path.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    default = wb.active
    wb.remove(default)

    for model, headers, rows in model_tables:
        title = _sanitize_sheet_title(model.model_slug)
        ws = wb.create_sheet(title)
        ws.append(headers)
        for row in rows:
            ws.append(row)
        _style_provider_sheet(ws, headers)

    meta_ws = wb.create_sheet(METADATA_SHEET)
    meta_ws.append(["字段", "内容"])
    for key, value in metadata_rows:
        meta_ws.append([key, value])
    _style_metadata_sheet(meta_ws)

    wb.save(out_path)
    logger.info("Saved Core Model Provider Excel to %s", out_path)
    return out_path


def generate_core_models_provider_excel(
    week_start: date,
    models: list[MonitoredModel] | None = None,
    price_uptime_dir: Path | None = None,
    output_dir: Path | None = None,
    updated_at: date | None = None,
) -> Path:
    monitored_models = models or load_monitored_models()
    input_root = price_uptime_dir or (config.OUTPUT_DIR / PRICE_UPTIME_INPUT_SUBDIR)

    model_tables: list[tuple[MonitoredModel, list[str], list[list[Any]]]] = []
    model_sources: list[tuple[MonitoredModel, Path | None, list[str]]] = []

    for model in monitored_models:
        source_path = workbook_path_for_model(model, week_start, output_dir=input_root)
        if not source_path.exists():
            logger.warning("Price&Uptime workbook not found for %s: %s", model.model_slug, source_path)
            model_sources.append((model, None, []))
            model_tables.append((model, [DATE_COLUMN], []))
            continue

        headers, rows, day_sheets = load_model_week_rows(source_path, week_start)
        model_sources.append((model, source_path, day_sheets))
        model_tables.append((model, headers, rows))
        logger.info(
            "Loaded %s rows from %s (%d day tabs)",
            model.model_slug,
            source_path.name,
            len(day_sheets),
        )

    snapshot_dates = _collect_snapshot_dates_from_sources(model_sources, week_start)
    metadata = build_metadata(
        week_start,
        model_sources,
        updated_at=updated_at,
        snapshot_dates=snapshot_dates,
    )
    return save_core_models_provider_workbook(
        week_start,
        model_tables,
        metadata,
        output_dir=output_dir,
    )
