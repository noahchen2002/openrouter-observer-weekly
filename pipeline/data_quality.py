"""Data quality checks for Core Models weekly pipeline outputs."""

from __future__ import annotations

from dataclasses import dataclass
from datetime import date, timedelta
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from pipeline import config
from pipeline.core_models_usage import output_path_for_week as usage_output_path_for_week
from pipeline.data_availability import EXPIRED_UNAVAILABLE_TEXT
from pipeline.model_provider_price_uptime import (
    MonitoredModel,
    load_monitored_models,
    list_missing_snapshot_pairs,
    workbook_path_for_model,
)
from pipeline.model_provider_usage import week_dates
from pipeline.utils import parse_compact_number

DISPLAY_STATUS_SHOWN = "已展示"
PRICE_UPTIME_INPUT_SUBDIR = "Price&Uptime&Usage"


@dataclass(frozen=True)
class DataQualityIssue:
    category: str
    model_slug: str | None
    snapshot_date: date | None
    detail: str


@dataclass(frozen=True)
class DataQualityReport:
    week_start: date
    through_date: date
    issues: tuple[DataQualityIssue, ...]

    @property
    def ok(self) -> bool:
        return not self.issues

    def format_summary(self) -> str:
        if self.ok:
            return (
                "Data quality OK.\n"
                f"week_start={self.week_start.isoformat()} through={self.through_date.isoformat()}\n"
            )

        lines = [
            "Data quality anomalies detected.",
            f"week_start={self.week_start.isoformat()} through={self.through_date.isoformat()}",
            f"{len(self.issues)} issue(s):",
        ]
        for issue in self.issues:
            prefix = f"[{issue.category}]"
            if issue.model_slug and issue.snapshot_date:
                lines.append(f"- {prefix} {issue.model_slug} @ {issue.snapshot_date.isoformat()}: {issue.detail}")
            elif issue.model_slug:
                lines.append(f"- {prefix} {issue.model_slug}: {issue.detail}")
            elif issue.snapshot_date:
                lines.append(f"- {prefix} {issue.snapshot_date.isoformat()}: {issue.detail}")
            else:
                lines.append(f"- {prefix} {issue.detail}")
        return "\n".join(lines) + "\n"


def _allowed_days(week_start: date, through_date: date) -> list[date]:
    return [day for day in week_dates(week_start) if day <= through_date]


def _parse_usage_activity(value: Any) -> float | None:
    if value is None or value == "":
        return None
    if isinstance(value, (int, float)):
        return float(value) if value > 0 else None
    text = str(value).strip()
    if text == EXPIRED_UNAVAILABLE_TEXT:
        return None
    parsed = parse_compact_number(text)
    return parsed if parsed and parsed > 0 else None


def _load_model_activity_by_day(
    usage_path: Path,
    week_start: date,
    through_date: date,
) -> dict[str, dict[date, float]]:
    """Return model_id -> {day -> positive activity tokens}."""
    if not usage_path.exists():
        return {}

    allowed = {day.isoformat() for day in _allowed_days(week_start, through_date)}
    activity: dict[str, dict[date, float]] = {}

    wb = load_workbook(usage_path, data_only=True, read_only=True)
    try:
        ws = wb["Usage"]
        headers = [str(cell.value or "") for cell in next(ws.iter_rows(min_row=1, max_row=1))]
        day_cols = [header for header in headers if header in allowed]
        model_id_col = headers.index("模型ID") + 1 if "模型ID" in headers else None
        if model_id_col is None:
            return {}

        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row:
                continue
            model_id = str(row[model_id_col - 1] or "").strip()
            if not model_id:
                continue
            by_day: dict[date, float] = {}
            for header in day_cols:
                col_idx = headers.index(header)
                tokens = _parse_usage_activity(row[col_idx] if col_idx < len(row) else None)
                if tokens:
                    by_day[date.fromisoformat(header)] = tokens
            if by_day:
                activity[model_id] = by_day
    finally:
        wb.close()
    return activity


def _list_incomplete_usage_pairs(
    week_start: date,
    *,
    through_date: date,
    models: list[MonitoredModel],
    price_uptime_output_dir: Path,
) -> list[tuple[MonitoredModel, date]]:
    """Return (model, date) pairs whose tab exists but lacks Usage columns."""
    pairs: list[tuple[MonitoredModel, date]] = []
    missing_pairs = {
        (model.model_slug, snap_date)
        for model, snap_date in list_missing_snapshot_pairs(
            week_start,
            through_date=through_date,
            models=models,
            output_dir=price_uptime_output_dir,
        )
    }
    for snap_date in _allowed_days(week_start, through_date):
        sheet_name = snap_date.isoformat()
        for model in models:
            if (model.model_slug, snap_date) in missing_pairs:
                continue
            path = workbook_path_for_model(model, week_start, output_dir=price_uptime_output_dir)
            if not path.exists():
                continue
            wb = load_workbook(path, read_only=True, data_only=True)
            try:
                if sheet_name not in wb.sheetnames:
                    continue
                headers = [cell.value for cell in wb[sheet_name][1]]
                if "展示状态" not in headers:
                    pairs.append((model, snap_date))
            finally:
                wb.close()
    return pairs


def _list_header_mismatch_pairs(
    week_start: date,
    *,
    through_date: date,
    models: list[MonitoredModel],
    price_uptime_output_dir: Path,
) -> list[tuple[MonitoredModel, date]]:
    pairs: list[tuple[MonitoredModel, date]] = []
    for model in models:
        path = workbook_path_for_model(model, week_start, output_dir=price_uptime_output_dir)
        if not path.exists():
            continue
        wb = load_workbook(path, read_only=True, data_only=True)
        try:
            base_headers: list[Any] | None = None
            for snap_date in _allowed_days(week_start, through_date):
                sheet_name = snap_date.isoformat()
                if sheet_name not in wb.sheetnames:
                    continue
                headers = [cell.value for cell in wb[sheet_name][1]]
                if base_headers is None:
                    base_headers = headers
                elif headers != base_headers:
                    pairs.append((model, snap_date))
        finally:
            wb.close()
    return pairs


def _list_zero_shown_with_activity(
    week_start: date,
    *,
    through_date: date,
    models: list[MonitoredModel],
    price_uptime_output_dir: Path,
    model_activity: dict[str, dict[date, float]],
) -> list[tuple[MonitoredModel, date, int, int, float]]:
    """Return (model, date, shown_count, total_rows, activity_tokens) anomalies."""
    anomalies: list[tuple[MonitoredModel, date, int, int, float]] = []
    for model in models:
        activity_by_day = model_activity.get(model.model_id, {})
        path = workbook_path_for_model(model, week_start, output_dir=price_uptime_output_dir)
        if not path.exists():
            continue
        wb = load_workbook(path, read_only=True, data_only=True)
        try:
            for snap_date in _allowed_days(week_start, through_date):
                activity = activity_by_day.get(snap_date)
                if not activity:
                    continue
                sheet_name = snap_date.isoformat()
                if sheet_name not in wb.sheetnames:
                    continue
                ws = wb[sheet_name]
                headers = [cell.value for cell in ws[1]]
                if "展示状态" not in headers or "Provider" not in headers:
                    continue
                status_col = headers.index("展示状态") + 1
                provider_col = headers.index("Provider") + 1
                shown = 0
                total = 0
                for row_idx in range(2, ws.max_row + 1):
                    provider = str(ws.cell(row_idx, provider_col).value or "").strip()
                    if not provider:
                        continue
                    total += 1
                    if str(ws.cell(row_idx, status_col).value or "").strip() == DISPLAY_STATUS_SHOWN:
                        shown += 1
                if shown == 0 and total > 0:
                    anomalies.append((model, snap_date, shown, total, activity))
        finally:
            wb.close()
    return anomalies


def check_data_quality(
    week_start: date,
    *,
    through_date: date | None = None,
    models: list[MonitoredModel] | None = None,
    price_uptime_output_dir: Path | None = None,
    usage_path: Path | None = None,
) -> DataQualityReport:
    """Run data-quality checks for a week through the given UTC snapshot date."""
    end = through_date or (date.today() - timedelta(days=1))
    monitored_models = models or load_monitored_models()
    input_root = price_uptime_output_dir or (config.OUTPUT_DIR / PRICE_UPTIME_INPUT_SUBDIR)
    resolved_usage = usage_path or usage_output_path_for_week(week_start)
    model_activity = _load_model_activity_by_day(resolved_usage, week_start, end)

    issues: list[DataQualityIssue] = []

    for model, snap_date in list_missing_snapshot_pairs(
        week_start,
        through_date=end,
        models=monitored_models,
        output_dir=input_root,
    ):
        issues.append(
            DataQualityIssue(
                category="missing_tab",
                model_slug=model.model_slug,
                snapshot_date=snap_date,
                detail="daily Price&Uptime tab missing",
            )
        )

    for model, snap_date in _list_incomplete_usage_pairs(
        week_start,
        through_date=end,
        models=monitored_models,
        price_uptime_output_dir=input_root,
    ):
        issues.append(
            DataQualityIssue(
                category="incomplete_usage",
                model_slug=model.model_slug,
                snapshot_date=snap_date,
                detail="tab missing 展示状态 / Usage columns",
            )
        )

    for model, snap_date in _list_header_mismatch_pairs(
        week_start,
        through_date=end,
        models=monitored_models,
        price_uptime_output_dir=input_root,
    ):
        issues.append(
            DataQualityIssue(
                category="header_mismatch",
                model_slug=model.model_slug,
                snapshot_date=snap_date,
                detail="daily tab headers differ from first tab (Provider charts may be empty)",
            )
        )

    for model, snap_date, shown, total, activity in _list_zero_shown_with_activity(
        week_start,
        through_date=end,
        models=monitored_models,
        price_uptime_output_dir=input_root,
        model_activity=model_activity,
    ):
        issues.append(
            DataQualityIssue(
                category="zero_shown_with_activity",
                model_slug=model.model_slug,
                snapshot_date=snap_date,
                detail=(
                    f"0/{total} providers shown as 已展示 but model activity is "
                    f"{activity:,.0f} tokens; re-run Step 3.2 or check provider charts"
                ),
            )
        )

    return DataQualityReport(week_start=week_start, through_date=end, issues=tuple(issues))
