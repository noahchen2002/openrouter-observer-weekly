"""Build a self-contained HTML dashboard from Core Models weekly Excel outputs."""

from __future__ import annotations

import html
import json
import re
import shutil
from datetime import date, datetime, timedelta
from pathlib import Path
from typing import Any
from urllib.parse import quote

from openpyxl import load_workbook

from pipeline import config
from pipeline.core_models_income import output_path_for_week as income_workbook_path
from pipeline.core_models_provider import output_path_for_week as provider_workbook_path
from pipeline.core_models_usage import (
    OUTPUT_SUBDIR,
    _format_tokens_compact,
    output_path_for_week as usage_workbook_path,
)
from pipeline.model_provider_price_uptime import (
    MonitoredModel,
    iso_week_label,
    load_monitored_models,
)
from pipeline.model_provider_usage import week_dates
from pipeline.logger import get_logger
from pipeline.utils import parse_compact_number

logger = get_logger()

DASHBOARD_BASENAME = "Core Models Dashboard"
METADATA_SHEET = "Metadata"
CHART_VENDOR_FILENAME = "chart.umd.min.js"
CHART_VENDOR_SOURCE = config.PROJECT_ROOT / "static" / "vendor" / CHART_VENDOR_FILENAME
DISPLAY_STATUS_SHOWN = "已展示"
TOP_PROVIDER_COUNT = 8
TOP_PROVIDER_SUMMARY_COUNT = 3
TOKEN_T = 1_000_000_000_000
PROVIDER_NO_DATA_TEXT = "暂无「已展示」承接数据，可能需补跑 Usage 或检查 provider 图表。"
_PROVIDER_PRICE_COLUMNS = ("Input Price", "Output Price", "Cache Read")
_PROVIDER_PRICE_LABELS = {
    "Input Price": "Input",
    "Output Price": "Output",
    "Cache Read": "Cache Read",
}

# Model author (OpenRouter model_id prefix) display names and favicon source sites.
_MODEL_AUTHOR_LABELS: dict[str, str] = {
    "deepseek": "DeepSeek",
    "moonshotai": "Moonshot AI",
    "z-ai": "Z.ai",
}
_MODEL_AUTHOR_HOME_URLS: dict[str, str] = {
    "deepseek": "https://www.deepseek.com",
    "moonshotai": "https://www.moonshot.cn",
    "z-ai": "https://z.ai",
}

# Fixed bar colors per provider display name (stable across all model charts).
_PROVIDER_FIXED_COLORS: dict[str, str] = {
    "SiliconFlow": "#6366f1",
    "DeepSeek": "#0891b2",
    "GMICloud": "#0d9488",
    "Alibaba Cloud Int.": "#f97316",
    "StepFun": "#8b5cf6",
    "Moonshot AI": "#ec4899",
    "MiniMax": "#14b8a6",
    "AtlasCloud": "#3b82f6",
    "Xiaomi": "#f59e0b",
    "Z.ai": "#22c55e",
    "Baidu Qianfan": "#2563eb",
    "NovitaAI": "#a855f7",
    "StreamLake": "#64748b",
    "Google": "#4285f4",
    "Anthropic": "#d97706",
    "OpenAI": "#10b981",
    "Groq": "#ef4444",
    "Together": "#7c3aed",
    "Fireworks": "#e11d48",
    "Parasail": "#06b6d4",
}
_PROVIDER_FALLBACK_PALETTE = (
    "#4f46e5",
    "#dc2626",
    "#16a34a",
    "#ca8a04",
    "#9333ea",
    "#db2777",
    "#0d9488",
    "#ea580c",
    "#0369a1",
    "#65a30d",
    "#7c2d12",
    "#5b21b6",
)


def dashboard_output_path(week_start: date, output_dir: Path | None = None) -> Path:
    base = output_dir or (config.OUTPUT_DIR / OUTPUT_SUBDIR)
    return base / f"{DASHBOARD_BASENAME} {iso_week_label(week_start)}.html"


def install_chart_vendor(output_dir: Path) -> str:
    """Copy Chart.js next to the dashboard HTML for offline file:// viewing."""
    if not CHART_VENDOR_SOURCE.exists():
        raise FileNotFoundError(
            f"Chart.js vendor file missing: {CHART_VENDOR_SOURCE}. "
            "Restore static/vendor/chart.umd.min.js in the repository."
        )
    output_dir.mkdir(parents=True, exist_ok=True)
    shutil.copy2(CHART_VENDOR_SOURCE, output_dir / CHART_VENDOR_FILENAME)
    return CHART_VENDOR_FILENAME


def _read_sheet_table(ws) -> tuple[list[str], list[list[Any]]]:
    rows_iter = ws.iter_rows(values_only=True)
    try:
        header_row = next(rows_iter)
    except StopIteration:
        return [], []
    headers = [str(h) if h is not None else "" for h in header_row]
    rows: list[list[Any]] = []
    width = len(headers)
    for values in rows_iter:
        row = list(values[:width]) + [None] * max(0, width - len(values or ()))
        if any(v is not None and str(v).strip() != "" for v in row):
            rows.append(row)
    return headers, rows


def _read_metadata(ws) -> dict[str, str]:
    meta: dict[str, str] = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or row[0] is None:
            continue
        key = str(row[0]).strip()
        value = "" if row[1] is None else str(row[1])
        meta[key] = value
    return meta


def _day_columns(headers: list[str]) -> list[str]:
    return [h for h in headers if re.fullmatch(r"\d{4}-\d{2}-\d{2}", h)]


def _rows_as_dicts(headers: list[str], rows: list[list[Any]]) -> list[dict[str, Any]]:
    return [dict(zip(headers, row)) for row in rows]


def _parse_usage_cell(value: Any) -> float | None:
    if value is None or value == "":
        return None
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).strip()
    if text == "过期无法获取该数据":
        return None
    parsed = parse_compact_number(text)
    return parsed


def _parse_provider_usage_cell(value: Any) -> float | None:
    if value is None or value == "":
        return None
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).strip()
    if text in ("过期无法获取该数据", "未展示"):
        return None
    return parse_compact_number(text)


def _parse_price_cell(value: Any) -> float | None:
    if value is None or value == "":
        return None
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).strip()
    if not text or text in ANOMALY_TEXTS:
        return None
    try:
        return float(text.replace("$", "").replace(",", ""))
    except ValueError:
        return None


def _format_price_usd(value: float) -> str:
    return f"${value:.4f}"


def _parse_income_cell(value: Any) -> float | None:
    if value is None or value == "":
        return None
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).strip()
    if text == "过期无法获取该数据":
        return None
    try:
        return float(text.replace(",", ""))
    except ValueError:
        return None


def load_usage_workbook(path: Path) -> dict[str, Any]:
    if not path.exists():
        raise FileNotFoundError(f"Usage workbook not found: {path}")
    wb = load_workbook(path, data_only=True, read_only=True)
    ws = wb["Usage"]
    headers, rows = _read_sheet_table(ws)
    metadata = _read_metadata(wb[METADATA_SHEET]) if METADATA_SHEET in wb.sheetnames else {}
    wb.close()
    days = _day_columns(headers)
    dict_rows = _rows_as_dicts(headers, rows)
    return {
        "source_path": str(path),
        "headers": headers,
        "rows": rows,
        "dict_rows": dict_rows,
        "days": days,
        "metadata": metadata,
    }


def load_income_workbook(path: Path) -> dict[str, Any]:
    if not path.exists():
        raise FileNotFoundError(f"Income workbook not found: {path}")
    wb = load_workbook(path, data_only=True, read_only=True)
    ws = wb["Income"]
    headers, rows = _read_sheet_table(ws)
    metadata = _read_metadata(wb[METADATA_SHEET]) if METADATA_SHEET in wb.sheetnames else {}
    wb.close()
    days = _day_columns(headers)
    return {
        "source_path": str(path),
        "headers": headers,
        "rows": rows,
        "dict_rows": _rows_as_dicts(headers, rows),
        "days": days,
        "metadata": metadata,
    }


def load_provider_workbook(path: Path) -> dict[str, Any]:
    if not path.exists():
        raise FileNotFoundError(f"Provider workbook not found: {path}")
    wb = load_workbook(path, data_only=True, read_only=True)
    metadata = _read_metadata(wb[METADATA_SHEET]) if METADATA_SHEET in wb.sheetnames else {}
    models: list[dict[str, Any]] = []
    for sheet_name in wb.sheetnames:
        if sheet_name == METADATA_SHEET:
            continue
        headers, rows = _read_sheet_table(wb[sheet_name])
        days = sorted({str(r[0]) for r in rows if r and r[0] and re.fullmatch(r"\d{4}-\d{2}-\d{2}", str(r[0]))})
        models.append({
            "slug": sheet_name,
            "headers": headers,
            "rows": rows,
            "dict_rows": _rows_as_dicts(headers, rows),
            "days": days,
        })
    wb.close()
    return {"source_path": str(path), "models": models, "metadata": metadata}


def _usage_chart_payload(usage: dict[str, Any]) -> dict[str, Any]:
    days = usage["days"]
    labels = [d[5:] for d in days]
    datasets = []
    pie_labels = []
    pie_values = []
    for row in usage["dict_rows"]:
        name = str(row.get("模型名称") or row.get("模型ID") or "")
        daily_t = []
        week_total = 0.0
        for day in days:
            raw = _parse_usage_cell(row.get(day))
            tokens = raw or 0.0
            week_total += tokens
            daily_t.append(round(tokens / TOKEN_T, 4) if tokens else 0.0)
        datasets.append({"label": name, "data": daily_t})
        pie_labels.append(name)
        pie_values.append(round(week_total / TOKEN_T, 4) if week_total else 0.0)
    return {
        "daily": {"labels": labels, "datasets": datasets},
        "share": {"labels": pie_labels, "values": pie_values},
    }


def _income_chart_payload(income: dict[str, Any]) -> dict[str, Any]:
    days = income["days"]
    labels = [d[5:] for d in days]
    datasets = []
    pie_labels = []
    pie_values = []
    for row in income["dict_rows"]:
        name = str(row.get("模型名称") or row.get("模型ID") or "")
        daily_usd = []
        week_total = 0.0
        for day in days:
            raw = _parse_income_cell(row.get(day))
            amount = raw or 0.0
            week_total += amount
            daily_usd.append(round(amount, 2))
        datasets.append({"label": name, "data": daily_usd})
        pie_labels.append(name)
        pie_values.append(round(week_total, 2))
    return {
        "daily": {"labels": labels, "datasets": datasets},
        "share": {"labels": pie_labels, "values": pie_values},
    }


def model_author_slug(model_id: str) -> str:
    clean = model_id.strip().strip("/")
    if "/" not in clean:
        return clean.lower()
    return clean.split("/", 1)[0].lower()


def model_author_label(author_slug: str) -> str:
    return _MODEL_AUTHOR_LABELS.get(author_slug, author_slug.replace("-", " ").title())


def model_author_icon_url(author_slug: str) -> str:
    homepage = _MODEL_AUTHOR_HOME_URLS.get(author_slug) or f"https://openrouter.ai/{author_slug}"
    encoded = quote(homepage, safe="")
    return (
        "https://t0.gstatic.com/faviconV2?client=SOCIAL&type=FAVICON"
        f"&fallback_opts=TYPE,SIZE,URL&url={encoded}&size=128"
    )


def provider_color(provider_name: str) -> str:
    if provider_name in _PROVIDER_FIXED_COLORS:
        return _PROVIDER_FIXED_COLORS[provider_name]
    digest = 0
    for char in provider_name:
        digest = (digest * 31 + ord(char)) & 0xFFFFFFFF
    return _PROVIDER_FALLBACK_PALETTE[digest % len(_PROVIDER_FALLBACK_PALETTE)]


def collect_provider_names(provider: dict[str, Any]) -> list[str]:
    names: set[str] = set()
    for model in provider["models"]:
        names.update(_aggregate_provider_totals(model).keys())
    return sorted(names)


def build_provider_color_map(provider: dict[str, Any]) -> dict[str, str]:
    return {name: provider_color(name) for name in collect_provider_names(provider)}


def build_model_meta_by_slug(monitored_models: list[MonitoredModel]) -> dict[str, dict[str, str]]:
    meta: dict[str, dict[str, str]] = {}
    for monitored in monitored_models:
        author_slug = model_author_slug(monitored.model_id)
        meta[monitored.model_slug] = {
            "model_id": monitored.model_id,
            "author_slug": author_slug,
            "author_label": model_author_label(author_slug),
            "icon_url": model_author_icon_url(author_slug),
        }
    return meta


def _render_model_provider_heading(slug: str, model_meta: dict[str, dict[str, str]]) -> str:
    meta = model_meta.get(slug)
    if not meta:
        return f"<span>{_escape(slug)}</span>"
    icon_url = _escape(meta["icon_url"])
    author_label = _escape(meta["author_label"])
    model_id = _escape(meta["model_id"])
    slug_text = _escape(slug)
    return f"""
          <span class="flex min-w-0 items-center gap-3">
            <img src="{icon_url}" alt="" width="32" height="32"
              class="h-8 w-8 shrink-0 rounded-lg border border-slate-200 bg-white object-contain p-0.5"
              loading="lazy" decoding="async" />
            <span class="min-w-0">
              <span class="block truncate">{slug_text}</span>
              <span class="block truncate text-xs font-normal text-slate-500">{author_label} · {model_id}</span>
            </span>
          </span>
    """


def _aggregate_provider_totals(model: dict[str, Any]) -> dict[str, float]:
    totals: dict[str, float] = {}
    for row in model["dict_rows"]:
        if str(row.get("展示状态") or "") != DISPLAY_STATUS_SHOWN:
            continue
        provider = str(row.get("Provider") or "")
        if not provider:
            continue
        amount = _parse_provider_usage_cell(row.get("Provider 承接用量")) or 0.0
        totals[provider] = totals.get(provider, 0.0) + amount
    return totals


def _provider_model_chart_payload(
    model: dict[str, Any],
    *,
    provider_color_map: dict[str, str] | None = None,
) -> dict[str, Any]:
    days = model["days"]
    labels = [d[5:] for d in days]
    totals_by_provider = _aggregate_provider_totals(model)
    by_day_provider: dict[str, dict[str, float]] = {d: {} for d in days}

    for row in model["dict_rows"]:
        if str(row.get("展示状态") or "") != DISPLAY_STATUS_SHOWN:
            continue
        day = str(row.get("日期") or "")
        provider = str(row.get("Provider") or "")
        amount = _parse_provider_usage_cell(row.get("Provider 承接用量")) or 0.0
        if not day or not provider:
            continue
        by_day_provider.setdefault(day, {})[provider] = amount

    top_providers = [
        name
        for name, _ in sorted(totals_by_provider.items(), key=lambda item: item[1], reverse=True)[
            :TOP_PROVIDER_COUNT
        ]
    ]
    datasets = []
    for provider in top_providers:
        data = []
        for day in days:
            data.append(
                round((by_day_provider.get(day, {}).get(provider) or 0.0) / 1_000_000_000, 4)
            )
        color = (provider_color_map or {}).get(provider) or provider_color(provider)
        datasets.append({"label": provider, "data": data, "color": color})
    return {"labels": labels, "datasets": datasets, "slug": model["slug"]}


def _escape(value: Any) -> str:
    if value is None:
        return ""
    return html.escape(str(value))


def _format_cell(header: str, value: Any) -> str:
    if value is None or value == "":
        return ""
    text = str(value)
    if header == "Provider URL" and text.startswith("http"):
        safe = _escape(text)
        return (
            f'<a href="{safe}" target="_blank" rel="noopener" '
            f'class="text-indigo-600 hover:text-indigo-800 hover:underline">{safe}</a>'
        )
    if header == "承接占比（%）":
        return _escape(text)
    if header in ("占比（%）", "承接占比") and isinstance(value, (int, float)):
        return f"{float(value) * 100:.2f}%"
    if isinstance(value, float):
        if header in ("Input Price", "Output Price", "Cache Read"):
            return f"${value:.4f}"
        if re.fullmatch(r"\d{4}-\d{2}-\d{2}", header) or header in ("周合计",):
            return f"{value:,.2f}"
        return f"{value:.4g}"
    return _escape(text)


ANOMALY_TEXTS = frozenset({"过期无法获取该数据", "未展示", "用量少未查询"})
NUMERIC_LEFT_HEADERS = frozenset(
    {"模型ID", "模型名称", "Provider", "Provider URL", "Region", "Quantization", "展示状态", "日期"},
)


def _is_anomaly_cell(header: str, value: Any) -> bool:
    text = str(value or "").strip()
    if text in ANOMALY_TEXTS:
        return True
    if header == "展示状态" and text and text != DISPLAY_STATUS_SHOWN:
        return True
    if header in ("承接占比", "占比（%）") and text in ("", "—"):
        return True
    return False


# 英文表头 → 中文展示名（仅影响表头文字；列逻辑/取数仍用原始英文 header）。
_HEADER_ZH = {
    "Provider": "厂商",
    "Provider URL": "厂商链接",
    "Input Price": "输入价",
    "Output Price": "输出价",
    "Cache Read": "缓存价",
    "Latency": "延迟",
    "Throughput": "吞吐",
    "Uptime": "稳定性",
    "Total Context": "上下文",
    "Max Output": "最大输出",
}


def _render_table(headers: list[str], rows: list[list[Any]], *, numeric_right: bool = True) -> str:
    th_cells = "".join(
        f'<th scope="col" class="px-4 py-3 text-left text-xs font-semibold uppercase tracking-wide text-slate-600">{_escape(_HEADER_ZH.get(h, h))}</th>'
        for h in headers
    )
    body_rows = []
    for row in rows:
        tds = []
        for header, value in zip(headers, row):
            classes = ["px-4", "py-2.5", "text-sm", "text-slate-700", "border-t", "border-slate-100"]
            if numeric_right and header not in NUMERIC_LEFT_HEADERS:
                classes.extend(["text-right", "tabular-nums"])
            if _is_anomaly_cell(header, value):
                classes.extend(["bg-amber-50", "text-amber-900", "font-medium"])
            tds.append(f'<td class="{" ".join(classes)}">{_format_cell(header, value)}</td>')
        body_rows.append(
            f'<tr class="transition-colors hover:bg-slate-50/80">{"".join(tds)}</tr>',
        )
    return (
        '<div class="mt-4 overflow-x-auto rounded-lg border border-slate-200">'
        '<table class="min-w-full divide-y divide-slate-200">'
        f'<thead class="bg-slate-50"><tr>{th_cells}</tr></thead>'
        f'<tbody class="divide-y divide-slate-100 bg-white">{"".join(body_rows)}</tbody>'
        "</table></div>"
    )


def _provider_rows_for_day(model: dict[str, Any], day: str) -> list[list[Any]]:
    usage_idx = model["headers"].index("Provider 承接用量") if "Provider 承接用量" in model["headers"] else -1
    provider_idx = model["headers"].index("Provider") if "Provider" in model["headers"] else -1
    rows = [row for row in model["rows"] if row and str(row[0]) == day]

    def sort_key(row: list[Any]) -> tuple[float, str]:
        raw_usage = row[usage_idx] if usage_idx >= 0 and usage_idx < len(row) else None
        usage = _parse_provider_usage_cell(raw_usage)
        provider = str(row[provider_idx] or "") if provider_idx >= 0 and provider_idx < len(row) else ""
        return (usage if usage is not None else -1.0, provider)

    return sorted(rows, key=sort_key, reverse=True)


_PROVIDER_TABLE_HIDDEN_HEADERS = frozenset({"展示状态", "Region", "Quantization"})

_PROVIDER_TABLE_HEADER_GROUPS: tuple[tuple[str, ...], ...] = (
    ("日期", "Provider"),
    ("Provider 承接用量", "Provider当日总量", "承接占比"),
    ("Input Price", "Output Price", "Cache Read"),
    ("Region", "Quantization", "Latency", "Throughput", "Uptime", "Total Context", "Max Output"),
    ("Provider URL",),
)


def _provider_display_header_order(headers: list[str]) -> list[str]:
    """用量 → 价格 → 性能 → Provider 链接；其余列插在性能组之后、链接之前。"""
    visible_headers = [h for h in headers if h not in _PROVIDER_TABLE_HIDDEN_HEADERS]
    known = {header for group in _PROVIDER_TABLE_HEADER_GROUPS for header in group}
    ordered: list[str] = []
    seen: set[str] = set()
    for group in _PROVIDER_TABLE_HEADER_GROUPS:
        for header in group:
            if header in visible_headers and header not in seen:
                ordered.append(header)
                seen.add(header)
    for header in visible_headers:
        if header not in seen and header not in known:
            ordered.append(header)
            seen.add(header)
    return ordered


def _provider_table_for_display(model: dict[str, Any], rows: list[list[Any]]) -> str:
    headers = list(model["headers"])
    display_headers = _provider_display_header_order(headers)

    header_index = {header: idx for idx, header in enumerate(headers)}
    reordered_rows = []
    for row in rows:
        padded = list(row[: len(headers)]) + [None] * max(0, len(headers) - len(row))
        reordered_rows.append([padded[header_index[header]] for header in display_headers])
    return _render_table(display_headers, reordered_rows)


def _render_provider_daily_tabs(model: dict[str, Any]) -> str:
    if not model["days"]:
        return _provider_table_for_display(model, model["rows"])

    safe_slug = re.sub(r"[^a-zA-Z0-9_-]+", "-", str(model["slug"]))
    buttons = []
    panels = []
    for idx, day in enumerate(model["days"]):
        active_class = "bg-indigo-600 text-white shadow-sm"
        inactive_class = "bg-white text-slate-600 hover:bg-slate-50"
        buttons.append(
            f'<button type="button" class="provider-tab rounded-lg border border-slate-200 px-3 py-1.5 text-sm font-medium '
            f'{" ".join([active_class if idx == 0 else inactive_class])}" '
            f'data-provider-tab="{_escape(safe_slug)}-{idx}" aria-pressed="{str(idx == 0).lower()}">{_escape(day[5:])}</button>'
        )
        rows = _provider_rows_for_day(model, day)
        panels.append(
            f'<div class="provider-tab-panel{" hidden" if idx else ""}" data-provider-panel="{_escape(safe_slug)}-{idx}">'
            f'{_provider_table_for_display(model, rows)}</div>'
        )

    return (
        '<div class="mt-4">'
        '<div class="flex flex-wrap gap-2" role="tablist" aria-label="按日期查看 Provider 承接排名">'
        f'{"".join(buttons)}'
        "</div>"
        f'{"".join(panels)}'
        "</div>"
    )


def _row_week_usage_t(row: dict[str, Any]) -> float:
    t_cell = row.get("换算后用量（T）")
    if isinstance(t_cell, (int, float)):
        return float(t_cell)
    parsed = _parse_usage_cell(row.get("周合计"))
    return (parsed or 0.0) / TOKEN_T


def _row_week_income(row: dict[str, Any]) -> float:
    week = row.get("周合计")
    if isinstance(week, (int, float)):
        return float(week)
    return _parse_income_cell(week) or 0.0


def _top_model_row(rows: list[dict[str, Any]], metric_fn) -> tuple[str, float]:
    best_name = ""
    best_value = -1.0
    for row in rows:
        name = str(row.get("模型名称") or row.get("模型ID") or "")
        value = metric_fn(row)
        if value > best_value:
            best_value = value
            best_name = name
    return best_name, max(best_value, 0.0)


def _compute_kpis(usage: dict[str, Any], income: dict[str, Any], provider: dict[str, Any]) -> dict[str, Any]:
    usage_rows = usage["dict_rows"]
    income_rows = income["dict_rows"]
    total_usage_t = sum(_row_week_usage_t(r) for r in usage_rows)
    total_income = sum(_row_week_income(r) for r in income_rows)
    top_usage_name, top_usage_t = _top_model_row(usage_rows, _row_week_usage_t)
    top_income_name, top_income = _top_model_row(income_rows, _row_week_income)
    usage_share_pct = (top_usage_t / total_usage_t * 100) if total_usage_t else 0.0
    income_share_pct = (top_income / total_income * 100) if total_income else 0.0
    provider_shown = sum(
        1
        for model in provider["models"]
        for row in model["dict_rows"]
        if str(row.get("展示状态") or "") == DISPLAY_STATUS_SHOWN
    )
    return {
        "model_count": len(usage_rows),
        "total_usage_t": total_usage_t,
        "total_income": total_income,
        "top_usage_name": top_usage_name,
        "top_usage_t": top_usage_t,
        "usage_share_pct": usage_share_pct,
        "top_income_name": top_income_name,
        "top_income": top_income,
        "income_share_pct": income_share_pct,
        "provider_shown": provider_shown,
    }


def _build_headline(kpis: dict[str, Any]) -> str:
    if kpis["model_count"] == 0:
        return "本周暂无核心模型监控数据，请先运行 Usage / Provider / Income 流水线。"
    parts = [
        f"共监控 {kpis['model_count']} 个核心模型，全周总用量 {kpis['total_usage_t']:.2f}T、总收入 ${kpis['total_income']:,.0f}。",
    ]
    if kpis["top_usage_name"]:
        parts.append(
            f"{kpis['top_usage_name']} 用量领先（{kpis['usage_share_pct']:.1f}%），"
            f"收入榜首为 {kpis['top_income_name']}（{kpis['income_share_pct']:.1f}%）。",
        )
    if kpis["provider_shown"]:
        parts.append(f"Provider 维度共 {kpis['provider_shown']} 条「已展示」承接记录可供对比。")
    return "".join(parts)


def _usage_daily_insight(usage: dict[str, Any]) -> str:
    days = usage["days"]
    if len(days) < 2:
        return "观察各模型在日粒度上的用量波动，识别突发流量。"
    last_day, first_day = days[-1], days[0]
    max_jump_name, max_jump = "", 0.0
    for row in usage["dict_rows"]:
        first = _parse_usage_cell(row.get(first_day)) or 0.0
        last = _parse_usage_cell(row.get(last_day)) or 0.0
        jump = last - first
        if jump > max_jump:
            max_jump = jump
            max_jump_name = str(row.get("模型名称") or "")
    if max_jump_name and max_jump > 0:
        return f"{max_jump_name} 从周初到周末增量最大，需关注是否为新上线或活动放量。"
    return "周内各模型用量相对平稳，未见单一模型极端拉升。"


def _usage_share_insight(kpis: dict[str, Any]) -> str:
    if kpis["usage_share_pct"] >= 50:
        return f"用量高度集中于 {kpis['top_usage_name']}，其余模型份额有限，路由策略偏单一。"
    if kpis["usage_share_pct"] >= 35:
        return f"{kpis['top_usage_name']} 占主导但未过半，核心模型之间仍有一定分流。"
    return "用量分布较分散，多模型并行承接平台流量。"


def _income_daily_insight(income: dict[str, Any]) -> str:
    days = income["days"]
    if not days:
        return "对照每日 paid_usd 走势，评估商业化变现节奏。"
    peak_name, peak_val = "", 0.0
    for row in income["dict_rows"]:
        for day in days:
            val = _parse_income_cell(row.get(day)) or 0.0
            if val > peak_val:
                peak_val = val
                peak_name = str(row.get("模型名称") or "")
    if peak_name:
        return f"单日收入峰值出现在 {peak_name}（约 ${peak_val:,.0f}），可结合当日用量判断单价走势。"
    return "本周收入曲线较平，建议结合上游模型调价与路由变化复盘。"


def _income_share_insight(kpis: dict[str, Any]) -> str:
    return (
        f"{kpis['top_income_name']} 贡献 {kpis['income_share_pct']:.1f}% 周收入，"
        f"与用量份额（{kpis['usage_share_pct']:.1f}%）对照可看出变现效率差异。"
    )


def _sorted_provider_totals(model: dict[str, Any]) -> list[tuple[str, float]]:
    return sorted(_aggregate_provider_totals(model).items(), key=lambda item: item[1], reverse=True)


def _provider_price_snapshot(row: dict[str, Any]) -> tuple[float | None, float | None, float | None]:
    return tuple(_parse_price_cell(row.get(column)) for column in _PROVIDER_PRICE_COLUMNS)


def _provider_price_for_day(
    model: dict[str, Any],
    provider: str,
    day: str,
    *,
    require_shown: bool = True,
) -> tuple[float | None, float | None, float | None] | None:
    """Read provider price snapshot for a day.

    Usage rankings still require 「已展示」; price insight may read any row that
  carries price fields (e.g. prior Sunday was 「未展示」 but prices exist).
    """
    for row in model["dict_rows"]:
        if require_shown and str(row.get("展示状态") or "") != DISPLAY_STATUS_SHOWN:
            continue
        if str(row.get("Provider") or "") != provider:
            continue
        if str(row.get("日期") or "") != day:
            continue
        snapshot = _provider_price_snapshot(row)
        if any(value is not None for value in snapshot):
            return snapshot
    return None


def _providers_for_price_insight(
    model: dict[str, Any],
    *,
    week_start: date,
    prior_model: dict[str, Any] | None,
) -> list[str]:
    """Top3 by shown usage, plus any other provider with Monday vs prior-Sunday price change."""
    ranked = _sorted_provider_totals(model)
    providers = [name for name, _ in ranked[:TOP_PROVIDER_SUMMARY_COUNT]]
    if prior_model is None:
        return providers

    monday = week_start.isoformat()
    prior_sunday = (week_start - timedelta(days=1)).isoformat()
    seen = set(providers)
    monday_providers = {
        str(row.get("Provider") or "").strip()
        for row in model["dict_rows"]
        if str(row.get("日期") or "") == monday and str(row.get("Provider") or "").strip()
    }
    for provider in sorted(monday_providers):
        if provider in seen:
            continue
        monday_prices = _provider_price_for_day(
            model,
            provider,
            monday,
            require_shown=True,
        )
        prior_sunday_prices = _provider_price_for_day(
            prior_model,
            provider,
            prior_sunday,
            require_shown=False,
        )
        if monday_prices is None or prior_sunday_prices is None:
            continue
        if _describe_price_field_changes(prior_sunday_prices, monday_prices):
            providers.append(provider)
            seen.add(provider)
    return providers


def load_prior_week_provider_models_by_slug(
    week_start: date,
    *,
    output_dir: Path | None = None,
) -> dict[str, dict[str, Any]]:
    prior_week_start = week_start - timedelta(days=7)
    prior_path = provider_workbook_path(prior_week_start, output_dir=output_dir)
    if not prior_path.exists():
        return {}
    prior_provider = load_provider_workbook(prior_path)
    return {model["slug"]: model for model in prior_provider["models"]}


def _describe_price_field_changes(
    previous: tuple[float | None, float | None, float | None],
    current: tuple[float | None, float | None, float | None],
) -> list[str]:
    changes: list[str] = []
    for column, previous_value, current_value in zip(_PROVIDER_PRICE_COLUMNS, previous, current):
        if previous_value is None or current_value is None:
            continue
        if abs(previous_value - current_value) > 1e-9:
            label = _PROVIDER_PRICE_LABELS[column]
            changes.append(
                f"{label} {_format_price_usd(previous_value)}→{_format_price_usd(current_value)}",
            )
    return changes


def _describe_monday_vs_prior_sunday_price_change(
    provider: str,
    *,
    prior_sunday_prices: tuple[float | None, float | None, float | None],
    monday_prices: tuple[float | None, float | None, float | None],
    monday_label: str,
) -> str | None:
    field_changes = _describe_price_field_changes(prior_sunday_prices, monday_prices)
    if not field_changes:
        return None
    return f"{provider} 本周一 {monday_label[5:]} 较上周日（{'、'.join(field_changes)}）"


def _build_provider_top3_price_insight(
    model: dict[str, Any],
    *,
    week_start: date,
    prior_model: dict[str, Any] | None = None,
) -> str:
    ranked = _sorted_provider_totals(model)
    if not ranked:
        return PROVIDER_NO_DATA_TEXT

    monday = week_start.isoformat()
    prior_sunday = (week_start - timedelta(days=1)).isoformat()
    insight_providers = _providers_for_price_insight(
        model,
        week_start=week_start,
        prior_model=prior_model,
    )
    changed: list[str] = []
    stable: list[str] = []
    insufficient: list[str] = []

    for provider in insight_providers:
        monday_prices = _provider_price_for_day(model, provider, monday, require_shown=True)
        if monday_prices is None:
            insufficient.append(provider)
            continue
        if prior_model is None:
            insufficient.append(provider)
            continue
        prior_sunday_prices = _provider_price_for_day(
            prior_model,
            provider,
            prior_sunday,
            require_shown=False,
        )
        if prior_sunday_prices is None:
            insufficient.append(provider)
            continue
        description = _describe_monday_vs_prior_sunday_price_change(
            provider,
            prior_sunday_prices=prior_sunday_prices,
            monday_prices=monday_prices,
            monday_label=monday,
        )
        if description:
            changed.append(description)
        else:
            stable.append(provider)

    if insufficient and not changed and not stable:
        names = "、".join(insufficient)
        return (
            f"Top{TOP_PROVIDER_SUMMARY_COUNT}（{names}）缺少本周一或上周日价格记录，"
            "无法对比跨周变动。"
        )

    parts: list[str] = []
    if changed:
        parts.append("；".join(changed))
    if stable:
        parts.append(f"{'、'.join(stable)} 本周一较上周日价格未变")
    if insufficient:
        parts.append(f"{'、'.join(insufficient)} 缺少本周一或上周日价格，未纳入跨周对比")

    body = "；".join(parts)
    return body if body.endswith("。") else f"{body}。"


def _build_provider_top3_summary(model: dict[str, Any]) -> str:
    ranked = _sorted_provider_totals(model)
    if not ranked:
        return PROVIDER_NO_DATA_TEXT

    total = sum(amount for _, amount in ranked)
    top_n = ranked[:TOP_PROVIDER_SUMMARY_COUNT]
    label = f"Top{len(top_n)}" if len(top_n) < TOP_PROVIDER_SUMMARY_COUNT else "Top3"

    parts: list[str] = []
    for index, (name, amount) in enumerate(top_n, start=1):
        share = (amount / total * 100) if total else 0.0
        parts.append(f"{index}) {name} 累计 {_format_tokens_compact(amount)}（{share:.1f}%）")

    top_sum = sum(amount for _, amount in top_n)
    top_share = (top_sum / total * 100) if total else 0.0
    body = "；".join(parts)
    return f"{body}。{label} 合计占已展示承接量 {top_share:.1f}%。"


def _provider_insight(model: dict[str, Any]) -> str:
    ranked = _sorted_provider_totals(model)
    if not ranked:
        return PROVIDER_NO_DATA_TEXT
    total = sum(amount for _, amount in ranked)
    top_name, top_val = ranked[0]
    top_share = (top_val / total * 100) if total else 0.0
    top3 = ranked[:TOP_PROVIDER_SUMMARY_COUNT]
    top3_share = (sum(v for _, v in top3) / total * 100) if total else 0.0
    if top_share >= 70:
        concentration = "供应高度集中"
    elif top3_share >= 50:
        concentration = "Top3 占主导"
    else:
        concentration = "供应较为分散"
    return (
        f"{top_name} 为最大承接方（{top_share:.1f}%），{concentration}；"
        f"下图展示 Top {TOP_PROVIDER_COUNT} 的日趋势。"
    )


def _render_provider_top3_summaries(usage_summary: str, price_insight: str) -> str:
    return f"""
          <p class="mb-4 rounded-md border border-slate-200 bg-slate-50 px-3 py-3 text-sm leading-relaxed text-slate-700">
            <span class="font-medium text-slate-900">Top3 承接总结：</span>{_escape(usage_summary)}
          </p>
          <p class="mb-4 rounded-md border border-slate-200 bg-slate-50 px-3 py-3 text-sm leading-relaxed text-slate-700">
            <span class="font-medium text-slate-900">价格洞察：</span>{_escape(price_insight)}
          </p>
    """


def _render_kpi_card(label: str, value: str, hint: str) -> str:
    return f"""
    <div class="rounded-xl border border-slate-200 bg-white p-5 shadow-sm">
      <p class="text-xs font-medium uppercase tracking-wide text-slate-500">{_escape(label)}</p>
      <p class="mt-2 text-2xl font-semibold tracking-tight text-slate-900">{_escape(value)}</p>
      <p class="mt-1 text-sm text-slate-500">{_escape(hint)}</p>
    </div>
    """


def _render_chart_card(title: str, description: str, insight: str, canvas_id: str) -> str:
    return f"""
    <div class="rounded-xl border border-slate-200 bg-white p-5 shadow-sm">
      <h3 class="text-base font-semibold text-slate-900">{_escape(title)}</h3>
      <p class="mt-1 text-sm leading-relaxed text-slate-500">{_escape(description)}</p>
      <p class="mt-2 rounded-md bg-indigo-50 px-3 py-2 text-sm text-indigo-900"><span class="font-medium">洞察：</span>{_escape(insight)}</p>
      <div class="relative mt-4 h-72 w-full">
        <canvas id="{_escape(canvas_id)}" aria-label="{_escape(title)}"></canvas>
      </div>
    </div>
    """


def _json_script(element_id: str, payload: Any) -> str:
    return (
        f'<script type="application/json" id="{element_id}">'
        f"{json.dumps(payload, ensure_ascii=False)}"
        f"</script>"
    )


def render_dashboard_html(
    week_label: str,
    usage: dict[str, Any],
    provider: dict[str, Any],
    income: dict[str, Any],
    *,
    week_start: date,
    chart_script_src: str = CHART_VENDOR_FILENAME,
    generated_at: datetime | None = None,
    model_meta_by_slug: dict[str, dict[str, str]] | None = None,
    provider_color_map: dict[str, str] | None = None,
    prior_provider_models_by_slug: dict[str, dict[str, Any]] | None = None,
) -> str:
    generated_at = generated_at or datetime.now()
    usage_charts = _usage_chart_payload(usage)
    income_charts = _income_chart_payload(income)
    kpis = _compute_kpis(usage, income, provider)
    headline = _build_headline(kpis)
    meta_range = usage["metadata"].get("数据范围") or income["metadata"].get("数据范围") or ""
    data_updated = (
        usage["metadata"].get("数据更新时间")
        or income["metadata"].get("数据更新时间")
        or ""
    )
    data_source_line = "OpenRouter Activity API · model_income 导出 · Price&Uptime&Usage 汇总"

    kpi_cards = "".join(
        [
            _render_kpi_card("监控模型数", str(kpis["model_count"]), "config/core_models.json"),
            _render_kpi_card("全周总用量", f"{kpis['total_usage_t']:.2f}T", "模型总体用量"),
            _render_kpi_card("全周总收入", f"${kpis['total_income']:,.0f}", "paid_usd 周合计"),
            _render_kpi_card(
                "用量榜首",
                kpis["top_usage_name"] or "—",
                f"{kpis['usage_share_pct']:.1f}% 份额 · {kpis['top_usage_t']:.2f}T",
            ),
            _render_kpi_card(
                "收入榜首",
                kpis["top_income_name"] or "—",
                f"{kpis['income_share_pct']:.1f}% 份额 · ${kpis['top_income']:,.0f}",
            ),
            _render_kpi_card(
                "Provider 已展示行",
                str(kpis["provider_shown"]),
                f"图表取 Top {TOP_PROVIDER_COUNT} 承接方",
            ),
        ],
    )

    model_meta = model_meta_by_slug or {}
    resolved_provider_colors = provider_color_map or build_provider_color_map(provider)
    prior_models = prior_provider_models_by_slug or {}

    provider_sections = []
    for model in provider["models"]:
        chart = _provider_model_chart_payload(
            model,
            provider_color_map=resolved_provider_colors,
        )
        chart_id = f"provider-chart-{model['slug']}"
        provider_sections.append(
            f"""
      <details class="group rounded-xl border border-slate-200 bg-white shadow-sm open:ring-1 open:ring-indigo-100" open>
        <summary class="cursor-pointer list-none px-5 py-4 font-semibold text-slate-900 marker:content-none flex items-center justify-between gap-4">
          {_render_model_provider_heading(model["slug"], model_meta)}
          <span class="text-xs font-normal text-slate-400 group-open:hidden">展开</span>
        </summary>
        <div class="border-t border-slate-100 px-5 pb-5 pt-4">
          {_render_provider_top3_summaries(
              _build_provider_top3_summary(model),
              _build_provider_top3_price_insight(
                  model,
                  week_start=week_start,
                  prior_model=prior_models.get(model["slug"]),
              ),
          )}
          {_render_chart_card(
              f"{model['slug']} · Provider 承接趋势",
              f"按日展示 Top {TOP_PROVIDER_COUNT} provider 的模型承接用量（B），仅统计「已展示」。",
              _provider_insight(model),
              chart_id,
          )}
          {_json_script(chart_id + "-data", chart)}
          <h4 class="mt-6 text-sm font-semibold text-slate-800">原始数据</h4>
          {_render_provider_daily_tabs(model)}
        </div>
      </details>
            """,
        )

    source_paths = "<br />".join(
        f'<span class="break-all font-mono text-xs text-slate-500">{_escape(p)}</span>'
        for p in (usage["source_path"], provider["source_path"], income["source_path"])
    )

    return f"""<!DOCTYPE html>
<html lang="zh-CN">
<head>
  <meta charset="utf-8" />
  <meta name="viewport" content="width=device-width, initial-scale=1" />
  <title>Core Models Dashboard {week_label}</title>
  <script src="https://cdn.tailwindcss.com"></script>
  <script>
    tailwind.config = {{
      theme: {{
        extend: {{
          fontFamily: {{
            sans: ['"Inter"', 'ui-sans-serif', 'system-ui', '"PingFang SC"', 'sans-serif'],
          }},
        }},
      }},
    }};
  </script>
  <style id="editorial-visual-tuning">
    /* Editorial theme: hierarchy + layout tuning (weekly report / research note) */
    :root {{
      --radius-xl: 16px;
      --radius-lg: 12px;
      --shadow-1: 0 1px 2px rgba(15, 23, 42, 0.06), 0 8px 24px rgba(15, 23, 42, 0.06);
      --shadow-2: 0 2px 6px rgba(15, 23, 42, 0.10), 0 18px 48px rgba(15, 23, 42, 0.10);

      --line: rgba(148, 163, 184, 0.45);
      --panel: #ffffff;

      --accent: #9a3412; /* editorial brown */
      --accent2: #14532d; /* deep green */
      --insight-bg: rgba(154, 52, 18, 0.08);
      --insight-ink: rgb(69 26 3);
    }}

    html {{ text-rendering: optimizeLegibility; }}
    body {{ letter-spacing: -0.01em; }}
    @supports (text-wrap: pretty) {{
      h1, h2, h3, h4, p {{ text-wrap: pretty; }}
    }}

    html:not(.dark) body {{
      background:
        radial-gradient(1200px 600px at 10% 0%, rgba(154, 52, 18, 0.05), transparent),
        radial-gradient(1200px 600px at 90% 0%, rgba(20, 83, 45, 0.045), transparent),
        rgb(248 250 252);
    }}

    html.dark body {{
      background:
        radial-gradient(1200px 600px at 10% 0%, rgba(154, 52, 18, 0.08), transparent),
        radial-gradient(1200px 600px at 90% 0%, rgba(20, 83, 45, 0.06), transparent),
        rgb(2 6 23);
    }}

    section[id] {{ scroll-margin-top: 110px; }}
    header + section {{ margin-top: 2.25rem; }}

    section[id] > div > h2 {{
      position: relative;
      padding-bottom: 0.65rem;
    }}
    section[id] > div > h2::after {{
      content: "";
      position: absolute;
      left: 0;
      bottom: 0;
      width: 136px;
      height: 2px;
      border-radius: 999px;
      background: linear-gradient(90deg, var(--accent), rgba(20, 83, 45, 0.35), transparent);
      opacity: 0.95;
    }}

    .rounded-xl.border.border-slate-200.bg-white.shadow-sm,
    header.rounded-2xl.border.border-slate-200.bg-white.shadow-sm {{
      border-color: var(--line) !important;
      border-radius: var(--radius-xl) !important;
      box-shadow: var(--shadow-1) !important;
      background: var(--panel) !important;
    }}
    .rounded-xl.border.border-slate-200.bg-white.shadow-sm:hover {{
      box-shadow: var(--shadow-2) !important;
      transform: translateY(-1px);
      transition: box-shadow 180ms ease, transform 180ms ease;
    }}

    .rounded-md.bg-indigo-50 {{
      background: var(--insight-bg) !important;
      color: var(--insight-ink) !important;
      border-left: 3px solid rgba(154, 52, 18, 0.55);
    }}

    table {{ border-collapse: separate; border-spacing: 0; }}
    thead th {{
      position: sticky;
      top: 62px;
      z-index: 1;
      background: rgba(248, 250, 252, 0.92);
      backdrop-filter: blur(10px);
    }}
    tbody tr:nth-child(2n) {{ background: rgba(248, 250, 252, 0.55); }}
    tbody tr:hover {{ background: rgba(254, 243, 199, 0.30); }}

    nav[aria-label="章节导航"] {{
      border-color: rgba(148, 163, 184, 0.35) !important;
      box-shadow: 0 1px 2px rgba(15, 23, 42, 0.06), 0 10px 30px rgba(15, 23, 42, 0.06) !important;
    }}
  </style>
  <script src="{_escape(chart_script_src)}"></script>
</head>
<body class="min-h-screen bg-slate-50 text-slate-900 antialiased">
  <div class="mx-auto max-w-7xl px-4 py-8 sm:px-6 lg:px-8">
    <header class="rounded-2xl border border-slate-200 bg-white px-6 py-8 shadow-sm sm:px-8">
      <p class="text-sm font-medium text-indigo-600">OpenRouter · Core Models</p>
      <h1 class="mt-2 text-2xl font-semibold tracking-tight text-slate-900 sm:text-3xl">核心模型周报 · {_escape(week_label)}</h1>
      <dl class="mt-6 grid gap-4 text-sm sm:grid-cols-2">
        <div>
          <dt class="font-medium text-slate-500">时间范围</dt>
          <dd class="mt-1 text-slate-900">{_escape(meta_range)}</dd>
        </div>
        <div>
          <dt class="font-medium text-slate-500">数据更新时间</dt>
          <dd class="mt-1 text-slate-900">{_escape(data_updated)} · 看板生成 {generated_at.strftime("%Y-%m-%d %H:%M")}</dd>
        </div>
        <div class="sm:col-span-2">
          <dt class="font-medium text-slate-500">数据来源</dt>
          <dd class="mt-1 text-slate-700">{_escape(data_source_line)}</dd>
          <dd class="mt-2 space-y-1">{source_paths}</dd>
        </div>
      </dl>
      <p class="mt-6 rounded-lg border border-indigo-100 bg-indigo-50 px-4 py-3 text-sm leading-relaxed text-indigo-950">
        <span class="font-semibold">核心结论：</span>{_escape(headline)}
      </p>
    </header>

    <section class="mt-8" aria-label="关键指标">
      <h2 class="sr-only">关键指标</h2>
      <div class="grid gap-4 sm:grid-cols-2 lg:grid-cols-3">
        {kpi_cards}
      </div>
    </section>

    <nav class="sticky top-0 z-10 mt-8 flex flex-wrap gap-2 rounded-xl border border-slate-200 bg-white/95 px-4 py-3 shadow-sm backdrop-blur" aria-label="章节导航">
      <a href="#usage" class="rounded-lg px-3 py-1.5 text-sm font-medium text-slate-600 hover:bg-slate-100 hover:text-indigo-600">用量</a>
      <a href="#provider" class="rounded-lg px-3 py-1.5 text-sm font-medium text-slate-600 hover:bg-slate-100 hover:text-indigo-600">Provider</a>
      <a href="#income" class="rounded-lg px-3 py-1.5 text-sm font-medium text-slate-600 hover:bg-slate-100 hover:text-indigo-600">收入</a>
    </nav>

    <section id="usage" class="mt-10 scroll-mt-24">
      <div class="mb-6">
        <h2 class="text-xl font-semibold text-slate-900">Core Model Usage</h2>
        <p class="mt-1 text-sm text-slate-500">模型 Activity 全平台 token 用量（T），对比周内结构与模型份额。</p>
      </div>
      <div class="grid gap-6 lg:grid-cols-2">
        {_render_chart_card(
            "每日用量趋势",
            "按自然日对比各核心模型的用量（T），用于发现周内峰值与尾部模型。",
            _usage_daily_insight(usage),
            "usage-daily-chart",
        )}
        {_render_chart_card(
            "周合计用量占比",
            "各模型占核心模型总用量的比例，衡量路由集中度。",
            _usage_share_insight(kpis),
            "usage-share-chart",
        )}
      </div>
      {_json_script("usage-daily-data", usage_charts["daily"])}
      {_json_script("usage-share-data", usage_charts["share"])}
      <h3 class="mt-8 text-sm font-semibold text-slate-800">原始数据</h3>
      {_render_table(usage["headers"], usage["rows"])}
    </section>

    <section id="provider" class="mt-14 scroll-mt-24">
      <div class="mb-6">
        <h2 class="text-xl font-semibold text-slate-900">Core Model Provider</h2>
        <p class="mt-1 text-sm text-slate-500">分模型查看 provider 承接用量（B）；下表含全部行，高亮异常状态。</p>
      </div>
      <div class="space-y-6">
        {''.join(provider_sections)}
      </div>
    </section>

    <section id="income" class="mt-14 scroll-mt-24 pb-12">
      <div class="mb-6">
        <h2 class="text-xl font-semibold text-slate-900">Core Model Income</h2>
        <p class="mt-1 text-sm text-slate-500">每日 paid_usd 与周合计，评估核心模型商业化贡献。</p>
      </div>
      <div class="grid gap-6 lg:grid-cols-2">
        {_render_chart_card(
            "每日收入趋势",
            "各模型按日的 paid_usd，观察变现节奏是否与用量同步。",
            _income_daily_insight(income),
            "income-daily-chart",
        )}
        {_render_chart_card(
            "周合计收入占比",
            "核心模型之间的收入份额，可与用量占比交叉验证变现效率。",
            _income_share_insight(kpis),
            "income-share-chart",
        )}
      </div>
      {_json_script("income-daily-data", income_charts["daily"])}
      {_json_script("income-share-data", income_charts["share"])}
      <h3 class="mt-8 text-sm font-semibold text-slate-800">原始数据</h3>
      {_render_table(income["headers"], income["rows"])}
    </section>
  </div>

  <script>
    const CHART_COLORS = ["#4f46e5", "#6366f1", "#818cf8", "#a5b4fc", "#cbd5e1", "#94a3b8"];

    function readJson(id) {{
      return JSON.parse(document.getElementById(id).textContent);
    }}

    function colorAt(i) {{
      return CHART_COLORS[i % CHART_COLORS.length];
    }}

    function datasetBarColors(ds, i) {{
      const base = ds.color || colorAt(i);
      return {{ fill: base + "cc", stroke: base }};
    }}

    function groupedBarChart(canvasId, payload, yTitle) {{
      const ctx = document.getElementById(canvasId);
      return new Chart(ctx, {{
        type: "bar",
        data: {{
          labels: payload.labels,
          datasets: payload.datasets.map((ds, i) => {{
            const colors = datasetBarColors(ds, i);
            return {{
            label: ds.label,
            data: ds.data,
            backgroundColor: colors.fill,
            borderColor: colors.stroke,
            borderWidth: 1,
            borderRadius: 4,
          }};
          }}),
        }},
        options: {{
          responsive: true,
          maintainAspectRatio: false,
          plugins: {{
            legend: {{ position: "bottom", labels: {{ boxWidth: 12, padding: 16 }} }},
          }},
          scales: {{
            x: {{
              grid: {{ display: false }},
              ticks: {{ color: "#64748b", font: {{ size: 11 }} }},
            }},
            y: {{
              beginAtZero: true,
              title: {{ display: true, text: yTitle, color: "#475569" }},
              grid: {{ color: "#f1f5f9" }},
              ticks: {{ color: "#64748b" }},
            }},
          }},
        }},
      }});
    }}

    const SHARE_PERCENT_MIN = 5;

    function shareSliceTotal(dataset) {{
      return (dataset.data || []).reduce((sum, v) => sum + (Number(v) || 0), 0);
    }}

    function sharePercentLabel(value, total) {{
      if (!total) return "0.0%";
      return ((Number(value) || 0) / total * 100).toFixed(1) + "%";
    }}

    const sharePercentLabelsPlugin = {{
      id: "sharePercentLabels",
      afterDatasetsDraw(chart) {{
        const meta = chart.getDatasetMeta(0);
        if (!meta || !meta.data.length) return;
        const dataset = chart.data.datasets[0];
        const total = shareSliceTotal(dataset);
        const ctx = chart.ctx;
        ctx.save();
        ctx.textAlign = "center";
        ctx.textBaseline = "middle";
        ctx.fillStyle = "#1e293b";
        ctx.font = "600 11px Inter, system-ui, sans-serif";
        meta.data.forEach((arc, i) => {{
          const value = dataset.data[i];
          const pct = total ? (Number(value) || 0) / total * 100 : 0;
          if (pct < SHARE_PERCENT_MIN) return;
          const pos = arc.tooltipPosition();
          ctx.fillText(pct.toFixed(1) + "%", pos.x, pos.y);
        }});
        ctx.restore();
      }},
    }};

    function doughnutShareChart(canvasId, payload, options) {{
      const opts = options || {{}};
      const showPercent = !!opts.showPercent;
      const valueSuffix = opts.unit === "USD" ? "" : "T";
      const ctx = document.getElementById(canvasId);
      const plugins = showPercent ? [sharePercentLabelsPlugin] : [];
      return new Chart(ctx, {{
        type: "doughnut",
        plugins,
        data: {{
          labels: payload.labels,
          datasets: [{{
            data: payload.values,
            backgroundColor: payload.labels.map((_, i) => colorAt(i) + "dd"),
            borderColor: "#ffffff",
            borderWidth: 2,
          }}],
        }},
        options: {{
          responsive: true,
          maintainAspectRatio: false,
          cutout: "58%",
          plugins: {{
            legend: {{
              position: "bottom",
              labels: {{
                boxWidth: 12,
                padding: 14,
                generateLabels(chart) {{
                  const dataset = chart.data.datasets[0];
                  const total = shareSliceTotal(dataset);
                  return chart.data.labels.map((label, i) => {{
                    const value = dataset.data[i];
                    const pctText = showPercent ? "  " + sharePercentLabel(value, total) : "";
                    return {{
                      text: label + pctText,
                      fillStyle: dataset.backgroundColor[i],
                      strokeStyle: dataset.borderColor,
                      lineWidth: dataset.borderWidth,
                      hidden: false,
                      index: i,
                    }};
                  }});
                }},
              }},
            }},
            tooltip: {{
              callbacks: {{
                label(context) {{
                  const label = context.label || "";
                  const value = context.parsed;
                  const dataset = context.chart.data.datasets[0];
                  const total = shareSliceTotal(dataset);
                  const pct = sharePercentLabel(value, total);
                  const unit = opts.unit === "USD" ? "$" + Number(value).toLocaleString() : value + valueSuffix;
                  return showPercent ? label + ": " + unit + " (" + pct + ")" : label + ": " + unit;
                }},
              }},
            }},
          }},
        }},
      }});
    }}

    document.addEventListener("DOMContentLoaded", () => {{
      document.querySelectorAll(".provider-tab").forEach((button) => {{
        button.addEventListener("click", () => {{
          const tabId = button.dataset.providerTab;
          const group = button.closest("details");
          if (!tabId || !group) return;
          group.querySelectorAll(".provider-tab").forEach((tabButton) => {{
            const active = tabButton === button;
            tabButton.setAttribute("aria-pressed", active ? "true" : "false");
            tabButton.classList.toggle("bg-indigo-600", active);
            tabButton.classList.toggle("text-white", active);
            tabButton.classList.toggle("shadow-sm", active);
            tabButton.classList.toggle("bg-white", !active);
            tabButton.classList.toggle("text-slate-600", !active);
            tabButton.classList.toggle("hover:bg-slate-50", !active);
          }});
          group.querySelectorAll(".provider-tab-panel").forEach((panel) => {{
            const active = panel.dataset.providerPanel === tabId;
            panel.classList.toggle("hidden", !active);
            panel.hidden = !active;
          }});
        }});
      }});

      const canvases = document.querySelectorAll("canvas");
      if (typeof Chart === "undefined") {{
        canvases.forEach((el) => {{
          el.parentElement.innerHTML =
            '<p class="rounded-lg bg-amber-50 px-4 py-3 text-sm text-amber-900">Chart.js 未加载。请确认同目录存在 chart.umd.min.js。</p>';
        }});
        return;
      }}
      groupedBarChart("usage-daily-chart", readJson("usage-daily-data"), "用量 (T)");
      doughnutShareChart("usage-share-chart", readJson("usage-share-data"), {{ showPercent: true }});
      groupedBarChart("income-daily-chart", readJson("income-daily-data"), "收入 (USD)");
      doughnutShareChart("income-share-chart", readJson("income-share-data"), {{ showPercent: true, unit: "USD" }});

      document.querySelectorAll("script[id$='-data']").forEach((node) => {{
        if (!node.id.startsWith("provider-chart-")) return;
        const canvasId = node.id.replace(/-data$/, "");
        if (!document.getElementById(canvasId)) return;
        const payload = JSON.parse(node.textContent);
        groupedBarChart(canvasId, payload, "承接量 (B)");
      }});
    }});
  </script>
</body>
</html>
"""


def generate_core_models_dashboard(
    week_start: date,
    *,
    output_dir: Path | None = None,
    usage_path: Path | None = None,
    provider_path: Path | None = None,
    income_path: Path | None = None,
) -> Path:
    config.ensure_dirs()
    resolved_usage = usage_path or usage_workbook_path(week_start, output_dir=output_dir)
    resolved_provider = provider_path or provider_workbook_path(week_start, output_dir=output_dir)
    resolved_income = income_path or income_workbook_path(week_start, output_dir=output_dir)

    usage = load_usage_workbook(resolved_usage)
    provider = load_provider_workbook(resolved_provider)
    if resolved_income.exists():
        income = load_income_workbook(resolved_income)
    else:
        # Income is manually supplied per week and may not exist yet (e.g. a brand
        # new ISO week, or intraday refresh). Render the dashboard without income
        # rather than failing the whole build.
        logger.warning("Income workbook missing (%s); rendering dashboard without income section", resolved_income)
        income = {
            "source_path": str(resolved_income),
            "headers": [],
            "rows": [],
            "dict_rows": [],
            "days": [],
            "metadata": {},
        }
    model_meta_by_slug = build_model_meta_by_slug(load_monitored_models())
    provider_color_map = build_provider_color_map(provider)
    prior_provider_models_by_slug = load_prior_week_provider_models_by_slug(
        week_start,
        output_dir=output_dir,
    )

    week_label = iso_week_label(week_start)
    days = [d.isoformat() for d in week_dates(week_start)]
    for block, name in ((usage, "Usage"), (income, "Income")):
        if block["days"] != days:
            block["days"] = days

    out_path = dashboard_output_path(week_start, output_dir=output_dir)
    chart_src = install_chart_vendor(out_path.parent)
    html_text = render_dashboard_html(
        week_label,
        usage,
        provider,
        income,
        week_start=week_start,
        chart_script_src=chart_src,
        model_meta_by_slug=model_meta_by_slug,
        provider_color_map=provider_color_map,
        prior_provider_models_by_slug=prior_provider_models_by_slug,
    )
    out_path.write_text(html_text, encoding="utf-8")
    return out_path
