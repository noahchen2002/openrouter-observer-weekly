"""Build a human-readable daily summary from the Core Model Provider workbook.

The daily Feishu push needs a readable digest of *yesterday's* data (the
"原始数据" view ld asked for): for each core model, which providers carried how
much, with SiliconFlow highlighted, plus a data-quality footnote distinguishing
the three kinds of blank uptake (truly none / not scraped / intentionally
skipped by the 90% rule).

This module only *reads* existing output workbooks; it never scrapes or writes.
"""

from __future__ import annotations

from dataclasses import dataclass, field
from datetime import date
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from pipeline.core_models_provider import (
    DATE_COLUMN,
    output_path_for_week as provider_output_path_for_week,
)
from pipeline.data_availability import EXPIRED_UNAVAILABLE_TEXT
from pipeline.logger import get_logger
from pipeline.model_provider_price_uptime import load_monitored_models
from pipeline.utils import normalize_for_match, parse_compact_number

logger = get_logger()

# 展示状态 vocabulary (kept in sync with model_provider_usage.py).
STATUS_SHOWN = "已展示"
STATUS_MISSING = "未展示"            # provider tooltip had no row for this model
STATUS_SKIPPED = "用量少未查询"       # not scraped on purpose (90% coverage rule)
EXPIRED_TEXT = EXPIRED_UNAVAILABLE_TEXT  # OpenRouter no longer exposes that past day

# How the SiliconFlow provider is identified across name/slug spellings.
SILICONFLOW_KEYS = {normalize_for_match("SiliconFlow"), normalize_for_match("siliconflow")}

TOP_N_PROVIDERS = 5


@dataclass
class ProviderUptake:
    provider: str
    uptake_text: str
    uptake_tokens: float | None
    share: float | None  # 承接占比, 0..1
    status: str


@dataclass
class ModelDaily:
    model_slug: str
    display_name: str
    rows: list[ProviderUptake] = field(default_factory=list)

    @property
    def shown(self) -> list[ProviderUptake]:
        return [r for r in self.rows if r.status == STATUS_SHOWN]

    @property
    def n_missing(self) -> int:
        return sum(1 for r in self.rows if r.status == STATUS_MISSING)

    @property
    def n_skipped(self) -> int:
        return sum(1 for r in self.rows if r.status == STATUS_SKIPPED)


def _display_name_by_slug() -> dict[str, str]:
    mapping: dict[str, str] = {}
    for model in load_monitored_models():
        # MonitoredModel has model_slug; display name falls back to slug.
        mapping[model.model_slug] = getattr(model, "model_name", None) or model.model_slug
    return mapping


def _col_index(headers: list[Any], name: str) -> int | None:
    for idx, header in enumerate(headers):
        if str(header or "").strip() == name:
            return idx
    return None


def _read_model_daily(ws, snapshot_date: date, display_name: str, slug: str) -> ModelDaily:
    rows_iter = ws.iter_rows(values_only=True)
    try:
        headers = list(next(rows_iter))
    except StopIteration:
        return ModelDaily(model_slug=slug, display_name=display_name)

    date_col = _col_index(headers, DATE_COLUMN)
    provider_col = _col_index(headers, "Provider")
    uptake_col = _col_index(headers, "Provider 承接用量")
    share_col = _col_index(headers, "承接占比")
    status_col = _col_index(headers, "展示状态")
    if provider_col is None:
        return ModelDaily(model_slug=slug, display_name=display_name)

    target = snapshot_date.isoformat()
    model = ModelDaily(model_slug=slug, display_name=display_name)
    for values in rows_iter:
        row = list(values)

        def cell(idx: int | None) -> Any:
            if idx is None or idx >= len(row):
                return None
            return row[idx]

        # 日期 cell may be a date or an iso string; match on iso prefix.
        day_val = cell(date_col)
        day_text = day_val.isoformat() if isinstance(day_val, date) else str(day_val or "").strip()[:10]
        if date_col is not None and day_text != target:
            continue

        provider = str(cell(provider_col) or "").strip()
        if not provider:
            continue
        uptake_text = str(cell(uptake_col) or "").strip()
        share_val = cell(share_col)
        share = float(share_val) if isinstance(share_val, (int, float)) else None
        status = str(cell(status_col) or "").strip()
        model.rows.append(
            ProviderUptake(
                provider=provider,
                uptake_text=uptake_text,
                uptake_tokens=parse_compact_number(uptake_text),
                share=share,
                status=status,
            )
        )
    return model


def collect_daily_models(week_start: date, snapshot_date: date, *, provider_path: Path | None = None) -> list[ModelDaily]:
    path = provider_path or provider_output_path_for_week(week_start)
    if not path.exists():
        raise FileNotFoundError(f"Core Model Provider workbook not found: {path}")

    names = _display_name_by_slug()
    wb = load_workbook(path, data_only=True, read_only=True)
    try:
        models: list[ModelDaily] = []
        for sheet_name in wb.sheetnames:
            if sheet_name == "Metadata":
                continue
            display = names.get(sheet_name, sheet_name)
            models.append(_read_model_daily(wb[sheet_name], snapshot_date, display, sheet_name))
        return models
    finally:
        wb.close()


def _fmt_share(share: float | None) -> str:
    return f"{share * 100:.1f}%" if isinstance(share, (int, float)) else "—"


def _siliconflow_row(model: ModelDaily) -> ProviderUptake | None:
    for r in model.rows:
        if normalize_for_match(r.provider) in SILICONFLOW_KEYS:
            return r
    return None


def build_daily_summary_text(
    week_start: date,
    snapshot_date: date,
    *,
    provider_path: Path | None = None,
    top_n: int = TOP_N_PROVIDERS,
) -> str:
    """Return a Feishu-friendly plain-text digest of one day's provider uptake."""
    models = collect_daily_models(week_start, snapshot_date, provider_path=provider_path)

    lines: list[str] = [f"📊 OpenRouter 核心模型日报 · {snapshot_date.isoformat()}"]
    any_data = False

    for model in models:
        shown = sorted(
            model.shown,
            key=lambda r: (r.uptake_tokens if r.uptake_tokens is not None else -1),
            reverse=True,
        )
        sf = _siliconflow_row(model)
        if not shown and sf is None:
            continue
        any_data = True

        lines.append("")
        lines.append(f"▍{model.display_name}")

        for rank, r in enumerate(shown[:top_n], start=1):
            uptake = r.uptake_text or "—"
            lines.append(f"  {rank}. {r.provider}：{uptake}（{_fmt_share(r.share)}）")

        # Always surface SiliconFlow's standing, even if outside Top N or not shown.
        if sf is not None:
            if sf.status == STATUS_SHOWN:
                in_top = any(normalize_for_match(r.provider) in SILICONFLOW_KEYS for r in shown[:top_n])
                if not in_top:
                    lines.append(f"  · SiliconFlow：{sf.uptake_text or '—'}（{_fmt_share(sf.share)}）")
            elif sf.status == STATUS_SKIPPED:
                lines.append("  · SiliconFlow：用量少未单独抓取（核心 provider 已覆盖≥90%）")
            elif sf.status == STATUS_MISSING:
                lines.append("  · SiliconFlow：该模型当日未在其页面展示用量")

        notes = []
        if model.n_missing:
            notes.append(f"{model.n_missing} 个未展示")
        if model.n_skipped:
            notes.append(f"{model.n_skipped} 个用量少未查询")
        if notes:
            lines.append(f"  （另有 {'、'.join(notes)}）")

    if not any_data:
        lines.append("")
        lines.append("⚠️ 未读到当日任何 provider 承接数据，请检查爬取是否完成。")

    lines.append("")
    lines.append("完整看板见附件 HTML / 内网看板。承接量为爬取值，最终以看板为准。")
    return "\n".join(lines)
