"""Competitor-focused view of the Core Models provider data.

Reframes the daily card around competitors: for each core model, list every
provider ranked ABOVE SiliconFlow (by uptake), with their call volume, prices
(input / output / cache) and uptime — so we watch who is beating us and on what
terms. SiliconFlow's own row is shown as the reference line.

Read-only over the Core Model Provider workbook.
"""

from __future__ import annotations

from dataclasses import dataclass, field
from datetime import date, timedelta
from pathlib import Path

from openpyxl import load_workbook

from pipeline.core_models_provider import (
    DATE_COLUMN,
    output_path_for_week as provider_output_path_for_week,
)
from pipeline.daily_summary import SILICONFLOW_KEYS, STATUS_SHOWN, _display_name_by_slug
from pipeline.logger import get_logger
from pipeline.openrouter_status import load_status_for_day
from pipeline.utils import normalize_for_match, parse_compact_number

logger = get_logger()


@dataclass
class ProviderRow:
    provider: str
    uptake_text: str
    uptake_tokens: float
    input_price: float | None
    output_price: float | None
    cache_price: float | None
    uptime: float | None
    is_sf: bool = False


@dataclass
class ModelCompetition:
    model_slug: str
    display_name: str
    rows: list[ProviderRow] = field(default_factory=list)  # all 已展示, sorted desc by uptake
    sf_rank: int | None = None
    shown_count: int = 0

    @property
    def sf_row(self) -> ProviderRow | None:
        for r in self.rows:
            if r.is_sf:
                return r
        return None

    @property
    def competitors_above(self) -> list[ProviderRow]:
        """Providers ranked strictly above SiliconFlow."""
        if self.sf_rank is None:
            return self.rows
        return self.rows[: self.sf_rank - 1]

    @property
    def sf_captured(self) -> bool:
        """True when SiliconFlow serves the model AND its daily uptake was captured.

        A 未展示 SF row (endpoint exists — has price/uptime — but the day's uptake
        wasn't scraped) has uptake_tokens<0: SF 仍承接，只是当日用量未知。
        """
        sf = self.sf_row
        return sf is not None and sf.uptake_tokens >= 0


def iso_week_start(target_date: date) -> date:
    return target_date - timedelta(days=target_date.weekday())


def _col(headers: list, name: str) -> int | None:
    for i, h in enumerate(headers):
        if str(h or "").strip() == name:
            return i
    return None


def _num(v) -> float | None:
    return float(v) if isinstance(v, (int, float)) else None


def collect_competition(week_start: date, day: date, *, provider_path: Path | None = None) -> list[ModelCompetition]:
    path = provider_path or provider_output_path_for_week(week_start)
    if not path.exists():
        raise FileNotFoundError(f"Core Model Provider workbook not found: {path}")
    names = _display_name_by_slug()
    target = day.isoformat()
    out: list[ModelCompetition] = []

    wb = load_workbook(path, data_only=True, read_only=True)
    try:
        for sheet in wb.sheetnames:
            if sheet == "Metadata":
                continue
            ws = wb[sheet]
            it = ws.iter_rows(values_only=True)
            try:
                headers = list(next(it))
            except StopIteration:
                continue
            di = _col(headers, DATE_COLUMN)
            pi = _col(headers, "Provider")
            ui = _col(headers, "Provider 承接用量")
            ti = _col(headers, "展示状态")
            ipi = _col(headers, "Input Price")
            opi = _col(headers, "Output Price")
            ci = _col(headers, "Cache Read")
            upi = _col(headers, "Uptime")
            if pi is None:
                continue

            rows: list[ProviderRow] = []
            for r in it:
                def c(idx):
                    return r[idx] if idx is not None and idx < len(r) else None
                day_val = c(di)
                day_text = day_val.isoformat() if isinstance(day_val, date) else str(day_val or "").strip()[:10]
                if di is not None and day_text != target:
                    continue
                prov = str(c(pi) or "").strip()
                if not prov:
                    continue
                is_sf = normalize_for_match(prov) in SILICONFLOW_KEYS
                # SiliconFlow 是监控主体：只要它在 provider 列表里有行（=有 endpoint，
                # 带价格/uptime），即便当日用量未抓到（展示状态=未展示）也保留，否则会被
                # 误判成"SF未承接"。其它 provider 仍只取"已展示"。
                if not is_sf and str(c(ti) or "").strip() != STATUS_SHOWN:
                    continue
                tok = parse_compact_number(str(c(ui) or "").strip())
                rows.append(
                    ProviderRow(
                        provider=prov,
                        uptake_text=str(c(ui) or "").strip(),
                        uptake_tokens=tok if tok is not None else -1.0,
                        input_price=_num(c(ipi)),
                        output_price=_num(c(opi)),
                        cache_price=_num(c(ci)),
                        uptime=_num(c(upi)),
                        is_sf=is_sf,
                    )
                )
            rows.sort(key=lambda x: x.uptake_tokens, reverse=True)
            mc = ModelCompetition(model_slug=sheet, display_name=names.get(sheet, sheet), rows=rows, shown_count=len(rows))
            for i, rw in enumerate(rows, start=1):
                if rw.is_sf:
                    mc.sf_rank = i
                    break
            out.append(mc)
        return out
    finally:
        wb.close()


def _fmt_price(p: float | None) -> str:
    if p is None:
        return "-"
    return f"${p:g}"


def _fmt_uptime(u: float | None) -> str:
    return f"{u:.1f}%" if u is not None else "-"


def _status_by_slug(day: date) -> dict[str, dict[str, int]]:
    """{model slug -> {"429":n,"502":n,"504":n}} keyed by the last path segment of
    the model_id, to match ModelCompetition.model_slug. {} when snapshot absent."""
    return {mid.split("/")[-1]: codes for mid, codes in load_status_for_day(day).items()}


def _prior_uptake_lookup(day: date) -> dict[tuple[str, str], float]:
    """(model_slug, provider) -> 前天调用量 tokens, for day-over-day 环比.

    Loads from 前天's own ISO-week workbook (handles the Monday cross-week case);
    returns {} gracefully if that workbook doesn't exist yet (no delta shown).
    """
    prior_day = day - timedelta(days=1)
    try:
        prior_models = collect_competition(iso_week_start(prior_day), prior_day)
    except FileNotFoundError:
        logger.info("Prior-day provider workbook missing for %s; 环比 will be blank", prior_day.isoformat())
        return {}
    out: dict[tuple[str, str], float] = {}
    for m in prior_models:
        for r in m.rows:
            out[(m.model_slug, r.provider)] = r.uptake_tokens
    return out


def _fmt_delta(cur: float, prior: float | None) -> str:
    """Day-over-day vs 前天: ▲/▼ with %, 新 for no prior, '' when n/a."""
    if cur is None or cur < 0:
        return ""
    if prior is None or prior <= 0:
        return "新" if (cur or 0) > 0 else ""
    pct = (cur - prior) / prior * 100
    if abs(pct) < 0.5:
        return "≈0%"
    return f"{'▲' if pct > 0 else '▼'}{abs(pct):.0f}%"


def _dw(s: str) -> int:
    """Display width: CJK chars take 2 cells, others 1 (for monospace alignment)."""
    return sum(2 if ord(c) > 0x2E80 else 1 for c in s)


def _pad(s: str, width: int, align: str = "left") -> str:
    gap = max(0, width - _dw(s))
    return (s + " " * gap) if align == "left" else (" " * gap + s)


def _render_table(rows_data: list[list[str]], aligns: list[str]) -> str:
    """Render a fixed-width monospace table (rendered inside a ``` code block)."""
    ncol = len(rows_data[0])
    widths = [max(_dw(r[c]) for r in rows_data) for c in range(ncol)]
    out = []
    for ri, row in enumerate(rows_data):
        cells = [_pad(row[c], widths[c], aligns[c]) for c in range(ncol)]
        out.append("  ".join(cells))
        if ri == 0:  # separator under header
            out.append("  ".join("-" * widths[c] for c in range(ncol)))
    return "\n".join(out)


# Short provider names so the table fits on mobile.
PROVIDER_ABBR = {
    "deepseek": "Dp", "siliconflow": "SF", "novitaai": "Novita", "novita": "Novita",
    "gmicloud": "GMI", "alibabacloudint": "Ali", "alibaba": "Ali", "atlascloud": "Atlas",
    "baiduqianfan": "Baidu", "baidu": "Baidu", "streamlake": "Stream", "moonshotai": "Moon",
    "cloudflare": "CF", "deepinfra": "DpInfra", "parasail": "Parasail", "friendli": "Friendli",
    "chutes": "Chutes", "digitalocean": "DO", "venice": "Venice", "akashml": "Akash",
    "inceptron": "Incep", "wandb": "WandB", "morph": "Morph",
}
# Color dots cycled per model to visually separate groups.
MODEL_DOTS = ["🔵", "🟣", "🟠", "🟡", "🟤", "🔴", "⚪️", "🟢"]


def _abbr(provider: str) -> str:
    return PROVIDER_ABBR.get(normalize_for_match(provider), provider)


def _abbr_model(name: str) -> str:
    """Shorten model display name for the narrow 模型 column (DeepSeek->Dp etc.)."""
    return (
        name.replace("DeepSeek", "Dp").replace("deepseek", "Dp")
        .replace("MoonshotAI", "Kimi").replace("Z.ai", "GLM").replace("Z.AI", "GLM")
        .replace(": ", ":").strip()
    )


def _vol_cell(row: "ProviderRow", prior: dict, model_slug: str) -> str:
    """调用量 + 环比前天（塞进同一格）：如「40.4B ▲16%」。"""
    vol = row.uptake_text or "-"
    delta = _fmt_delta(row.uptake_tokens, prior.get((model_slug, row.provider)))
    return f"{vol} {delta}" if delta else vol


def _fmt_429_cell(codes: dict | None) -> str:
    """当日429 格：无数据→—，0→0，>0→⚠️粗体千分位。"""
    if not codes or codes.get("429") is None:
        return "—"
    n = int(codes.get("429") or 0)
    return f"⚠️{n:,}" if n > 0 else "0"


def _fmt_5xx_cell(codes: dict | None) -> str:
    """502/504 格：显示「502数/504数」，任一>0 加⚠️；无数据→—。"""
    if not codes or (codes.get("502") is None and codes.get("504") is None):
        return "—"
    a, b = int(codes.get("502") or 0), int(codes.get("504") or 0)
    s = f"{a:,}/{b:,}"
    return f"⚠️{s}" if (a or b) else s


def _merged_table(models: list["ModelCompetition"], prior: dict | None = None, status: dict | None = None) -> dict:
    """One native table for all models (Feishu caps tables-per-card). Each model's
    competitors-above-SF plus the SF row, grouped, with a colored 模型 column. The
    调用量 cell carries a day-over-day 环比 (▲/▼ vs 前天); the 当日429 and 502/504
    columns are SiliconFlow's error counts for the model, shown on the ✅SF row only."""
    prior = prior or {}
    status = status or {}
    rows = []
    dot_i = 0
    # 硅基尚未承接的模型（如新上线模型）只展示头部这么多承接厂商，避免卡片过长。
    NO_SF_TOP_N = 5
    for m in models:
        dot = MODEL_DOTS[dot_i % len(MODEL_DOTS)]
        dot_i += 1
        label = f"{dot}{_abbr_model(m.display_name)}"
        sf = m.sf_row
        codes = status.get(m.model_slug)
        # 模型名每行都填（不留空），否则手机上分组的空行看不到归属。
        if sf is None:
            # 硅基未承接：仍展示该模型，列出头部承接厂商，末行标「❌SF未承接」，
            # 这样新上线 / 硅基暂未承接的模型也能持续监控竞争格局。
            for r in m.rows[:NO_SF_TOP_N]:
                rows.append({
                    "model": label,
                    "prov": _abbr(r.provider),
                    "vol": _vol_cell(r, prior, m.model_slug),
                    "inp": _fmt_price(r.input_price),
                    "out": _fmt_price(r.output_price),
                    "up": _fmt_uptime(r.uptime),
                    "e429": "",
                    "e5xx": "",
                })
            rows.append({
                "model": label,
                "prov": "❌SF未承接",
                "vol": "—",
                "inp": "-",
                "out": "-",
                "up": "-",
                "e429": _fmt_429_cell(codes),
                "e5xx": _fmt_5xx_cell(codes),
            })
            continue
        for r in m.competitors_above[:NO_SF_TOP_N]:
            rows.append({
                "model": label,
                "prov": _abbr(r.provider),
                "vol": _vol_cell(r, prior, m.model_slug),
                "inp": _fmt_price(r.input_price),
                "out": _fmt_price(r.output_price),
                "up": _fmt_uptime(r.uptime),
                "e429": "",  # 错误码是硅基自己的，只在 SF 行显示
                "e5xx": "",
            })
        rows.append({
            "model": label,
            "prov": "✅SF",
            # 有 endpoint 但当日用量未抓到 → 明确标注，避免被读成"0"或"未承接"。
            "vol": _vol_cell(sf, prior, m.model_slug) if sf.uptake_tokens >= 0 else "用量未抓到",
            "inp": _fmt_price(sf.input_price),
            "out": _fmt_price(sf.output_price),
            "up": _fmt_uptime(sf.uptime),
            "e429": _fmt_429_cell(codes),
            "e5xx": _fmt_5xx_cell(codes),
        })
    return {
        "tag": "table",
        "page_size": len(rows) or 1,
        "row_height": "low",
        "header_style": {"background_style": "grey", "bold": True},
        "columns": [
            {"name": "model", "display_name": "模型", "data_type": "text"},
            {"name": "prov", "display_name": "厂商", "data_type": "text"},
            {"name": "vol", "display_name": "调用量(环比)", "data_type": "text"},
            {"name": "inp", "display_name": "In", "data_type": "text"},
            {"name": "out", "display_name": "Out", "data_type": "text"},
            {"name": "up", "display_name": "Uptime", "data_type": "text"},
            {"name": "e429", "display_name": "当日429", "data_type": "text"},
            {"name": "e5xx", "display_name": "502/504", "data_type": "text"},
        ],
        "rows": rows,
    }


def build_competitor_card(
    day: date, dashboard_url: str | None, *, as_of: str, provisional: bool, ai_text: str | None = None
) -> dict:
    models = collect_competition(iso_week_start(day), day)
    prior = _prior_uptake_lookup(day)
    prior_day = day - timedelta(days=1)

    if provisional:
        title = "🎯 竞品监控 · OpenRouter 日报（实时）"
        template = "orange"
        note = f"⚠️ 截至 {as_of} · 今日（{day.strftime('%m-%d')}）数据**未结算**，仅供参考 · 环比对比前天（{prior_day.strftime('%m-%d')}）"
    else:
        # 推送日 = 数据日+1（每天 9 点推送当天，标题随之每天更新）。
        push_day = day + timedelta(days=1)
        title = "🎯 竞品监控 · OpenRouter 日报"
        template = "blue"
        note = (
            f"📅 {push_day.month}月{push_day.day:02d}日推送 · 数据截至 {day.strftime('%m-%d')} · "
            f"环比对比前天（{prior_day.strftime('%m-%d')}）· 关注排在硅基流动前面的厂商"
        )

    lead = sum(1 for m in models if m.sf_captured and m.sf_rank == 1)
    trail = sum(1 for m in models if m.sf_captured and m.sf_rank and m.sf_rank > 1)
    uncap = sum(1 for m in models if m.sf_row is not None and not m.sf_captured)
    no_sf = sum(1 for m in models if m.sf_row is None)
    overall = (
        f"**硅基流动**：{len(models)} 个核心模型中领跑 {lead} 个，落后 {trail} 个"
        + (f"，用量未抓 {uncap} 个" if uncap else "")
        + (f"，未承接 {no_sf} 个" if no_sf else "")
        + "（下列为各模型领先我们的竞品）"
    )

    elements: list[dict] = [
        {"tag": "markdown", "content": note},
        {"tag": "markdown", "content": overall},
        {"tag": "hr"},
    ]

    # Per-model rank one-liners (dots match the table groups), then table (with 错误码).
    status = _status_by_slug(day)
    rank_lines = []
    di = 0
    for m in models:
        dot = MODEL_DOTS[di % len(MODEL_DOTS)]
        di += 1
        if m.sf_row is None:
            base = f"{dot} **{m.display_name}**：硅基未承接（共 {m.shown_count} 家承接，详见下表）"
        elif not m.sf_captured:
            base = f"{dot} **{m.display_name}**：硅基承接，但当日用量未抓到（价格/Uptime 见下表）"
        elif m.sf_rank == 1:
            base = f"{dot} **{m.display_name}**：🥇 硅基第 1/{m.shown_count}（领跑）"
        else:
            base = f"{dot} **{m.display_name}**：硅基第 {m.sf_rank}/{m.shown_count}，前面 {len(m.competitors_above)} 家"
        rank_lines.append(base)
    elements.append({"tag": "markdown", "content": "\n".join(rank_lines)})
    elements.append({"tag": "hr"})
    elements.append({"tag": "markdown", "content": "**各模型领先硅基的竞品明细**（✅SF=硅基流动，色点区分模型；调用量格内 ▲▼=环比前天；当日429/502/504=硅基错误码数，仅 SF 行）"})
    elements.append(_merged_table(models, prior, status))
    elements.append({"tag": "hr"})

    if ai_text:
        elements.append({"tag": "markdown", "content": "🤖 **AI 竞品洞察**（DeepSeek V4 Pro）\n" + ai_text})
        elements.append({"tag": "hr"})

    if dashboard_url:
        # Card v2: button is a top-level element (no more "action" wrapper).
        elements.append({
            "tag": "button",
            "text": {"tag": "plain_text", "content": "📊 打开完整看板（公网）"},
            "type": "primary",
            "width": "default",
            "behaviors": [{"type": "open_url", "default_url": dashboard_url}],
        })
    elements.append({
        "tag": "markdown",
        "content": "<font color='grey'>价格 $/M tokens · ✅SF=硅基流动 · Dp=DeepSeek · 调用量格内 ▲▼%=环比前天 · 数值为爬取值，最终以看板为准。</font>",
    })

    # Card v2 schema (required for the native table element).
    return {
        "schema": "2.0",
        "config": {"wide_screen_mode": True},
        "header": {"template": template, "title": {"tag": "plain_text", "content": title}},
        "body": {"elements": elements},
    }
