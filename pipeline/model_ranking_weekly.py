"""Generate the weekly AI Model Rankings Excel workbook."""

from __future__ import annotations

import re
from datetime import date, timedelta
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from playwright.sync_api import Page, sync_playwright

from pipeline import config
from pipeline.logger import get_logger
from pipeline.utils import (
    extract_chart_tooltip_structured,
    navigate_with_retry,
    parse_compact_number,
)
from pipeline.week_format import format_cn_date, format_data_range, iso_week_label
from pipeline.workbook_metadata import (
    build_standard_metadata,
    resolve_updated_at,
    style_metadata_sheet,
)

RANKINGS_URL = config.RANKINGS_URL

logger = get_logger()

RANKING_HEADERS = ["模型名称", "原始用量", "换算后用量（T）", "占比（%）"]
METADATA_SOURCE_URL = "https://openrouter.ai/rankings"
OUTPUT_SUBDIR = "Ranking"


def get_previous_complete_week_start(today: date | None = None) -> date:
    """Return the Monday for the most recent fully completed ISO week."""
    if today is None:
        today = date.today()
    current_monday = today - timedelta(days=today.weekday())
    return current_monday - timedelta(days=7)


def parse_week_start(value: str | None) -> date:
    if not value:
        return get_previous_complete_week_start()
    parsed = date.fromisoformat(value)
    if parsed.weekday() != 0:
        raise ValueError(f"--week must be a Monday ISO date, got {value}")
    return parsed


def _clean_usage_text(value: str | None) -> str:
    if not value:
        return ""
    return value.replace("tokens", "").replace("Tokens", "").strip()


def _tokens_to_t(tokens: Any) -> float | None:
    if tokens is None:
        return None
    return round(float(tokens) / 1_000_000_000_000, 4)


def _pct(part: Any, total: Any) -> float | None:
    if part is None or not total:
        return None
    return round(float(part) / float(total), 4)


def build_ranking_rows(payload: dict[str, Any]) -> list[dict[str, Any]]:
    """Convert a Top Models tooltip payload into the required four-column rows."""
    model_items = payload.get("models") or []
    total_tokens = payload.get("total_tokens")
    if total_tokens is None:
        total_tokens = sum(float(item.get("tokens") or 0) for item in model_items)
        logger.warning("Tooltip payload did not expose Total; using model row sum as Total")

    rows: list[dict[str, Any]] = []
    regular_items = [
        item for item in model_items
        if str(item.get("model_name", "")).strip().lower() != "others"
    ]
    others_items = [
        item for item in model_items
        if str(item.get("model_name", "")).strip().lower() == "others"
    ]

    for item in regular_items + others_items:
        tokens = item.get("tokens")
        rows.append({
            "模型名称": item.get("model_name", ""),
            "原始用量": _clean_usage_text(item.get("tokens_text")),
            "换算后用量（T）": _tokens_to_t(tokens),
            "占比（%）": _pct(tokens, total_tokens),
        })

    total_text = payload.get("total_tokens_text")
    if not total_text:
        total_text = _format_tokens_compact(float(total_tokens or 0))
    rows.append({
        "模型名称": "Total",
        "原始用量": _clean_usage_text(total_text),
        "换算后用量（T）": _tokens_to_t(total_tokens),
        "占比（%）": 1,
    })
    return rows


def _format_tokens_compact(tokens: float) -> str:
    units = [
        ("T", 1_000_000_000_000),
        ("B", 1_000_000_000),
        ("M", 1_000_000),
        ("K", 1_000),
    ]
    for suffix, multiplier in units:
        if abs(tokens) >= multiplier:
            value = tokens / multiplier
            text = f"{value:.2f}".rstrip("0").rstrip(".")
            return f"{text}{suffix}"
    return str(int(tokens))


def build_metadata(week_start: date, updated_at: date | None = None) -> list[tuple[str, str]]:
    resolved = resolve_updated_at(updated_at, snapshot_dates=[week_start])
    return build_standard_metadata(
        week_start,
        updated_at=resolved,
        data_source=METADATA_SOURCE_URL,
    )


def output_path_for_week(week_start: date, output_dir: Path | None = None) -> Path:
    base_dir = output_dir or (config.OUTPUT_DIR / OUTPUT_SUBDIR)
    return base_dir / f"AI Model Rankings {iso_week_label(week_start)}.xlsx"


def save_rankings_workbook(
    rows: list[dict[str, Any]],
    week_start: date,
    output_dir: Path | None = None,
    updated_at: date | None = None,
) -> Path:
    out_path = output_path_for_week(week_start, output_dir=output_dir)
    out_path.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    ws = wb.active
    ws.title = "Ranking"
    ws.append(RANKING_HEADERS)
    for row in rows:
        ws.append([row.get(header) for header in RANKING_HEADERS])

    meta_ws = wb.create_sheet("Metadata")
    meta_ws.append(["字段", "内容"])
    for key, value in build_metadata(week_start, updated_at=updated_at):
        meta_ws.append([key, value])

    _style_ranking_sheet(ws)
    style_metadata_sheet(meta_ws, column_b_width=50)
    wb.save(out_path)
    logger.info("Saved AI Model Rankings Excel to %s", out_path)
    return out_path


def _style_ranking_sheet(ws) -> None:
    header_fill = PatternFill("solid", fgColor="D9EAF7")
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    widths = [24, 16, 18, 14]
    for index, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(index)].width = width

    for row in ws.iter_rows(min_row=2, min_col=3, max_col=4):
        for cell in row:
            if isinstance(cell.value, (int, float)):
                cell.number_format = "0.0000"
    ws.freeze_panes = "A2"


def _locate_top_models_section(page: Page):
    candidates = [
        page.locator("section").filter(has_text=re.compile(r"Top Models|Weekly usage of models", re.IGNORECASE)),
        page.locator("article").filter(has_text=re.compile(r"Top Models|Weekly usage of models", re.IGNORECASE)),
        page.locator("div").filter(has_text=re.compile(r"Top Models|Weekly usage of models", re.IGNORECASE)),
    ]
    for locator in candidates:
        try:
            if locator.count() > 0:
                return locator.first
        except Exception:
            continue
    return None


def _scrape_top_models_chart_payloads(page: Page) -> list[dict[str, Any]]:
    section = _locate_top_models_section(page)
    if section is None:
        logger.warning("Top Models chart section not found")
        return []

    try:
        section.scroll_into_view_if_needed(timeout=5_000)
    except Exception:
        pass
    page.wait_for_timeout(1_000)

    chart = section.locator(".recharts-wrapper").first
    bars = chart.locator(".recharts-bar-rectangles .recharts-rectangle")

    # One JS round trip for all bar center-x (per-bar bounding_box() with a 2s
    # timeout over hundreds of stacked rects cost ~10 minutes here).
    try:
        x_positions = bars.evaluate_all(
            """
            (els) => {
              const seen = new Set();
              const out = [];
              for (const el of els) {
                const b = el.getBoundingClientRect();
                if (!b.width && !b.height) continue;
                const cx = Math.round(b.x + b.width / 2);
                if (!seen.has(cx)) { seen.add(cx); out.push(cx); }
              }
              return out.sort((a, b) => a - b);
            }
            """
        )
    except Exception as exc:
        logger.warning("Top Models bar enumeration failed: %s", exc)
        x_positions = []
    if not x_positions:
        logger.warning("Top Models chart has no bars")
        return []

    # Only the recent dates matter (we just need the target week's Total). Hovering
    # all ~52 history bars is wasteful, so keep the right-most dozen.
    x_positions = x_positions[-12:]

    svg = chart.locator("svg").first
    try:
        svg_box = svg.bounding_box(timeout=2_000)
    except Exception:
        svg_box = None
    if not svg_box:
        logger.warning("Cannot get Top Models chart SVG bounding box")
        return []

    center_y = svg_box["y"] + svg_box["height"] / 2
    payloads: list[dict[str, Any]] = []
    seen_dates: set[str] = set()
    for x in x_positions:
        page.mouse.move(x, center_y)
        page.wait_for_timeout(200)
        payload = extract_chart_tooltip_structured(page)
        if not payload:
            continue
        chart_date = payload.get("chart_date")
        if not chart_date or chart_date in seen_dates:
            continue
        seen_dates.add(chart_date)
        payloads.append(payload)

    logger.info("Collected %d Top Models chart payloads", len(payloads))
    return payloads


def scrape_top_models_payload_for_week(week_start: date) -> dict[str, Any]:
    """Scrape the Top Models chart tooltip payload for a specific data week."""
    target_date = week_start.isoformat()
    config.ensure_dirs()

    with sync_playwright() as playwright:
        browser = playwright.chromium.launch(headless=config.HEADLESS)
        page = browser.new_page()
        page.set_viewport_size({"width": 1600, "height": 1000})
        try:
            if not navigate_with_retry(page, RANKINGS_URL):
                raise RuntimeError(f"Could not open {RANKINGS_URL}")
            payloads = _scrape_top_models_chart_payloads(page)
            payload = next((item for item in payloads if item.get("chart_date") == target_date), None)
            if not payload:
                available = [item.get("chart_date") for item in payloads]
                raise RuntimeError(
                    f"Top Models chart did not contain {target_date}; available dates: {available[-10:]}"
                )
            return payload
        finally:
            browser.close()


def generate_ai_model_rankings_excel(
    week_start: date,
    output_dir: Path | None = None,
    updated_at: date | None = None,
) -> Path:
    payload = scrape_top_models_payload_for_week(week_start)
    rows = build_ranking_rows(payload)
    return save_rankings_workbook(rows, week_start, output_dir=output_dir, updated_at=updated_at)


def build_rows_from_text_pairs(items: list[tuple[str, str]]) -> list[dict[str, Any]]:
    """Test/helper path for converting model usage text into workbook rows."""
    models = [
        {
            "model_name": name,
            "tokens_text": usage,
            "tokens": parse_compact_number(usage),
        }
        for name, usage in items
    ]
    payload = {
        "models": models,
        "total_tokens": sum(float(item["tokens"] or 0) for item in models),
        "total_tokens_text": None,
    }
    return build_ranking_rows(payload)
