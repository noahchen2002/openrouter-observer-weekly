"""Shared Metadata sheet helpers for weekly Excel workbooks."""

from __future__ import annotations

import re
from datetime import date
from typing import Any, Iterable

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.worksheet.worksheet import Worksheet

from pipeline.week_format import format_cn_date, format_data_range, iso_week_label

METADATA_SHEET = "Metadata"
_ISO_DATE_TAB = re.compile(r"^\d{4}-\d{2}-\d{2}$")


def resolve_updated_at(
    explicit: date | None,
    *,
    snapshot_dates: Iterable[date] | None = None,
) -> date:
    if explicit is not None:
        return explicit
    dates = list(snapshot_dates or [])
    if dates:
        return max(dates)
    return date.today()


def build_standard_metadata(
    week_start: date,
    *,
    updated_at: date,
    data_source: str,
    extra_rows: Iterable[tuple[str, str]] | None = None,
) -> list[tuple[str, str]]:
    rows: list[tuple[str, str]] = [
        ("数据范围", format_data_range(week_start)),
        ("数据更新时间", format_cn_date(updated_at)),
        ("数据来源", data_source),
        ("数据周", iso_week_label(week_start)),
    ]
    if extra_rows:
        rows.extend(extra_rows)
    return rows


def list_snapshot_tab_dates(wb: Workbook) -> list[date]:
    dates: list[date] = []
    for name in wb.sheetnames:
        if name == METADATA_SHEET:
            continue
        if not _ISO_DATE_TAB.match(name):
            continue
        dates.append(date.fromisoformat(name))
    return dates


def upsert_metadata_sheet(wb: Workbook, rows: list[tuple[str, str]]) -> None:
    if METADATA_SHEET in wb.sheetnames:
        wb.remove(wb[METADATA_SHEET])
    meta_ws = wb.create_sheet(METADATA_SHEET)
    meta_ws.append(["字段", "内容"])
    for key, value in rows:
        meta_ws.append([key, value])
    style_metadata_sheet(meta_ws)


def read_metadata_map(ws: Worksheet) -> dict[str, str]:
    result: dict[str, str] = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or not row[0]:
            continue
        result[str(row[0])] = str(row[1] or "")
    return result


def style_metadata_sheet(ws: Worksheet, *, column_b_width: int = 56) -> None:
    header_fill = PatternFill("solid", fgColor="D9EAD3")
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = header_fill
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = column_b_width
