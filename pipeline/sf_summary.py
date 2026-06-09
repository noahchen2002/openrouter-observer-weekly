"""SiliconFlow-focused summary of the Core Models provider data.

Produces, per core model, SiliconFlow's standing and its change vs the previous
day with data: ranking (up/down), call volume (承接用量) and share, and service
stability (uptime / latency). Read-only over the Core Model Provider workbooks.
"""

from __future__ import annotations

from dataclasses import dataclass
from datetime import date, timedelta
from pathlib import Path

from openpyxl import load_workbook

from pipeline.core_models_provider import (
    DATE_COLUMN,
    output_path_for_week as provider_output_path_for_week,
)
from pipeline.daily_summary import (
    SILICONFLOW_KEYS,
    STATUS_SHOWN,
    _display_name_by_slug,
)
from pipeline.logger import get_logger
from pipeline.utils import normalize_for_match, parse_compact_number

logger = get_logger()


@dataclass
class SFMetric:
    model_slug: str
    display_name: str
    status: str
    rank: int | None          # SiliconFlow's rank among 已展示 providers by uptake (1 = top)
    shown_count: int
    uptake_text: str
    uptake_tokens: float | None
    share: float | None       # 0..1
    uptime: float | None      # percent
    latency: str | None


def iso_week_start(target_date: date) -> date:
    return target_date - timedelta(days=target_date.weekday())


def _col(headers: list, name: str) -> int | None:
    for i, h in enumerate(headers):
        if str(h or "").strip() == name:
            return i
    return None


def _is_sf(provider: str) -> bool:
    return normalize_for_match(provider) in SILICONFLOW_KEYS


def collect_sf_metrics(week_start: date, day: date, *, provider_path: Path | None = None) -> dict[str, SFMetric]:
    path = provider_path or provider_output_path_for_week(week_start)
    if not path.exists():
        return {}
    names = _display_name_by_slug()
    out: dict[str, SFMetric] = {}
    wb = load_workbook(path, data_only=True, read_only=True)
    try:
        target = day.isoformat()
        for sheet in wb.sheetnames:
            if sheet == "Metadata":
                continue
            ws = wb[sheet]
            it = ws.iter_rows(values_only=True)
            try:
                headers = list(next(it))
            except StopIteration:
                continue
            di = _col(headers, DATE_COLUMN)
            pi = _col(headers, "Provider")
            ui = _col(headers, "Provider 承接用量")
            si = _col(headers, "承接占比")
            ti = _col(headers, "展示状态")
            upi = _col(headers, "Uptime")
            li = _col(headers, "Latency")
            if pi is None:
                continue

            shown: list[tuple[str, float]] = []  # (provider, uptake_tokens) for ranking
            sf_row = None
            for row in it:
                def cell(idx):
                    return row[idx] if idx is not None and idx < len(row) else None
                day_val = cell(di)
                day_text = day_val.isoformat() if isinstance(day_val, date) else str(day_val or "").strip()[:10]
                if di is not None and day_text != target:
                    continue
                provider = str(cell(pi) or "").strip()
                if not provider:
                    continue
                status = str(cell(ti) or "").strip()
                tokens = parse_compact_number(str(cell(ui) or "").strip())
                if status == STATUS_SHOWN:
                    shown.append((provider, tokens if tokens is not None else -1.0))
                if _is_sf(provider):
                    share_val = cell(si)
                    uptime_val = cell(upi)
                    sf_row = SFMetric(
                        model_slug=sheet,
                        display_name=names.get(sheet, sheet),
                        status=status,
                        rank=None,
                        shown_count=0,
                        uptake_text=str(cell(ui) or "").strip(),
                        uptake_tokens=tokens,
                        share=float(share_val) if isinstance(share_val, (int, float)) else None,
                        uptime=float(uptime_val) if isinstance(uptime_val, (int, float)) else None,
                        latency=str(cell(li) or "").strip() or None,
                    )
            if sf_row is None:
                continue
            shown.sort(key=lambda x: x[1], reverse=True)
            sf_row.shown_count = len(shown)
            for i, (prov, _) in enumerate(shown, start=1):
                if _is_sf(prov):
                    sf_row.rank = i
                    break
            out[sheet] = sf_row
    finally:
        wb.close()
    return out


def find_previous_day_metrics(day: date, *, max_back: int = 10) -> tuple[date | None, dict[str, SFMetric]]:
    """Walk back day-by-day until a provider workbook with SF data is found."""
    for delta in range(1, max_back + 1):
        cand = day - timedelta(days=delta)
        metrics = collect_sf_metrics(iso_week_start(cand), cand)
        if metrics:
            return cand, metrics
    return None, {}


def _fmt_pct(x: float | None) -> str:
    return f"{x * 100:.1f}%" if isinstance(x, (int, float)) else "—"


def _rank_arrow(cur: int | None, prev: int | None) -> str:
    if cur is None or prev is None:
        return ""
    if cur < prev:
        return f" ↑{prev - cur}"
    if cur > prev:
        return f" ↓{cur - prev}"
    return " 持平"


def _vol_delta(cur: float | None, prev: float | None) -> str:
    if not cur or not prev:
        return ""
    pct = (cur - prev) / prev * 100
    sign = "+" if pct >= 0 else ""
    return f"（较前值 {sign}{pct:.0f}%）"


def _uptime_delta(cur: float | None, prev: float | None) -> str:
    if cur is None or prev is None:
        return ""
    d = cur - prev
    sign = "+" if d >= 0 else ""
    return f"（{sign}{d:.1f}pp）"


def build_sf_card(day: date, dashboard_url: str | None, *, as_of: str, provisional: bool) -> dict:
    cur = collect_sf_metrics(iso_week_start(day), day)
    prev_day, prev = find_previous_day_metrics(day)

    if provisional:
        title = f"🟢 硅基流动 · OpenRouter 承接概览（{day.isoformat()} 实时）"
        template = "orange"
        note = f"⚠️ 截至 {as_of} · 今日数据**未结算**，仅供参考"
    else:
        title = f"🟢 硅基流动 · OpenRouter 承接概览（{day.isoformat()}）"
        template = "blue"
        note = f"数据日期 {day.isoformat()}"
    if prev_day:
        note += f"｜对比基线：{prev_day.isoformat()}"

    # Overall line: how many core models SF is in top-3, total uptake.
    in_top3 = sum(1 for m in cur.values() if m.rank and m.rank <= 3)
    total_tokens = sum(m.uptake_tokens for m in cur.values() if m.uptake_tokens)
    def _fmt_b(t):
        return f"{t/1e9:.0f}B" if t >= 1e9 else (f"{t/1e6:.0f}M" if t else "—")
    overall = f"**硅基流动**在 {len(cur)} 个核心模型中：Top3 {in_top3} 个 · 合计承接约 {_fmt_b(total_tokens)}"

    elements: list[dict] = [
        {"tag": "markdown", "content": note},
        {"tag": "markdown", "content": overall},
        {"tag": "hr"},
    ]

    if not cur:
        elements.append({"tag": "markdown", "content": "⚠️ 未读到当日硅基流动数据，请检查爬取是否完成。"})
    for slug, m in cur.items():
        p = prev.get(slug)
        if m.status == STATUS_SHOWN and m.rank:
            rank_txt = f"第 {m.rank}/{m.shown_count} 名{_rank_arrow(m.rank, p.rank if p else None)}"
            vol_txt = f"调用量 `{m.uptake_text or '—'}`（占 {_fmt_pct(m.share)}）{_vol_delta(m.uptake_tokens, p.uptake_tokens if p else None)}"
            stab_bits = []
            if m.uptime is not None:
                stab_bits.append(f"uptime {m.uptime:.1f}%{_uptime_delta(m.uptime, p.uptime if p else None)}")
            if m.latency:
                stab_bits.append(f"延迟 {m.latency}")
            stab_txt = "稳定性 " + " · ".join(stab_bits) if stab_bits else ""
            lines = [f"**{m.display_name}** — 排名 {rank_txt}", vol_txt]
            if stab_txt:
                lines.append(stab_txt)
            elements.append({"tag": "markdown", "content": "\n".join(lines)})
        else:
            reason = {"用量少未查询": "用量少未单独抓取", "未展示": "当日未在页面展示"}.get(m.status, m.status or "无数据")
            elements.append({"tag": "markdown", "content": f"**{m.display_name}** — 硅基：{reason}"})

    if dashboard_url:
        elements.append({"tag": "hr"})
        elements.append(
            {
                "tag": "action",
                "actions": [
                    {
                        "tag": "button",
                        "text": {"tag": "plain_text", "content": "📊 打开完整看板（公网）"},
                        "url": dashboard_url,
                        "type": "primary",
                    }
                ],
            }
        )
    elements.append({"tag": "note", "elements": [{"tag": "plain_text", "content": "承接量/稳定性为爬取值，最终以看板为准。↑↓ 为较对比基线的排名变化。"}]})

    return {
        "config": {"wide_screen_mode": True},
        "header": {"template": template, "title": {"tag": "plain_text", "content": title}},
        "elements": elements,
    }
