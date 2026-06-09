"""Generate weekly Core Model Income Excel from model_income exports."""

from __future__ import annotations

from datetime import date, datetime, timedelta
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook

from pipeline import config
from pipeline.core_models_usage import (
    FIXED_HEADERS,
    OUTPUT_SUBDIR,
    _pct,
    _style_metadata_sheet,
    _style_usage_sheet,
)
from pipeline.logger import get_logger
from pipeline.model_ranking_weekly import format_cn_date, format_data_range
from pipeline.workbook_metadata import resolve_updated_at
from pipeline.model_provider_price_uptime import (
    MonitoredModel,
    iso_week_label,
    load_monitored_models,
)
from pipeline.model_provider_usage import week_dates

logger = get_logger()

INCOME_SHEET_NAME = "Income"
INCOME_WORKBOOK_BASENAME = "Core Model Income"
INCOME_SHEET_TITLE = "查询结果"
INCOME_COLUMNS = ("emitted_day", "model_name", "paid_usd")
INCOME_MODEL_NAME_BY_ID: dict[str, str] = {
    "deepseek/deepseek-v4-flash": "deepseek-ai/DeepSeek-V4-Flash",
    "deepseek/deepseek-v4-pro": "deepseek-ai/DeepSeek-V4-Pro",
    "deepseek/deepseek-v3.2": "deepseek-ai/DeepSeek-V3.2",
    "moonshotai/kimi-k2.6": "moonshotai/Kimi-K2.6",
    "z-ai/glm-5.1": "zai-org/GLM-5.1",
}
SUMMARY_HEADERS = ["周合计", "占比（%）"]


def default_income_path(week_start: date, input_dir: Path | None = None) -> Path:
    base = input_dir or (config.DATA_DIR / "input")
    iso = week_start.isocalendar()
    labeled = base / f"model_income_{iso_week_label(week_start)}.xlsx"
    short = base / f"model_income_W{iso.week:02d}.xlsx"
    if short.exists():
        return short
    return labeled


def output_path_for_week(week_start: date, output_dir: Path | None = None) -> Path:
    base_dir = output_dir or (config.OUTPUT_DIR / OUTPUT_SUBDIR)
    return base_dir / f"{INCOME_WORKBOOK_BASENAME} {iso_week_label(week_start)}.xlsx"


def income_model_name(model_id: str) -> str | None:
    return INCOME_MODEL_NAME_BY_ID.get(model_id.strip())


def income_headers(days: list[date]) -> list[str]:
    return FIXED_HEADERS + [day.isoformat() for day in days] + SUMMARY_HEADERS


def _parse_emitted_day(value: Any) -> date | None:
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    text = str(value).strip()
    if not text:
        return None
    return date.fromisoformat(text[:10])


def load_income_daily(
    income_path: Path,
    week_start: date,
    *,
    sheet_name: str = INCOME_SHEET_TITLE,
) -> tuple[dict[tuple[str, str], float], float]:
    """Return ((iso_day, model_name) -> paid_usd, platform_week_total_usd)."""
    if not income_path.exists():
        raise FileNotFoundError(f"Income workbook not found: {income_path}")

    wb = load_workbook(income_path, data_only=True, read_only=True)
    if sheet_name not in wb.sheetnames:
        raise ValueError(f"Income workbook {income_path} has no sheet {sheet_name!r}")

    ws = wb[sheet_name]
    headers = [ws.cell(1, col).value for col in range(1, ws.max_column + 1)]
    col_index = {str(header).strip(): idx + 1 for idx, header in enumerate(headers) if header}
    missing = [name for name in INCOME_COLUMNS if name not in col_index]
    if missing:
        raise ValueError(f"Income sheet missing columns: {missing}")

    days = {day.isoformat() for day in week_dates(week_start)}
    daily_index: dict[tuple[str, str], float] = {}
    platform_week_total = 0.0

    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row:
            continue
        day_value = _parse_emitted_day(row[col_index["emitted_day"] - 1])
        model_name = row[col_index["model_name"] - 1]
        paid = row[col_index["paid_usd"] - 1]
        if day_value is None or model_name is None or paid is None:
            continue
        day_key = day_value.isoformat()
        if day_key not in days:
            continue
        name_key = str(model_name).strip()
        amount = float(paid)
        daily_index[(day_key, name_key)] = daily_index.get((day_key, name_key), 0.0) + amount
        platform_week_total += amount

    wb.close()
    return daily_index, platform_week_total


def build_income_row(
    monitored_model: MonitoredModel,
    income_name: str,
    days: list[date],
    daily_index: dict[tuple[str, str], float],
    platform_week_total: float,
) -> dict[str, Any]:
    daily_values = [
        daily_index.get((day.isoformat(), income_name), 0.0)
        for day in days
    ]
    week_sum = sum(daily_values)
    row: dict[str, Any] = {
        "模型ID": monitored_model.model_id,
        "模型名称": income_name,
    }
    for day, amount in zip(days, daily_values):
        row[day.isoformat()] = round(amount, 2) if amount else 0.0
    row["周合计"] = round(week_sum, 2) if week_sum else 0.0
    row["占比（%）"] = _pct(week_sum, platform_week_total) if platform_week_total else 0.0
    return row


def build_income_rows(
    models: list[MonitoredModel],
    days: list[date],
    daily_index: dict[tuple[str, str], float],
    platform_week_total: float,
) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    for monitored_model in models:
        income_name = income_model_name(monitored_model.model_id)
        if not income_name:
            logger.warning("No income model_name mapping for %s", monitored_model.model_id)
            income_name = monitored_model.model_slug
        week_sum = sum(
            daily_index.get((day.isoformat(), income_name), 0.0)
            for day in days
        )
        if week_sum == 0:
            logger.warning(
                "No income rows in target week for %s (income name %s)",
                monitored_model.model_id,
                income_name,
            )
        rows.append(
            build_income_row(monitored_model, income_name, days, daily_index, platform_week_total)
        )
    return rows


def build_income_metadata(
    week_start: date,
    income_path: Path,
    platform_week_total: float,
    updated_at: date | None = None,
    *,
    snapshot_dates: list[date] | None = None,
) -> list[tuple[str, str]]:
    resolved = resolve_updated_at(
        updated_at,
        snapshot_dates=snapshot_dates or [week_start + timedelta(days=7)],
    )
    return [
        ("数据范围", format_data_range(week_start)),
        ("数据更新时间", format_cn_date(resolved)),
        ("数据来源", str(income_path)),
        ("数据周", iso_week_label(week_start)),
        (
            "收入占比分母",
            f"${platform_week_total:,.2f}（该周收入文件中全部模型 paid_usd 合计）",
        ),
        ("占比分母说明", "占比（%）= 模型周收入合计 / 收入文件该周全平台合计"),
    ]


def _style_income_sheet(ws, column_count: int, day_start_col: int, day_end_col: int) -> None:
    _style_usage_sheet(ws, column_count)
    for row in ws.iter_rows(min_row=2, min_col=day_start_col, max_col=day_end_col):
        for cell in row:
            if isinstance(cell.value, (int, float)):
                cell.number_format = "$#,##0.00"
    week_col = column_count - 1
    pct_col = column_count
    for col_cells in ws.iter_cols(min_col=week_col, max_col=week_col, min_row=2):
        for cell in col_cells:
            if isinstance(cell.value, (int, float)):
                cell.number_format = "$#,##0.00"
    for col_cells in ws.iter_cols(min_col=pct_col, max_col=pct_col, min_row=2):
        for cell in col_cells:
            if isinstance(cell.value, (int, float)):
                cell.number_format = "0.00%"


def save_core_models_income_workbook(
    rows: list[dict[str, Any]],
    week_start: date,
    days: list[date],
    income_path: Path,
    platform_week_total: float,
    output_dir: Path | None = None,
    updated_at: date | None = None,
) -> Path:
    out_path = output_path_for_week(week_start, output_dir=output_dir)
    out_path.parent.mkdir(parents=True, exist_ok=True)
    headers = income_headers(days)

    wb = Workbook()
    ws = wb.active
    ws.title = INCOME_SHEET_NAME
    ws.append(headers)
    for row in rows:
        ws.append([row.get(header) for header in headers])

    meta_ws = wb.create_sheet("Metadata")
    meta_ws.append(["字段", "内容"])
    for key, value in build_income_metadata(
        week_start,
        income_path,
        platform_week_total,
        updated_at=updated_at,
        snapshot_dates=days,
    ):
        meta_ws.append([key, value])

    day_start_col = 3
    day_end_col = 2 + len(days)
    _style_income_sheet(ws, len(headers), day_start_col, day_end_col)
    _style_metadata_sheet(meta_ws)
    wb.save(out_path)
    logger.info(
        "Saved Core Model Income Excel to %s (platform week total $%s)",
        out_path,
        f"{platform_week_total:,.2f}",
    )
    return out_path


def generate_core_models_income_excel(
    week_start: date,
    income_path: Path | None = None,
    models: list[MonitoredModel] | None = None,
    output_dir: Path | None = None,
    updated_at: date | None = None,
) -> Path:
    resolved_income = income_path or default_income_path(week_start)
    monitored_models = models or load_monitored_models()
    daily_index, platform_week_total = load_income_daily(resolved_income, week_start)
    days = week_dates(week_start)
    rows = build_income_rows(monitored_models, days, daily_index, platform_week_total)
    return save_core_models_income_workbook(
        rows,
        week_start,
        days,
        resolved_income,
        platform_week_total,
        output_dir=output_dir,
        updated_at=updated_at,
    )


def fill_core_models_income(
    week_start: date,
    income_path: Path | None = None,
    output_dir: Path | None = None,
    models: list[MonitoredModel] | None = None,
) -> Path:
    config.ensure_dirs()
    return generate_core_models_income_excel(
        week_start,
        income_path=income_path,
        models=models,
        output_dir=output_dir,
    )
