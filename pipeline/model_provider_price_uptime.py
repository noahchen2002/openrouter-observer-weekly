"""Generate weekly model provider price and uptime monitoring workbooks."""

from __future__ import annotations

import json
import re
from dataclasses import dataclass
from datetime import date
from pathlib import Path
from typing import Any

import httpx
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from playwright.sync_api import Page, sync_playwright

from pipeline import config
from pipeline.logger import get_logger
from pipeline.utils import clean_lines, navigate_with_retry, normalize_for_match, safe_parse_float
from pipeline.workbook_metadata import (
    build_standard_metadata,
    list_snapshot_tab_dates,
    resolve_updated_at,
    upsert_metadata_sheet,
)

logger = get_logger()

CONFIG_PATH = config.PROJECT_ROOT / "config" / "core_models.json"
OUTPUT_SUBDIR = "Price&Uptime&Usage"
HEADERS = [
    "Provider",
    "Region",
    "Quantization",
    "Latency",
    "Throughput",
    "Uptime",
    "Total Context",
    "Max Output",
    "Input Price",
    "Output Price",
    "Cache Read",
]


@dataclass(frozen=True)
class MonitoredModel:
    model_id: str
    model_slug: str
    model_url: str


def parse_model_slug(model_id: str | None = None, model_url: str | None = None) -> str:
    if model_id and "/" in model_id:
        return model_id.rsplit("/", 1)[-1].strip()
    if model_url:
        clean = model_url.rstrip("/").split("?", 1)[0]
        return clean.rsplit("/", 1)[-1].strip()
    raise ValueError("model_id or model_url is required to derive model_slug")


def model_url_from_id(model_id: str) -> str:
    clean_model_id = model_id.strip().strip("/")
    if "/" not in clean_model_id:
        raise ValueError(f"model_id must use author/model format, got {model_id}")
    return f"https://openrouter.ai/{clean_model_id}"


def load_monitored_models(path: Path = CONFIG_PATH) -> list[MonitoredModel]:
    payload = json.loads(path.read_text(encoding="utf-8"))
    if not isinstance(payload, list):
        raise ValueError(f"{path} must contain a JSON list")

    models: list[MonitoredModel] = []
    for item in payload:
        if isinstance(item, str):
            model_id = item.strip()
            model_url = model_url_from_id(model_id)
            model_slug = parse_model_slug(model_id=model_id)
        elif isinstance(item, dict):
            model_id = str(item.get("model_id") or "").strip()
            if not model_id:
                raise ValueError(f"model_id is required for every monitored model: {item}")
            model_url = str(item.get("model_url") or "").strip() or model_url_from_id(model_id)
            model_slug = str(item.get("model_slug") or "").strip() or parse_model_slug(model_id=model_id)
        else:
            continue
        models.append(MonitoredModel(model_id=model_id, model_slug=model_slug, model_url=model_url))
    return models


def parse_week_start(value: str) -> date:
    parsed = date.fromisoformat(value)
    if parsed.weekday() != 0:
        raise ValueError(f"--week must be a Monday ISO date, got {value}")
    return parsed


def parse_snapshot_date(value: str | None = None) -> date:
    return date.fromisoformat(value) if value else date.today()


def iso_week_label(week_start: date) -> str:
    iso = week_start.isocalendar()
    return f"{iso.year}-W{iso.week:02d}"


def workbook_path_for_model(
    monitored_model: MonitoredModel,
    week_start: date,
    output_dir: Path | None = None,
) -> Path:
    base_dir = output_dir or (config.OUTPUT_DIR / OUTPUT_SUBDIR)
    return base_dir / monitored_model.model_slug / f"{monitored_model.model_slug} {iso_week_label(week_start)}.xlsx"


def price_per_million(raw_price: Any) -> float | None:
    parsed = safe_parse_float(raw_price)
    if parsed is None:
        return None
    return round(parsed * 1_000_000, 6)


def format_compact_count(value: Any) -> str:
    parsed = safe_parse_float(value)
    if parsed is None:
        return ""
    if parsed >= 1_000_000:
        return _trim_decimal(parsed / 1_000_000, 2) + "M"
    if parsed >= 1_000:
        return _trim_decimal(parsed / 1_000, 1) + "K"
    return _trim_decimal(parsed, 0)


def _trim_decimal(value: float, digits: int) -> str:
    if digits <= 0:
        return str(int(round(value)))
    return f"{value:.{digits}f}".rstrip("0").rstrip(".")


def fetch_model_endpoints(model_id: str) -> list[dict[str, Any]]:
    url = f"{config.API_BASE_URL}/models/{model_id}/endpoints"
    headers = {
        "Accept": "application/json",
        "User-Agent": "openrouter-observer-weekly/price-uptime",
    }
    response = httpx.get(url, headers=headers, timeout=30.0)
    response.raise_for_status()
    payload = response.json()
    data = payload.get("data", payload) if isinstance(payload, dict) else payload
    if isinstance(data, dict):
        endpoints = data.get("endpoints", [])
    else:
        endpoints = data
    return endpoints if isinstance(endpoints, list) else []


def build_rows_from_api_and_page(
    endpoints: list[dict[str, Any]],
    page_cards: list[dict[str, str]],
) -> list[dict[str, Any]]:
    page_lookup = _build_page_lookup(page_cards)
    rows: list[dict[str, Any]] = []

    for endpoint in endpoints:
        if not isinstance(endpoint, dict):
            continue
        pricing = endpoint.get("pricing") if isinstance(endpoint.get("pricing"), dict) else {}
        provider = str(endpoint.get("provider_name") or endpoint.get("provider_display_name") or endpoint.get("tag") or "").strip()
        tag = str(endpoint.get("tag") or "").strip()
        page_card = _find_page_card(provider, tag, page_lookup)

        rows.append({
            "Provider": page_card.get("Provider") or provider,
            "Region": endpoint.get("region") or page_card.get("Region") or "",
            "Quantization": endpoint.get("quantization") or page_card.get("Quantization") or "",
            "Latency": page_card.get("Latency") or "",
            "Throughput": page_card.get("Throughput") or "",
            "Uptime": _uptime_1d(endpoint),
            "Total Context": format_compact_count(endpoint.get("context_length")) or page_card.get("Total Context") or "",
            "Max Output": format_compact_count(endpoint.get("max_completion_tokens")) or page_card.get("Max Output") or "",
            "Input Price": price_per_million(pricing.get("prompt") or pricing.get("input") or pricing.get("input_price")),
            "Output Price": price_per_million(pricing.get("completion") or pricing.get("output") or pricing.get("output_price")),
            "Cache Read": price_per_million(
                pricing.get("input_cache_read")
                or pricing.get("cache_read")
                or pricing.get("cache_read_price")
            ),
        })

    return rows


def _uptime_1d(endpoint: dict[str, Any]) -> float | None:
    raw = safe_parse_float(endpoint.get("uptime_last_1d"))
    if raw is None:
        return None
    return round(raw * 100, 1) if 0 <= raw <= 1 else round(raw, 1)


def _build_page_lookup(page_cards: list[dict[str, str]]) -> dict[str, dict[str, str]]:
    lookup: dict[str, dict[str, str]] = {}
    for card in page_cards:
        provider = card.get("Provider", "")
        if provider:
            lookup[normalize_for_match(provider)] = card
    return lookup


def _find_page_card(provider: str, tag: str, page_lookup: dict[str, dict[str, str]]) -> dict[str, str]:
    provider_key = normalize_for_match(provider)
    tag_provider_key = normalize_for_match(tag.split("/", 1)[0]) if tag else ""
    for key, card in page_lookup.items():
        if key == provider_key or key.startswith(provider_key) or provider_key.startswith(key):
            return card
        if tag_provider_key and (key == tag_provider_key or key.startswith(tag_provider_key) or tag_provider_key.startswith(key)):
            return card
    return {}


def scrape_provider_page_cards(model_url: str) -> list[dict[str, str]]:
    with sync_playwright() as playwright:
        browser = playwright.chromium.launch(headless=config.HEADLESS)
        page = browser.new_page(viewport={"width": 1920, "height": 1200})
        try:
            if not navigate_with_retry(page, model_url):
                raise RuntimeError(f"Could not open {model_url}")
            _expand_all_provider_cards(page)
            return _extract_provider_cards(page)
        finally:
            browser.close()


def _expand_all_provider_cards(page: Page) -> None:
    for _ in range(5):
        clicked = page.evaluate(
            """
            () => {
                const visible = (el) => {
                    const style = window.getComputedStyle(el);
                    const rect = el.getBoundingClientRect();
                    return style.display !== "none"
                        && style.visibility !== "hidden"
                        && rect.width > 0
                        && rect.height > 0;
                };
                const buttons = Array.from(document.querySelectorAll("button"))
                    .filter((button) => /Show\\s+\\d+\\s+more|Show more/i.test((button.innerText || button.textContent || "").trim()))
                    .filter(visible);
                for (const button of buttons) { button.click(); }
                return buttons.length;
            }
            """
        )
        if not clicked:
            break
        page.wait_for_timeout(1_000)


def _extract_provider_cards(page: Page) -> list[dict[str, str]]:
    raw_cards = page.evaluate(
        """
        () => {
            const visible = (el) => {
                const style = window.getComputedStyle(el);
                const rect = el.getBoundingClientRect();
                return style.display !== "none"
                    && style.visibility !== "hidden"
                    && rect.width > 0
                    && rect.height > 0;
            };
            return Array.from(document.querySelectorAll("article"))
                .filter(visible)
                .map((el) => (el.innerText || el.textContent || "").trim())
                .filter((text) => text.includes("PRICING") && text.includes("Latency") && text.includes("Throughput"));
        }
        """
    )
    return [_parse_provider_card_text(text) for text in raw_cards if text]


def _parse_provider_card_text(text: str) -> dict[str, str]:
    lines = clean_lines(text)
    if not lines:
        return {}

    provider = lines[0]
    latency = _value_after(lines, "Latency")
    throughput = _value_after(lines, "Throughput")
    total_context = _value_after(lines, "Total Context")
    max_output = _value_after(lines, "Max Output")

    latency_index = _line_index(lines, "Latency")
    attributes = lines[1:latency_index] if latency_index is not None else lines[1:3]
    region = attributes[0] if attributes else ""
    quantization = attributes[1] if len(attributes) > 1 else ""

    return {
        "Provider": provider,
        "Region": region,
        "Quantization": quantization,
        "Latency": latency,
        "Throughput": throughput,
        "Total Context": total_context,
        "Max Output": max_output,
    }


def _line_index(lines: list[str], label: str) -> int | None:
    for index, line in enumerate(lines):
        if line.lower() == label.lower():
            return index
    return None


def _value_after(lines: list[str], label: str) -> str:
    index = _line_index(lines, label)
    if index is None or index + 1 >= len(lines):
        return ""
    return lines[index + 1]


def _refresh_price_uptime_metadata(
    wb: Workbook,
    monitored_model: MonitoredModel,
    week_start: date,
    *,
    snapshot_dates: list[date] | None = None,
) -> None:
    tab_dates = list_snapshot_tab_dates(wb)
    if snapshot_dates:
        tab_dates = sorted(set(tab_dates) | set(snapshot_dates))
    updated_at = resolve_updated_at(None, snapshot_dates=tab_dates)
    extra_rows: list[tuple[str, str]] = []
    if tab_dates:
        extra_rows.append(
            ("已写入日期", "、".join(day.isoformat() for day in sorted(tab_dates))),
        )
    metadata_rows = build_standard_metadata(
        week_start,
        updated_at=updated_at,
        data_source=monitored_model.model_url,
        extra_rows=extra_rows,
    )
    upsert_metadata_sheet(wb, metadata_rows)


def write_weekly_workbook(
    monitored_model: MonitoredModel,
    week_start: date,
    snapshot_date: date,
    rows: list[dict[str, Any]],
    output_dir: Path | None = None,
) -> Path:
    path = workbook_path_for_model(monitored_model, week_start, output_dir=output_dir)
    path.parent.mkdir(parents=True, exist_ok=True)
    sheet_name = snapshot_date.isoformat()

    if path.exists():
        wb = load_workbook(path)
        if sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            wb.remove(ws)
    else:
        wb = Workbook()
        default = wb.active
        wb.remove(default)

    ws = wb.create_sheet(sheet_name)
    ws.append(HEADERS)
    for row in rows:
        ws.append([row.get(header) for header in HEADERS])
    _style_sheet(ws)
    _refresh_price_uptime_metadata(
        wb,
        monitored_model,
        week_start,
        snapshot_dates=[snapshot_date],
    )
    wb.save(path)
    logger.info("Saved Price&Uptime workbook to %s", path)
    return path


def _style_sheet(ws) -> None:
    header_fill = PatternFill("solid", fgColor="FCE4D6")
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    widths = [20, 12, 14, 12, 14, 12, 16, 14, 14, 14, 14]
    for index, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(index)].width = width

    price_cols = {9, 10, 11}
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            if cell.column in price_cols and isinstance(cell.value, (int, float)):
                cell.number_format = "$0.####"
            elif cell.column == 6 and isinstance(cell.value, (int, float)):
                cell.number_format = '0.0"%"'
    ws.freeze_panes = "A2"


def generate_price_uptime_workbook_for_model(
    monitored_model: MonitoredModel,
    week_start: date,
    snapshot_date: date,
    output_dir: Path | None = None,
) -> Path:
    endpoints = fetch_model_endpoints(monitored_model.model_id)
    page_cards = scrape_provider_page_cards(monitored_model.model_url)
    rows = build_rows_from_api_and_page(endpoints, page_cards)
    return write_weekly_workbook(monitored_model, week_start, snapshot_date, rows, output_dir=output_dir)


def list_missing_snapshot_pairs(
    week_start: date,
    *,
    through_date: date | None = None,
    models: list[MonitoredModel] | None = None,
    output_dir: Path | None = None,
) -> list[tuple[MonitoredModel, date]]:
    """Return (model, date) pairs missing a daily tab within the ISO week through through_date."""
    from pipeline.model_provider_usage import week_dates

    monitored_models = models or load_monitored_models()
    end = through_date or (date.today() - timedelta(days=1))
    pairs: list[tuple[MonitoredModel, date]] = []
    for snap_date in week_dates(week_start):
        if snap_date > end:
            continue
        sheet_name = snap_date.isoformat()
        for monitored_model in monitored_models:
            path = workbook_path_for_model(monitored_model, week_start, output_dir=output_dir)
            if not path.exists():
                pairs.append((monitored_model, snap_date))
                continue
            wb = load_workbook(path, read_only=True)
            try:
                if sheet_name not in wb.sheetnames:
                    pairs.append((monitored_model, snap_date))
            finally:
                wb.close()
    return pairs


def generate_price_uptime_workbooks(
    week_start: date,
    snapshot_date: date,
    models: list[MonitoredModel] | None = None,
    output_dir: Path | None = None,
) -> list[Path]:
    monitored_models = models or load_monitored_models()
    paths: list[Path] = []
    for monitored_model in monitored_models:
        paths.append(generate_price_uptime_workbook_for_model(monitored_model, week_start, snapshot_date, output_dir=output_dir))
    return paths
