"""Generate weekly Core Model Usage Excel from Activity API daily totals."""

from __future__ import annotations

from datetime import date, timedelta
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from pipeline import config
from pipeline.model_ranking_weekly import (
    format_cn_date,
    format_data_range,
    scrape_top_models_payload_for_week,
)
from pipeline.logger import get_logger
from pipeline.model_activity import fetch_model_activity_totals, model_activity_url, model_display_name_from_endpoints
from pipeline.model_provider_price_uptime import (
    MonitoredModel,
    fetch_model_endpoints,
    iso_week_label,
    load_monitored_models,
    parse_week_start,
)
from pipeline.data_availability import activity_daily_display
from pipeline.model_provider_usage import week_dates
from pipeline.workbook_metadata import resolve_updated_at, style_metadata_sheet

logger = get_logger()

OUTPUT_SUBDIR = "Core_Models"
WORKBOOK_BASENAME = "Core Model Usage"
RANKINGS_SOURCE_URL = "https://openrouter.ai/rankings"
FIXED_HEADERS = ["模型ID", "模型名称"]


def _format_tokens_compact(tokens: float) -> str:
    units = [
        ("T", 1_000_000_000_000),
        ("B", 1_000_000_000),
        ("M", 1_000_000),
        ("K", 1_000),
    ]
    for suffix, multiplier in units:
        if abs(tokens) >= multiplier:
            value = tokens / multiplier
            text = f"{value:.2f}".rstrip("0").rstrip(".")
            return f"{text}{suffix}"
    return str(int(tokens))


def _tokens_to_t(tokens: int | float) -> float:
    return round(float(tokens) / 1_000_000_000_000, 4)


def _pct(part: int | float, total: int | float | None) -> float | None:
    if not total:
        return None
    return round(float(part) / float(total), 4)


def usage_headers(days: list[date]) -> list[str]:
    return FIXED_HEADERS + [day.isoformat() for day in days] + ["周合计", "换算后用量（T）", "占比（%）"]


def output_path_for_week(week_start: date, output_dir: Path | None = None) -> Path:
    base_dir = output_dir or (config.OUTPUT_DIR / OUTPUT_SUBDIR)
    return base_dir / f"{WORKBOOK_BASENAME} {iso_week_label(week_start)}.xlsx"


def build_usage_row(
    monitored_model: MonitoredModel,
    model_name: str,
    days: list[date],
    activity_totals: dict[str, int],
    rankings_total: int | float | None,
    *,
    activity_api_failed: bool = False,
) -> dict[str, Any]:
    week_sum = 0
    row: dict[str, Any] = {
        "模型ID": monitored_model.model_id,
        "模型名称": model_name,
    }
    for day in days:
        tokens = int(activity_totals.get(day.isoformat(), 0) or 0)
        cell = activity_daily_display(day, tokens, api_failed=activity_api_failed)
        if isinstance(cell, int):
            week_sum += cell
            row[day.isoformat()] = _format_tokens_compact(cell)
        else:
            row[day.isoformat()] = cell
    row["周合计"] = _format_tokens_compact(week_sum) if week_sum else ""
    row["换算后用量（T）"] = _tokens_to_t(week_sum) if week_sum else None
    row["占比（%）"] = _pct(week_sum, rankings_total)
    return row


def build_metadata(
    week_start: date,
    rankings_total: int | float | None,
    rankings_total_text: str | None,
    updated_at: date | None = None,
    *,
    snapshot_dates: list[date] | None = None,
) -> list[tuple[str, str]]:
    resolved = resolve_updated_at(
        updated_at,
        snapshot_dates=snapshot_dates or [week_start + timedelta(days=7)],
    )
    total_text = rankings_total_text or (
        _format_tokens_compact(float(rankings_total)) if rankings_total else ""
    )
    return [
        ("数据范围", format_data_range(week_start)),
        ("数据更新时间", format_cn_date(resolved)),
        ("数据来源", f"Model Activity API; Rankings: {RANKINGS_SOURCE_URL}"),
        ("数据周", iso_week_label(week_start)),
        ("排行榜 Total", total_text),
        ("排行榜 Total（T）", str(_tokens_to_t(rankings_total)) if rankings_total else ""),
        ("占比分母说明", "占比（%）= 模型周合计 / 该周 Top Models 图 Total"),
    ]


def save_core_models_usage_workbook(
    rows: list[dict[str, Any]],
    week_start: date,
    days: list[date],
    rankings_total: int | float | None,
    rankings_total_text: str | None = None,
    output_dir: Path | None = None,
    updated_at: date | None = None,
) -> Path:
    out_path = output_path_for_week(week_start, output_dir=output_dir)
    out_path.parent.mkdir(parents=True, exist_ok=True)
    headers = usage_headers(days)

    wb = Workbook()
    ws = wb.active
    ws.title = "Usage"
    ws.append(headers)
    for row in rows:
        ws.append([row.get(header) for header in headers])

    meta_ws = wb.create_sheet("Metadata")
    meta_ws.append(["字段", "内容"])
    for key, value in build_metadata(
        week_start,
        rankings_total,
        rankings_total_text,
        updated_at=updated_at,
        snapshot_dates=days,
    ):
        meta_ws.append([key, value])

    _style_usage_sheet(ws, len(headers))
    style_metadata_sheet(meta_ws)
    wb.save(out_path)
    logger.info("Saved Core Model Usage Excel to %s", out_path)
    return out_path


def _style_usage_sheet(ws, column_count: int) -> None:
    header_fill = PatternFill("solid", fgColor="D9EAF7")
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 24
    for index in range(3, column_count + 1):
        ws.column_dimensions[get_column_letter(index)].width = 14
    pct_col = column_count
    t_col = column_count - 1
    for row in ws.iter_rows(min_row=2, min_col=t_col, max_col=t_col):
        for cell in row:
            if isinstance(cell.value, (int, float)):
                cell.number_format = "0.0000"
    for row in ws.iter_rows(min_row=2, min_col=pct_col, max_col=pct_col):
        for cell in row:
            if isinstance(cell.value, (int, float)):
                cell.number_format = "0.00%"
    ws.freeze_panes = "A2"


_style_metadata_sheet = style_metadata_sheet


def collect_usage_rows(
    models: list[MonitoredModel],
    week_start: date,
    rankings_total: int | float | None,
) -> tuple[list[dict[str, Any]], list[date]]:
    days = week_dates(week_start)
    rows: list[dict[str, Any]] = []
    for monitored_model in models:
        logger.info("Fetching activity totals for %s", monitored_model.model_id)
        activity_api_failed = False
        try:
            activity_totals = fetch_model_activity_totals(monitored_model.model_id)
        except Exception as exc:
            logger.warning("Activity API failed for %s: %s", monitored_model.model_id, exc)
            activity_totals = {}
            activity_api_failed = True
        try:
            endpoints = fetch_model_endpoints(monitored_model.model_id)
            model_name = model_display_name_from_endpoints(endpoints, monitored_model.model_slug)
        except Exception as exc:
            logger.warning("Could not resolve display name for %s: %s", monitored_model.model_id, exc)
            model_name = monitored_model.model_slug.replace("-", " ").title()
        week_sum = sum(int(activity_totals.get(day.isoformat(), 0)) for day in days)
        if week_sum == 0:
            logger.warning(
                "No activity tokens in target week for %s (%s)",
                monitored_model.model_id,
                model_activity_url(monitored_model.model_id),
            )
        rows.append(
            build_usage_row(
                monitored_model,
                model_name,
                days,
                activity_totals,
                rankings_total,
                activity_api_failed=activity_api_failed,
            )
        )
    return rows, days


def generate_core_models_usage_excel(
    week_start: date,
    models: list[MonitoredModel] | None = None,
    output_dir: Path | None = None,
    rankings_payload: dict[str, Any] | None = None,
) -> Path:
    config.ensure_dirs()
    monitored_models = models or load_monitored_models()
    if rankings_payload is None:
        logger.info("Fetching Top Models rankings payload for week %s", week_start.isoformat())
        rankings_payload = scrape_top_models_payload_for_week(week_start)

    rankings_total = rankings_payload.get("total_tokens")
    if rankings_total is None:
        model_items = rankings_payload.get("models") or []
        rankings_total = sum(float(item.get("tokens") or 0) for item in model_items)
        logger.warning("Rankings payload had no Total; using model row sum as denominator")

    rows, days = collect_usage_rows(monitored_models, week_start, rankings_total)
    return save_core_models_usage_workbook(
        rows,
        week_start,
        days,
        rankings_total,
        rankings_total_text=rankings_payload.get("total_tokens_text"),
        output_dir=output_dir,
    )
