"""Generate weekly model provider usage monitoring workbooks."""

from __future__ import annotations

from datetime import date, timedelta
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from playwright.sync_api import sync_playwright

from pipeline import config
from pipeline.logger import get_logger
from pipeline.model_activity import fetch_model_activity_totals, model_display_name_from_endpoints
from pipeline.model_provider_price_uptime import (
    MonitoredModel,
    fetch_model_endpoints,
    load_monitored_models,
    parse_snapshot_date,
    parse_week_start,
    workbook_path_for_model as price_uptime_workbook_path_for_model,
)
from pipeline.core_providers import load_core_providers
from pipeline.data_availability import EXPIRED_UNAVAILABLE_TEXT, is_past_day
from pipeline.model_provider_price_uptime import _refresh_price_uptime_metadata
from pipeline.utils import navigate_with_retry, normalize_for_match, scrape_provider_chart_usage

logger = get_logger()

MISSING_DISPLAY_TEXT = "未展示"
SKIPPED_LOW_USAGE_TEXT = "用量少未查询"
CORE_COVERAGE_THRESHOLD = 0.9
PRICE_UPTIME_USAGE_COLUMNS = ["展示状态", "Provider 承接用量", "Provider当日总量", "承接占比"]


def week_dates(week_start: date) -> list[date]:
    return [week_start + timedelta(days=offset) for offset in range(7)]


def read_price_uptime_providers(
    monitored_model: MonitoredModel,
    week_start: date,
    snapshot_date: date,
    price_uptime_output_dir: Path | None = None,
) -> list[str]:
    path = price_uptime_workbook_path_for_model(
        monitored_model,
        week_start,
        output_dir=price_uptime_output_dir,
    )
    sheet_name = snapshot_date.isoformat()
    if not path.exists():
        raise FileNotFoundError(f"Price&Uptime workbook not found: {path}")

    wb = load_workbook(path, data_only=True)
    if sheet_name not in wb.sheetnames:
        raise ValueError(f"Price&Uptime workbook {path} does not contain daily tab {sheet_name}")

    ws = wb[sheet_name]
    headers = [cell.value for cell in ws[1]]
    try:
        provider_col = headers.index("Provider") + 1
    except ValueError as exc:
        raise ValueError(f"Price&Uptime tab {sheet_name} has no Provider column") from exc

    providers: list[str] = []
    seen: set[str] = set()
    for row in range(2, ws.max_row + 1):
        provider = str(ws.cell(row, provider_col).value or "").strip()
        key = normalize_for_match(provider)
        if provider and key not in seen:
            providers.append(provider)
            seen.add(key)
    return providers


def build_provider_slug_map(endpoints: list[dict[str, Any]]) -> dict[str, str]:
    mapping: dict[str, str] = {}
    for endpoint in endpoints:
        if not isinstance(endpoint, dict):
            continue
        provider_name = str(endpoint.get("provider_name") or endpoint.get("provider_display_name") or "").strip()
        tag = str(endpoint.get("tag") or endpoint.get("provider_slug") or "").strip()
        provider_slug = tag.split("/", 1)[0].strip().lower()
        if not provider_slug:
            continue
        for key in {provider_name, provider_slug, tag}:
            normalized = normalize_for_match(key)
            if normalized:
                mapping[normalized] = provider_slug
    return mapping


def resolve_provider_slug(provider_name: str, slug_map: dict[str, str]) -> str | None:
    key = normalize_for_match(provider_name)
    if key in slug_map:
        return slug_map[key]
    for candidate, slug in slug_map.items():
        if key.startswith(candidate) or candidate.startswith(key):
            return slug
    return None


def provider_url(provider_slug: str) -> str:
    return f"https://openrouter.ai/provider/{provider_slug}"


def _expired_usage_fields(provider_slug: str | None) -> dict[str, Any]:
    return {
        "Provider URL": provider_url(provider_slug) if provider_slug else "",
        "Provider 承接用量": EXPIRED_UNAVAILABLE_TEXT,
        "Provider当日总量": EXPIRED_UNAVAILABLE_TEXT,
        "承接占比": EXPIRED_UNAVAILABLE_TEXT,
        "展示状态": EXPIRED_UNAVAILABLE_TEXT,
    }


def build_usage_row(
    provider_name: str,
    provider_slug: str | None,
    chart_payload: dict[str, Any] | None,
    model_display_name: str,
    model_daily_total: int | None,
    snapshot_date: date | None = None,
) -> dict[str, Any]:
    day = snapshot_date or date.today()
    if chart_payload is None and is_past_day(day):
        return {
            "Provider": provider_name,
            "模型每日总量": model_daily_total if model_daily_total is not None else EXPIRED_UNAVAILABLE_TEXT,
            "模型用量": None,
            **_expired_usage_fields(provider_slug),
        }

    provider_total_text = chart_payload.get("total_tokens_text") if chart_payload else ""
    matched = None
    for item in (chart_payload or {}).get("models", []):
        if _chart_model_matches(item.get("model_name", ""), model_display_name):
            matched = item
            break

    if not matched:
        model_total_cell: Any = model_daily_total
        if model_daily_total is None and is_past_day(day):
            model_total_cell = EXPIRED_UNAVAILABLE_TEXT
        return {
            "Provider": provider_name,
            "Provider URL": provider_url(provider_slug) if provider_slug else "",
            "模型每日总量": model_total_cell,
            "模型用量": None,
            "Provider 承接用量": "",
            "承接占比": None,
            "展示状态": MISSING_DISPLAY_TEXT,
            "Provider当日总量": provider_total_text or "",
        }

    tokens = matched.get("tokens") or 0
    share = (float(tokens) / model_daily_total) if model_daily_total else None
    return {
        "Provider": provider_name,
        "Provider URL": provider_url(provider_slug) if provider_slug else "",
        "模型每日总量": model_daily_total,
        "模型用量": int(tokens),
        "Provider 承接用量": matched.get("tokens_text") or "",
        "承接占比": share,
        "展示状态": "已展示",
        "Provider当日总量": provider_total_text or "",
    }


def _chart_model_matches(chart_name: str, model_display_name: str) -> bool:
    chart_key = normalize_for_match(chart_name)
    candidates = {model_display_name}
    if ":" in model_display_name:
        candidates.add(model_display_name.split(":", 1)[1].strip())
    return chart_key in {normalize_for_match(candidate) for candidate in candidates if candidate}


def provider_in_core_list(
    provider_name: str,
    slug_map: dict[str, str],
    core_providers: list[dict[str, str]],
) -> bool:
    core_slugs = {entry["provider_slug"] for entry in core_providers}
    provider_slug = resolve_provider_slug(provider_name, slug_map)
    if provider_slug and provider_slug.lower() in core_slugs:
        return True
    provider_key = normalize_for_match(provider_name)
    for entry in core_providers:
        configured_name = entry.get("provider_name") or ""
        if configured_name and provider_key == normalize_for_match(configured_name):
            return True
    return False


def partition_providers_by_core(
    provider_names: list[str],
    slug_map: dict[str, str],
    core_providers: list[dict[str, str]],
) -> tuple[list[str], list[str]]:
    core_names: list[str] = []
    non_core_names: list[str] = []
    for provider_name in provider_names:
        if provider_in_core_list(provider_name, slug_map, core_providers):
            core_names.append(provider_name)
        else:
            non_core_names.append(provider_name)
    return core_names, non_core_names


def build_skipped_low_usage_row(
    provider_name: str,
    provider_slug: str | None,
    model_daily_total: int | None,
    snapshot_date: date,
) -> dict[str, Any]:
    day = snapshot_date
    model_total_cell: Any = model_daily_total
    if model_daily_total is None and is_past_day(day):
        model_total_cell = EXPIRED_UNAVAILABLE_TEXT
    return {
        "Provider": provider_name,
        "Provider URL": provider_url(provider_slug) if provider_slug else "",
        "模型每日总量": model_total_cell,
        "模型用量": None,
        "Provider 承接用量": "",
        "承接占比": None,
        "展示状态": SKIPPED_LOW_USAGE_TEXT,
        "Provider当日总量": "",
    }


def core_usage_coverage_ratio(rows: list[dict[str, Any]], model_daily_total: int | None) -> float:
    if not model_daily_total or model_daily_total <= 0:
        return 0.0
    covered = sum(int(row["模型用量"]) for row in rows if isinstance(row.get("模型用量"), int))
    return covered / model_daily_total


def core_coverage_sufficient(rows: list[dict[str, Any]], model_daily_total: int | None) -> bool:
    return core_usage_coverage_ratio(rows, model_daily_total) >= CORE_COVERAGE_THRESHOLD


def build_usage_rows_for_date(
    provider_names: list[str],
    provider_payloads: dict[str, dict[str, Any] | None],
    slug_map: dict[str, str],
    model_display_name: str,
    model_daily_total: int | None,
    snapshot_date: date,
) -> list[dict[str, Any]]:
    rows = []
    for provider_name in provider_names:
        provider_slug = resolve_provider_slug(provider_name, slug_map)
        payload = provider_payloads.get(provider_name)
        rows.append(
            build_usage_row(
                provider_name,
                provider_slug,
                payload,
                model_display_name,
                model_daily_total,
                snapshot_date=snapshot_date,
            )
        )
    return rows


def _merge_provider_payloads(
    target: dict[str, dict[str, dict[str, Any] | None]],
    incoming: dict[str, dict[str, dict[str, Any] | None]],
) -> None:
    for provider_name, by_date in incoming.items():
        target.setdefault(provider_name, {}).update(by_date)


def update_price_uptime_workbook_with_usage(
    monitored_model: MonitoredModel,
    week_start: date,
    rows_by_date: dict[date, list[dict[str, Any]]],
    price_uptime_output_dir: Path | None = None,
) -> Path:
    path = price_uptime_workbook_path_for_model(monitored_model, week_start, output_dir=price_uptime_output_dir)
    if not path.exists():
        raise FileNotFoundError(f"Price&Uptime workbook not found: {path}")

    wb = load_workbook(path)
    for snapshot_date, usage_rows in rows_by_date.items():
        sheet_name = snapshot_date.isoformat()
        if sheet_name not in wb.sheetnames:
            raise ValueError(f"Price&Uptime workbook {path} does not contain daily tab {sheet_name}")
        ws = wb[sheet_name]
        _ensure_price_uptime_usage_layout(ws)
        headers = [cell.value for cell in ws[1]]
        col_by_header = {header: index + 1 for index, header in enumerate(headers)}
        usage_by_provider = {
            normalize_for_match(str(row.get("Provider") or "")): row
            for row in usage_rows
            if row.get("Provider")
        }

        for row_index in range(2, ws.max_row + 1):
            provider = str(ws.cell(row_index, col_by_header["Provider"]).value or "").strip()
            usage_row = usage_by_provider.get(normalize_for_match(provider), {})
            ws.cell(row_index, col_by_header["Provider URL"]).value = usage_row.get("Provider URL")
            for header in PRICE_UPTIME_USAGE_COLUMNS:
                ws.cell(row_index, col_by_header[header]).value = usage_row.get(header)
        _style_price_uptime_usage_columns(ws)

    _refresh_price_uptime_metadata(
        wb,
        monitored_model,
        week_start,
        snapshot_dates=list(rows_by_date.keys()),
    )
    wb.save(path)
    logger.info("Updated Price&Uptime workbook with Usage columns: %s", path)
    return path


def _ensure_price_uptime_usage_layout(ws) -> None:
    headers = [cell.value for cell in ws[1]]
    if "Provider" not in headers:
        raise ValueError(f"Sheet {ws.title} has no Provider column")

    if "Provider URL" not in headers:
        provider_col = headers.index("Provider") + 1
        ws.insert_cols(provider_col + 1)
        ws.cell(1, provider_col + 1).value = "Provider URL"

    headers = [cell.value for cell in ws[1]]
    cache_read_col = headers.index("Cache Read") + 1 if "Cache Read" in headers else ws.max_column
    for offset, header in enumerate(PRICE_UPTIME_USAGE_COLUMNS, start=1):
        headers = [cell.value for cell in ws[1]]
        if header in headers:
            continue
        ws.insert_cols(cache_read_col + offset)
        ws.cell(1, cache_read_col + offset).value = header


def _style_price_uptime_usage_columns(ws) -> None:
    headers = [cell.value for cell in ws[1]]
    col_by_header = {header: index + 1 for index, header in enumerate(headers)}
    fill = PatternFill("solid", fgColor="E2F0D9")
    for header in ["Provider URL", *PRICE_UPTIME_USAGE_COLUMNS]:
        col = col_by_header.get(header)
        if not col:
            continue
        ws.cell(1, col).font = Font(bold=True)
        ws.cell(1, col).fill = fill
        ws.cell(1, col).alignment = Alignment(horizontal="center")
        ws.column_dimensions[get_column_letter(col)].width = 34 if header == "Provider URL" else 18

    share_col = col_by_header.get("承接占比")
    if share_col:
        for row in range(2, ws.max_row + 1):
            cell = ws.cell(row, share_col)
            if isinstance(cell.value, (int, float)):
                cell.number_format = "0.00%"


def scrape_provider_payloads_for_dates(
    provider_names: list[str],
    slug_map: dict[str, str],
    target_dates: list[date],
    *,
    latest_n: int | None = None,
) -> dict[str, dict[str, dict[str, Any] | None]]:
    wanted = {day.isoformat() for day in target_dates}
    results: dict[str, dict[str, dict[str, Any] | None]] = {
        provider_name: {day: None for day in wanted}
        for provider_name in provider_names
    }

    with sync_playwright() as playwright:
        browser = playwright.chromium.launch(headless=config.HEADLESS)
        page = browser.new_page(viewport={"width": 1600, "height": 1000})
        try:
            for provider_name in provider_names:
                provider_slug = resolve_provider_slug(provider_name, slug_map)
                if not provider_slug:
                    logger.warning("Could not resolve provider slug for %s", provider_name)
                    continue
                url = provider_url(provider_slug)
                logger.info("Scraping provider usage: %s (%s)", provider_name, url)
                # Retry the whole page when a scrape comes back empty: the most common
                # accuracy failure is the recharts chart not having rendered yet, which
                # leaves that provider blank for every date. Reload and try again before
                # giving up so a slow render doesn't silently drop a provider.
                chart_usage: list[dict[str, Any]] = []
                for attempt in range(1, 4):
                    if not navigate_with_retry(page, url):
                        logger.warning("Could not open provider page for %s (attempt %d/3)", provider_name, attempt)
                        continue
                    chart_usage = scrape_provider_chart_usage(page, provider_slug, latest_n=latest_n)
                    if chart_usage:
                        break
                    logger.warning("[%s] empty chart scrape (attempt %d/3); reloading", provider_name, attempt)
                if not chart_usage:
                    logger.warning("Provider %s returned no chart data after retries", provider_name)
                    continue
                by_date = {item.get("chart_date"): item for item in chart_usage}
                for day in wanted:
                    results[provider_name][day] = by_date.get(day)
        finally:
            browser.close()

    return results


def generate_usage_workbook_for_model(
    monitored_model: MonitoredModel,
    week_start: date,
    target_dates: list[date],
    price_uptime_output_dir: Path | None = None,
    core_providers: list[dict[str, str]] | None = None,
    *,
    latest_n: int | None = None,
    skip_noncore_fallback: bool = False,
) -> Path:
    endpoints = fetch_model_endpoints(monitored_model.model_id)
    slug_map = build_provider_slug_map(endpoints)
    activity_totals = fetch_model_activity_totals(monitored_model.model_id)
    core_list = core_providers if core_providers is not None else load_core_providers()

    provider_names_by_date = {
        snapshot_date: read_price_uptime_providers(monitored_model, week_start, snapshot_date, price_uptime_output_dir)
        for snapshot_date in target_dates
    }
    core_names_union = _dedupe_names(
        [
            name
            for snapshot_date in target_dates
            for name in partition_providers_by_core(
                provider_names_by_date[snapshot_date],
                slug_map,
                core_list,
            )[0]
        ]
    )
    payloads_by_provider: dict[str, dict[str, dict[str, Any] | None]] = {}
    if core_names_union:
        logger.info(
            "Scraping %d core providers first for %s",
            len(core_names_union),
            monitored_model.model_slug,
        )
        _merge_provider_payloads(
            payloads_by_provider,
            scrape_provider_payloads_for_dates(core_names_union, slug_map, target_dates, latest_n=latest_n),
        )

    model_display_name = model_display_name_from_endpoints(endpoints, monitored_model.model_slug)
    scraped_non_core: set[str] = set()
    rows_by_date: dict[date, list[dict[str, Any]]] = {}
    for snapshot_date, provider_names_for_date in provider_names_by_date.items():
        day = snapshot_date.isoformat()
        model_daily_total = activity_totals.get(day)
        core_names, non_core_names = partition_providers_by_core(
            provider_names_for_date,
            slug_map,
            core_list,
        )
        core_payloads = {
            provider_name: payloads_by_provider.get(provider_name, {}).get(day)
            for provider_name in core_names
        }
        core_rows = build_usage_rows_for_date(
            core_names,
            core_payloads,
            slug_map,
            model_display_name,
            model_daily_total,
            snapshot_date,
        )
        core_rows_by_name = {row["Provider"]: row for row in core_rows}

        if non_core_names and (skip_noncore_fallback or core_coverage_sufficient(core_rows, model_daily_total)):
            ratio = core_usage_coverage_ratio(core_rows, model_daily_total)
            reason = "intraday fast mode" if skip_noncore_fallback else f"core providers cover {ratio*100:.1f}%"
            logger.info(
                "Skipping %d non-core providers on %s for %s (%s)",
                len(non_core_names),
                day,
                monitored_model.model_slug,
                reason,
            )
            rows = []
            for provider_name in provider_names_for_date:
                if provider_name in non_core_names:
                    rows.append(
                        build_skipped_low_usage_row(
                            provider_name,
                            resolve_provider_slug(provider_name, slug_map),
                            model_daily_total,
                            snapshot_date,
                        )
                    )
                else:
                    rows.append(core_rows_by_name[provider_name])
        else:
            need_scrape = [name for name in non_core_names if name not in scraped_non_core]
            if need_scrape:
                logger.info(
                    "Core coverage below %.0f%% on %s for %s; scraping %d non-core providers",
                    CORE_COVERAGE_THRESHOLD * 100,
                    day,
                    monitored_model.model_slug,
                    len(need_scrape),
                )
                _merge_provider_payloads(
                    payloads_by_provider,
                    scrape_provider_payloads_for_dates(need_scrape, slug_map, target_dates, latest_n=latest_n),
                )
                scraped_non_core.update(need_scrape)
            provider_payloads = {
                provider_name: payloads_by_provider.get(provider_name, {}).get(day)
                for provider_name in provider_names_for_date
            }
            rows = build_usage_rows_for_date(
                provider_names_for_date,
                provider_payloads,
                slug_map,
                model_display_name,
                model_daily_total,
                snapshot_date,
            )
        rows_by_date[snapshot_date] = rows

    return update_price_uptime_workbook_with_usage(
        monitored_model,
        week_start,
        rows_by_date,
        price_uptime_output_dir=price_uptime_output_dir,
    )


def list_incomplete_usage_snapshot_dates(
    week_start: date,
    models: list[MonitoredModel] | None = None,
    price_uptime_output_dir: Path | None = None,
    *,
    through_date: date | None = None,
) -> list[date]:
    """Return week dates needing Step 3.2 (missing tab or tab without Usage columns)."""
    monitored_models = models or load_monitored_models()
    end = through_date or (date.today() - timedelta(days=1))
    allowed_days = [day for day in week_dates(week_start) if day <= end]
    incomplete: set[date] = set()
    for snap_date in allowed_days:
        sheet_name = snap_date.isoformat()
        for monitored_model in monitored_models:
            path = price_uptime_workbook_path_for_model(
                monitored_model,
                week_start,
                output_dir=price_uptime_output_dir,
            )
            if not path.exists():
                incomplete.add(snap_date)
                break
            wb = load_workbook(path, read_only=True, data_only=True)
            try:
                if sheet_name not in wb.sheetnames:
                    incomplete.add(snap_date)
                    break
                headers = [cell.value for cell in wb[sheet_name][1]]
                if "展示状态" not in headers:
                    incomplete.add(snap_date)
                    break
            finally:
                wb.close()
    return sorted(incomplete)


def generate_usage_workbooks(
    week_start: date,
    target_dates: list[date],
    models: list[MonitoredModel] | None = None,
    price_uptime_output_dir: Path | None = None,
    *,
    latest_n: int | None = None,
    skip_noncore_fallback: bool = False,
) -> list[Path]:
    monitored_models = models or load_monitored_models()
    return [
        generate_usage_workbook_for_model(
            monitored_model,
            week_start,
            target_dates,
            price_uptime_output_dir=price_uptime_output_dir,
            latest_n=latest_n,
            skip_noncore_fallback=skip_noncore_fallback,
        )
        for monitored_model in monitored_models
    ]


def _dedupe_names(names: list[str]) -> list[str]:
    result: list[str] = []
    seen: set[str] = set()
    for name in names:
        key = normalize_for_match(name)
        if name and key not in seen:
            result.append(name)
            seen.add(key)
    return result
