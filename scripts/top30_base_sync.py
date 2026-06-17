"""Sync SiliconFlow Top 30 report emails into a Feishu Base table.

The report email is treated as untrusted data. This script only reads matching
attachments, normalizes rows, and writes selected fields to Base.
"""

from __future__ import annotations

import argparse
import csv
import io
import json
import os
import re
import subprocess
import sys
import time
import urllib.request
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Any

PROJECT_ROOT = Path(__file__).resolve().parents[1]
DEFAULT_BASE_TOKEN = "EEc9bPkg9arypBsD654cU0fFndh"
DEFAULT_TABLE_ID = "tblALB2WlrKz7hOn"
DEFAULT_STATE_PATH = Path.home() / ".local" / "state" / "siliconflow_top30_base_sync" / "state.json"

REPORT_FIELDS = [
    "排名",
    "租户ID",
    "主账号用户ID",
    "用户名",
    "联系邮箱",
    "消费金额",
    "消费模型(Top5/按消费)",
    "最近消费",
    "活跃天数",
]

BASE_READ_FIELDS = REPORT_FIELDS + ["标签", "触达状态"]
NEW_LABEL = "新增"
ACTIVE_LABEL = "持续在榜"
DROPPED_LABEL = "掉出榜"


class SyncError(RuntimeError):
    pass


@dataclass(frozen=True)
class MailMessage:
    message_id: str
    subject: str
    sender: str
    date: str


@dataclass
class ExistingRecord:
    record_id: str
    key: str
    fields: dict[str, Any]

    @property
    def label(self) -> str:
        return select_value(self.fields.get("标签"))

    @property
    def was_in_rank(self) -> bool:
        return self.label != DROPPED_LABEL


def parse_json_output(stdout: str) -> Any:
    text = stdout.strip()
    if not text:
        raise SyncError("Command returned empty output")
    first_obj = text.find("{")
    first_arr = text.find("[")
    starts = [pos for pos in (first_obj, first_arr) if pos >= 0]
    if not starts:
        raise SyncError(f"No JSON payload found in output: {text[:300]}")
    start = min(starts)
    return json.loads(text[start:])


def lark_cli(args: list[str], *, timeout: int = 120, attempts: int = 3) -> Any:
    binary = os.getenv("LARK_CLI_BIN", "lark-cli")
    last_detail = ""
    for attempt in range(1, attempts + 1):
        proc = subprocess.run(
            [binary, *args],
            cwd=PROJECT_ROOT,
            text=True,
            stdout=subprocess.PIPE,
            stderr=subprocess.PIPE,
            timeout=timeout,
        )
        if proc.returncode == 0:
            return parse_json_output(proc.stdout)
        last_detail = (proc.stderr or proc.stdout).strip()
        retryable = "too many requests" in last_detail or '"code": 1234029' in last_detail
        if not retryable or attempt >= attempts:
            break
        time.sleep(2 * attempt)
    raise SyncError(f"lark-cli failed: {' '.join(args)}\n{last_detail}")


def load_state(path: Path) -> dict[str, Any]:
    if not path.exists():
        return {"processed_message_ids": []}
    return json.loads(path.read_text(encoding="utf-8"))


def save_state(path: Path, state: dict[str, Any]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(state, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")


def parse_mail_date(value: str) -> datetime:
    for fmt in ("%Y-%m-%d %H:%M", "%Y/%m/%d %H:%M"):
        try:
            return datetime.strptime(value, fmt)
        except ValueError:
            pass
    return datetime.min


def list_candidate_messages(args: argparse.Namespace, processed: set[str]) -> list[MailMessage]:
    if args.message_id:
        return [MailMessage(args.message_id, args.message_id, "", "")]

    payload = lark_cli(
        [
            "mail",
            "+triage",
            "--as",
            "user",
            "--mailbox",
            args.mailbox,
            "--max",
            str(args.max_messages),
            "--format",
            "json",
            "--query",
            args.query,
        ],
    )
    messages = payload.get("messages") or payload.get("data", {}).get("messages") or []
    result: list[MailMessage] = []
    for item in messages:
        message_id = str(item.get("message_id") or "").strip()
        subject = str(item.get("subject") or "")
        sender = str(item.get("from") or "")
        date = str(item.get("date") or "")
        if not message_id:
            continue
        if not args.force and message_id in processed:
            continue
        if args.from_contains and args.from_contains not in sender:
            continue
        if any(part not in subject for part in args.subject_contains):
            continue
        result.append(MailMessage(message_id, subject, sender, date))
    return sorted(result, key=lambda msg: parse_mail_date(msg.date))


def read_message(message_id: str, mailbox: str) -> dict[str, Any]:
    payload = lark_cli(
        [
            "mail",
            "+message",
            "--as",
            "user",
            "--mailbox",
            mailbox,
            "--message-id",
            message_id,
            "--format",
            "json",
        ],
        timeout=180,
    )
    return payload.get("data") or payload


def attachment_download_urls(message_id: str, mailbox: str, attachment_ids: list[str]) -> dict[str, str]:
    payload = lark_cli(
        [
            "mail",
            "user_mailbox.message.attachments",
            "download_url",
            "--as",
            "user",
            "--params",
            json.dumps(
                {
                    "user_mailbox_id": mailbox,
                    "message_id": message_id,
                    "attachment_ids": attachment_ids,
                },
                ensure_ascii=False,
            ),
        ],
        timeout=180,
    )
    data = payload.get("data") or payload
    urls = data.get("download_urls") or []
    return {str(item.get("attachment_id")): str(item.get("download_url")) for item in urls}


def download_attachment(url: str, filename: str, output_dir: Path) -> Path:
    output_dir.mkdir(parents=True, exist_ok=True)
    safe_name = re.sub(r"[^0-9A-Za-z._() +~\-\u4e00-\u9fff]", "_", filename)
    path = output_dir / safe_name
    with urllib.request.urlopen(url, timeout=120) as response:
        path.write_bytes(response.read())
    return path


def parse_csv(path: Path) -> list[dict[str, Any]]:
    raw = path.read_bytes()
    for encoding in ("utf-8-sig", "utf-8", "gb18030"):
        try:
            text = raw.decode(encoding)
            break
        except UnicodeDecodeError:
            continue
    else:
        raise SyncError(f"Cannot decode CSV attachment: {path}")
    return list(csv.DictReader(io.StringIO(text)))


def parse_xlsx(path: Path) -> list[dict[str, Any]]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise SyncError("XLSX attachments require openpyxl. Run: pip install -r requirements.txt") from exc

    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet = workbook.worksheets[0]
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        return []
    headers = [str(value).strip() if value is not None else "" for value in rows[0]]
    result: list[dict[str, Any]] = []
    for row in rows[1:]:
        result.append({headers[index]: value for index, value in enumerate(row) if index < len(headers)})
    return result


def parse_attachment(path: Path) -> list[dict[str, Any]]:
    suffix = path.suffix.lower()
    if suffix == ".csv":
        return parse_csv(path)
    if suffix == ".xlsx":
        return parse_xlsx(path)
    raise SyncError(f"Unsupported attachment type: {path.name}")


def clean_text(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def parse_int(value: Any) -> int | None:
    text = clean_text(value).replace(",", "")
    if not text:
        return None
    return int(float(text))


def parse_float(value: Any) -> float | None:
    text = clean_text(value).replace(",", "")
    if not text:
        return None
    return float(text)


def parse_date(value: Any) -> str | None:
    text = clean_text(value)
    if not text:
        return None
    if re.fullmatch(r"\d{4}-\d{2}-\d{2}", text):
        return f"{text} 00:00:00"
    return text


def normalize_rows(raw_rows: list[dict[str, Any]]) -> dict[str, dict[str, Any]]:
    deduped: dict[str, dict[str, Any]] = {}
    for row in raw_rows:
        key = clean_text(row.get("主账号用户ID")) or clean_text(row.get("租户ID")) or clean_text(row.get("联系邮箱"))
        if not key:
            continue
        normalized = {
            "排名": parse_int(row.get("排名")),
            "租户ID": clean_text(row.get("租户ID")),
            "主账号用户ID": clean_text(row.get("主账号用户ID")),
            "用户名": clean_text(row.get("用户名")),
            "联系邮箱": clean_text(row.get("联系邮箱")),
            "消费金额": parse_float(row.get("消费金额")),
            "消费模型(Top5/按消费)": clean_text(row.get("消费模型(Top5/按消费)")),
            "最近消费": parse_date(row.get("最近消费")),
            "活跃天数": parse_int(row.get("活跃天数")),
        }
        old = deduped.get(key)
        if old is None:
            deduped[key] = normalized
            continue
        old_rank = old.get("排名") or 10**9
        new_rank = normalized.get("排名") or 10**9
        if new_rank < old_rank:
            deduped[key] = normalized
    return deduped


def select_value(value: Any) -> str:
    if isinstance(value, list):
        return str(value[0]) if value else ""
    if value is None:
        return ""
    return str(value)


def fetch_existing_records(base_token: str, table_id: str) -> dict[str, ExistingRecord]:
    records: dict[str, ExistingRecord] = {}
    offset = 0
    while True:
        cmd = [
            "base",
            "+record-list",
            "--base-token",
            base_token,
            "--table-id",
            table_id,
            "--offset",
            str(offset),
            "--limit",
            "200",
            "--format",
            "json",
        ]
        for field in BASE_READ_FIELDS:
            cmd.extend(["--field-id", field])
        payload = lark_cli(cmd, timeout=180)
        data = payload.get("data") or payload
        field_names = data.get("fields") or BASE_READ_FIELDS
        rows = data.get("data") or []
        record_ids = data.get("record_id_list") or []
        for record_id, values in zip(record_ids, rows):
            field_map = dict(zip(field_names, values))
            key = clean_text(field_map.get("主账号用户ID")) or clean_text(field_map.get("租户ID")) or clean_text(
                field_map.get("联系邮箱"),
            )
            if key:
                records[key] = ExistingRecord(record_id=str(record_id), key=key, fields=field_map)
        if not data.get("has_more"):
            break
        offset += len(rows)
        if not rows:
            break
    return records


def record_payload(row: dict[str, Any], label: str, *, create: bool) -> dict[str, Any]:
    payload = {field: row.get(field) for field in REPORT_FIELDS if row.get(field) is not None}
    payload["标签"] = label
    if create:
        payload["触达状态"] = "待触达"
    return payload


def chunked(values: list[Any], size: int) -> list[list[Any]]:
    return [values[index : index + size] for index in range(0, len(values), size)]


def create_records(base_token: str, table_id: str, rows: list[dict[str, Any]], *, dry_run: bool) -> None:
    if not rows:
        return
    fields = REPORT_FIELDS + ["标签", "触达状态"]
    for batch in chunked(rows, 200):
        body = {"fields": fields, "rows": [[row.get(field) for field in fields] for row in batch]}
        if dry_run:
            continue
        lark_cli(
            [
                "base",
                "+record-batch-create",
                "--base-token",
                base_token,
                "--table-id",
                table_id,
                "--json",
                json.dumps(body, ensure_ascii=False, separators=(",", ":")),
            ],
            timeout=240,
        )


def update_record(base_token: str, table_id: str, record_id: str, patch: dict[str, Any], *, dry_run: bool) -> None:
    if dry_run:
        return
    lark_cli(
        [
            "base",
            "+record-upsert",
            "--base-token",
            base_token,
            "--table-id",
            table_id,
            "--record-id",
            record_id,
            "--json",
            json.dumps(patch, ensure_ascii=False, separators=(",", ":")),
        ],
        timeout=180,
    )


def mark_dropped(base_token: str, table_id: str, record_ids: list[str], *, dry_run: bool) -> None:
    if not record_ids:
        return
    for batch in chunked(record_ids, 200):
        if dry_run:
            continue
        lark_cli(
            [
                "base",
                "+record-batch-update",
                "--base-token",
                base_token,
                "--table-id",
                table_id,
                "--json",
                json.dumps(
                    {"record_id_list": batch, "patch": {"标签": DROPPED_LABEL}},
                    ensure_ascii=False,
                    separators=(",", ":"),
                ),
            ],
            timeout=180,
        )


def select_report_attachments(message: dict[str, Any]) -> list[dict[str, Any]]:
    attachments = message.get("attachments") or []
    selected = []
    for item in attachments:
        filename = str(item.get("filename") or "")
        content_type = str(item.get("content_type") or "")
        if item.get("is_inline"):
            continue
        if filename.lower().endswith((".csv", ".xlsx")) or "text/csv" in content_type:
            selected.append(item)
    return selected


def process_message(args: argparse.Namespace, message: MailMessage) -> dict[str, Any]:
    full_message = read_message(message.message_id, args.mailbox)
    attachments = select_report_attachments(full_message)
    if not attachments:
        raise SyncError(f"No CSV/XLSX report attachment found in message {message.message_id}")

    urls = attachment_download_urls(
        message.message_id,
        args.mailbox,
        [str(item["id"]) for item in attachments if item.get("id")],
    )
    output_dir = Path(args.download_dir).expanduser()
    raw_rows: list[dict[str, Any]] = []
    downloaded: list[str] = []
    for item in attachments:
        attachment_id = str(item.get("id") or "")
        url = urls.get(attachment_id)
        if not url:
            raise SyncError(f"Missing download URL for attachment {attachment_id}")
        path = download_attachment(url, str(item.get("filename") or attachment_id), output_dir)
        downloaded.append(str(path))
        raw_rows.extend(parse_attachment(path))

    incoming = normalize_rows(raw_rows)
    existing = fetch_existing_records(args.base_token, args.table_id)

    creates: list[dict[str, Any]] = []
    updates: list[tuple[str, dict[str, Any]]] = []
    incoming_keys = set(incoming)

    for key, row in incoming.items():
        record = existing.get(key)
        if record is None:
            creates.append(record_payload(row, NEW_LABEL, create=True))
        else:
            label = ACTIVE_LABEL if record.was_in_rank else NEW_LABEL
            updates.append((record.record_id, record_payload(row, label, create=False)))

    dropped = [
        record.record_id
        for key, record in existing.items()
        if record.was_in_rank and key not in incoming_keys
    ]

    create_records(args.base_token, args.table_id, creates, dry_run=args.dry_run)
    for record_id, patch in updates:
        update_record(args.base_token, args.table_id, record_id, patch, dry_run=args.dry_run)
        if args.write_delay > 0:
            time.sleep(args.write_delay)
    mark_dropped(args.base_token, args.table_id, dropped, dry_run=args.dry_run)

    return {
        "message_id": message.message_id,
        "subject": full_message.get("subject") or message.subject,
        "downloaded": downloaded,
        "incoming_rows": len(raw_rows),
        "deduped_users": len(incoming),
        "created": len(creates),
        "updated": len(updates),
        "dropped": len(dropped),
        "dry_run": args.dry_run,
    }


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Sync SiliconFlow Top 30 report emails into Feishu Base.")
    parser.add_argument("--base-token", default=os.getenv("TOP30_BASE_TOKEN", DEFAULT_BASE_TOKEN))
    parser.add_argument("--table-id", default=os.getenv("TOP30_TABLE_ID", DEFAULT_TABLE_ID))
    parser.add_argument("--mailbox", default=os.getenv("TOP30_MAILBOX", "me"))
    parser.add_argument("--query", default=os.getenv("TOP30_MAIL_QUERY", "Top 30 用户"))
    parser.add_argument("--from-contains", default=os.getenv("TOP30_FROM_CONTAINS", "no-reply@siliconflow.cn"))
    parser.add_argument(
        "--subject-contains",
        action="append",
        default=None,
        help="Required subject substring. Repeat for multiple checks.",
    )
    parser.add_argument("--max-messages", type=int, default=int(os.getenv("TOP30_MAX_MESSAGES", "50")))
    parser.add_argument("--message-id", default=None, help="Process one explicit mail message_id.")
    parser.add_argument("--force", action="store_true", help="Process messages even if state marks them processed.")
    parser.add_argument("--dry-run", action="store_true", help="Parse and compare without writing Base or state.")
    parser.add_argument(
        "--state-path",
        default=os.getenv("TOP30_STATE_PATH", str(DEFAULT_STATE_PATH)),
        help="Processed-message state JSON path.",
    )
    parser.add_argument(
        "--download-dir",
        default=os.getenv(
            "TOP30_DOWNLOAD_DIR",
            str(Path.home() / "Downloads" / "siliconflow_top30_base_sync"),
        ),
    )
    parser.add_argument("--write-delay", type=float, default=float(os.getenv("TOP30_WRITE_DELAY", "0.1")))
    return parser.parse_args()


def main() -> int:
    args = parse_args()
    if args.subject_contains is None:
        args.subject_contains = ["Top 30 用户", "有结果"]

    state_path = Path(args.state_path).expanduser()
    state = load_state(state_path)
    processed = set(state.get("processed_message_ids") or [])
    messages = list_candidate_messages(args, processed)
    if not messages:
        print(json.dumps({"processed": 0, "message": "No unprocessed matching report emails found."}, ensure_ascii=False))
        return 0

    summaries = []
    for message in messages:
        summary = process_message(args, message)
        summaries.append(summary)
        if not args.dry_run:
            processed.add(message.message_id)

    if not args.dry_run:
        state["processed_message_ids"] = sorted(processed)
        state["last_run_at"] = datetime.now().isoformat(timespec="seconds")
        state["last_summary"] = summaries[-1] if summaries else None
        save_state(state_path, state)

    print(json.dumps({"processed": len(summaries), "summaries": summaries}, ensure_ascii=False, indent=2))
    return 0


if __name__ == "__main__":
    try:
        raise SystemExit(main())
    except SyncError as exc:
        print(f"ERROR: {exc}", file=sys.stderr)
        raise SystemExit(1)
