"""Unit tests for pipeline utility and analysis functions."""

from __future__ import annotations

from datetime import date, datetime, timezone

from openpyxl import load_workbook

from pipeline.utils import normalize_for_match, parse_compact_number
from pipeline.model_ranking_weekly import (
    build_metadata,
    build_ranking_rows,
    get_previous_complete_week_start,
    iso_week_label,
    output_path_for_week,
    save_rankings_workbook,
)
from pipeline.model_provider_price_uptime import (
    MonitoredModel,
    build_rows_from_api_and_page,
    format_compact_count,
    load_monitored_models,
    model_url_from_id,
    parse_model_slug,
    price_per_million,
    workbook_path_for_model,
    write_weekly_workbook,
)
from pipeline.core_providers import load_core_providers
from pipeline.workbook_metadata import METADATA_SHEET, resolve_updated_at
from pipeline.model_provider_usage import (
    CORE_COVERAGE_THRESHOLD,
    MISSING_DISPLAY_TEXT,
    SKIPPED_LOW_USAGE_TEXT,
    _chart_model_matches,
    build_provider_slug_map,
    build_skipped_low_usage_row,
    build_usage_row,
    core_coverage_sufficient,
    core_usage_coverage_ratio,
    partition_providers_by_core,
    provider_in_core_list,
    read_price_uptime_providers,
    resolve_provider_slug,
    update_price_uptime_workbook_with_usage,
    week_dates,
)
from scripts.model_provider_daily import iso_week_start, run_daily_pipeline, utc_yesterday


class TestAIModelRankingsExcel:
    def test_week_label_uses_data_week(self):
        week_start = date(2026, 5, 4)
        assert iso_week_label(week_start) == "2026-W19"
        assert build_metadata(week_start, updated_at=date(2026, 5, 11))[0] == (
            "数据范围",
            "2026/5/4 ～ 2026/5/10",
        )

    def test_default_week_is_previous_complete_iso_week(self):
        assert get_previous_complete_week_start(date(2026, 5, 17)) == date(2026, 5, 4)

    def test_build_ranking_rows_converts_usage_and_pct(self):
        payload = {
            "models": [
                {"model_name": "Hy3 preview", "tokens_text": "2.68T", "tokens": 2_680_000_000_000},
                {"model_name": "DeepSeek V3.2", "tokens_text": "868B", "tokens": 868_000_000_000},
                {"model_name": "Others", "tokens_text": "14T", "tokens": 14_000_000_000_000},
            ],
            "total_tokens_text": "25.7T",
            "total_tokens": 25_700_000_000_000,
        }
        rows = build_ranking_rows(payload)

        assert rows[0] == {
            "模型名称": "Hy3 preview",
            "原始用量": "2.68T",
            "换算后用量（T）": 2.68,
            "占比（%）": 0.1043,
        }
        assert rows[1]["模型名称"] == "DeepSeek V3.2"
        assert rows[1]["换算后用量（T）"] == 0.868
        assert rows[-2]["模型名称"] == "Others"
        assert rows[-1] == {
            "模型名称": "Total",
            "原始用量": "25.7T",
            "换算后用量（T）": 25.7,
            "占比（%）": 1,
        }

    def test_save_workbook_has_required_sheets_and_output_path(self, tmp_path):
        week_start = date(2026, 5, 4)
        rows = build_ranking_rows({
            "models": [
                {"model_name": "Hy3 preview", "tokens_text": "2.68T tokens", "tokens": 2_680_000_000_000},
                {"model_name": "Others", "tokens_text": "14T tokens", "tokens": 14_000_000_000_000},
            ],
            "total_tokens_text": "16.68T tokens",
            "total_tokens": 16_680_000_000_000,
        })

        path = save_rankings_workbook(
            rows,
            week_start,
            output_dir=tmp_path / "Ranking",
            updated_at=date(2026, 5, 11),
        )

        assert path == tmp_path / "Ranking" / "AI Model Rankings 2026-W19.xlsx"
        assert output_path_for_week(week_start, output_dir=tmp_path / "Ranking") == path

        wb = load_workbook(path, data_only=True)
        assert wb.sheetnames == ["Ranking", "Metadata"]
        ranking = wb["Ranking"]
        assert [ranking.cell(1, col).value for col in range(1, 5)] == [
            "模型名称",
            "原始用量",
            "换算后用量（T）",
            "占比（%）",
        ]
        assert ranking.cell(2, 1).value == "Hy3 preview"
        assert ranking.cell(2, 2).value == "2.68T"

        metadata = wb["Metadata"]
        assert metadata.cell(2, 1).value == "数据范围"
        assert metadata.cell(2, 2).value == "2026/5/4 ～ 2026/5/10"
        assert metadata.cell(4, 2).value == "https://openrouter.ai/rankings"


class TestModelProviderPriceUptime:
    def test_parse_model_slug_from_id_and_url(self):
        assert parse_model_slug(model_id="deepseek/deepseek-v4-flash") == "deepseek-v4-flash"
        assert parse_model_slug(model_url="https://openrouter.ai/deepseek/deepseek-v4-flash") == "deepseek-v4-flash"

    def test_model_url_from_id(self):
        assert model_url_from_id("deepseek/deepseek-v4-flash") == "https://openrouter.ai/deepseek/deepseek-v4-flash"

    def test_load_monitored_models_from_id_only_config(self, tmp_path):
        config_path = tmp_path / "core_models.json"
        config_path.write_text('["deepseek/deepseek-v4-flash"]', encoding="utf-8")
        models = load_monitored_models(config_path)
        assert models == [
            MonitoredModel(
                model_id="deepseek/deepseek-v4-flash",
                model_slug="deepseek-v4-flash",
                model_url="https://openrouter.ai/deepseek/deepseek-v4-flash",
            )
        ]

    def test_price_per_million(self):
        assert price_per_million("0.000000112") == 0.112
        assert price_per_million("0.0000000028") == 0.0028
        assert price_per_million(None) is None

    def test_format_compact_count(self):
        assert format_compact_count(1_048_576) == "1.05M"
        assert format_compact_count(393_216) == "393.2K"
        assert format_compact_count(384_000) == "384K"
        assert format_compact_count(None) == ""

    def test_workbook_path_for_model(self, tmp_path):
        model = MonitoredModel(
            model_id="deepseek/deepseek-v4-flash",
            model_slug="deepseek-v4-flash",
            model_url="https://openrouter.ai/deepseek/deepseek-v4-flash",
        )
        path = workbook_path_for_model(model, date(2026, 5, 18), output_dir=tmp_path / "Price&Uptime&Usage")
        assert path == tmp_path / "Price&Uptime&Usage" / "deepseek-v4-flash" / "deepseek-v4-flash 2026-W21.xlsx"

    def test_build_rows_from_api_and_page(self):
        endpoints = [{
            "provider_name": "GMICloud",
            "tag": "gmicloud/fp8",
            "quantization": "fp8",
            "context_length": 1_048_575,
            "max_completion_tokens": None,
            "pricing": {
                "prompt": "0.000000112",
                "completion": "0.000000224",
                "input_cache_read": "0.000000022",
            },
            "uptime_last_1d": 99.0192244905798,
        }]
        page_cards = [{
            "Provider": "GMICloud",
            "Region": "US",
            "Quantization": "fp8",
            "Latency": "2.38s",
            "Throughput": "56tps",
            "Total Context": "1.05M",
            "Max Output": "1.05M",
        }]
        rows = build_rows_from_api_and_page(endpoints, page_cards)
        assert rows == [{
            "Provider": "GMICloud",
            "Region": "US",
            "Quantization": "fp8",
            "Latency": "2.38s",
            "Throughput": "56tps",
            "Uptime": 99.0,
            "Total Context": "1.05M",
            "Max Output": "1.05M",
            "Input Price": 0.112,
            "Output Price": 0.224,
            "Cache Read": 0.022,
        }]

    def test_daily_tab_overwrites_existing_tab(self, tmp_path):
        model = MonitoredModel(
            model_id="deepseek/deepseek-v4-flash",
            model_slug="deepseek-v4-flash",
            model_url="https://openrouter.ai/deepseek/deepseek-v4-flash",
        )
        week_start = date(2026, 5, 18)
        snapshot = date(2026, 5, 18)

        first_path = write_weekly_workbook(
            model,
            week_start,
            snapshot,
            [{"Provider": "Old", "Uptime": 1.0}],
            output_dir=tmp_path / "Price&Uptime&Usage",
        )
        second_path = write_weekly_workbook(
            model,
            week_start,
            snapshot,
            [{"Provider": "New", "Uptime": 99.9}],
            output_dir=tmp_path / "Price&Uptime&Usage",
        )

        assert first_path == second_path
        wb = load_workbook(second_path, data_only=True)
        assert wb.sheetnames == ["2026-05-18", METADATA_SHEET]
        ws = wb["2026-05-18"]
        assert ws.cell(2, 1).value == "New"
        assert ws.max_row == 2


class TestModelProviderDaily:
    def test_utc_yesterday_for_regular_day(self):
        assert utc_yesterday(datetime(2026, 5, 20, 0, 0, tzinfo=timezone.utc)) == date(2026, 5, 19)

    def test_utc_yesterday_for_monday_backfills_sunday(self):
        target_date = utc_yesterday(datetime(2026, 5, 25, 0, 0, tzinfo=timezone.utc))
        assert target_date == date(2026, 5, 24)
        assert iso_week_start(target_date) == date(2026, 5, 18)

    def test_iso_week_start_handles_year_boundary(self):
        target_date = utc_yesterday(datetime(2027, 1, 1, 0, 0, tzinfo=timezone.utc))
        assert target_date == date(2026, 12, 31)
        assert iso_week_start(target_date) == date(2026, 12, 28)

    def test_run_daily_pipeline_skips_income_when_input_missing(self, tmp_path, monkeypatch):
        week_start = date(2026, 5, 18)
        target_date = date(2026, 5, 20)
        calls: list[str] = []

        monkeypatch.setattr(
            "scripts.model_provider_daily.generate_price_uptime_workbooks",
            lambda ws, d: calls.append("price") or [tmp_path / "p.xlsx"],
        )
        monkeypatch.setattr(
            "scripts.model_provider_daily.generate_usage_workbooks",
            lambda ws, days: calls.append("usage") or [tmp_path / "u.xlsx"],
        )
        monkeypatch.setattr(
            "scripts.model_provider_daily.generate_core_models_provider_excel",
            lambda ws: calls.append("provider") or tmp_path / "provider.xlsx",
        )
        monkeypatch.setattr(
            "scripts.model_provider_daily.generate_core_models_usage_excel",
            lambda ws: calls.append("core_usage") or tmp_path / "usage.xlsx",
        )
        monkeypatch.setattr(
            "scripts.model_provider_daily.generate_core_models_income_excel",
            lambda *a, **k: calls.append("income") or tmp_path / "income.xlsx",
        )
        monkeypatch.setattr(
            "scripts.model_provider_daily.generate_core_models_dashboard",
            lambda ws: calls.append("dashboard") or tmp_path / "dash.html",
        )
        monkeypatch.setattr(
            "scripts.model_provider_daily.default_income_path",
            lambda ws, input_dir=None: tmp_path / "missing_income.xlsx",
        )
        out_income = tmp_path / "Core_Models" / "Core Model Income 2026-W21.xlsx"
        out_income.parent.mkdir(parents=True, exist_ok=True)
        out_income.write_text("placeholder")
        monkeypatch.setattr(
            "scripts.model_provider_daily.income_output_path_for_week",
            lambda ws, output_dir=None: out_income,
        )

        assert run_daily_pipeline(target_date) == 0
        assert calls == ["core_usage", "price", "usage", "provider", "dashboard"]
        assert "income" not in calls

    def test_run_daily_pipeline_runs_income_when_input_exists(self, tmp_path, monkeypatch):
        week_start = date(2026, 5, 18)
        target_date = date(2026, 5, 20)
        income_input = tmp_path / "model_income_W21.xlsx"
        income_input.write_text("x")
        calls: list[str] = []

        monkeypatch.setattr(
            "scripts.model_provider_daily.generate_price_uptime_workbooks",
            lambda ws, d: calls.append("price") or [],
        )
        monkeypatch.setattr(
            "scripts.model_provider_daily.generate_usage_workbooks",
            lambda ws, days: calls.append("usage") or [],
        )
        monkeypatch.setattr(
            "scripts.model_provider_daily.generate_core_models_provider_excel",
            lambda ws: calls.append("provider") or tmp_path / "p.xlsx",
        )
        monkeypatch.setattr(
            "scripts.model_provider_daily.generate_core_models_usage_excel",
            lambda ws: calls.append("core_usage") or tmp_path / "u.xlsx",
        )
        monkeypatch.setattr(
            "scripts.model_provider_daily.generate_core_models_income_excel",
            lambda *a, **k: calls.append("income") or tmp_path / "i.xlsx",
        )
        monkeypatch.setattr(
            "scripts.model_provider_daily.generate_core_models_dashboard",
            lambda ws: calls.append("dashboard") or tmp_path / "d.html",
        )
        monkeypatch.setattr(
            "scripts.model_provider_daily.default_income_path",
            lambda ws, input_dir=None: income_input,
        )
        monkeypatch.setattr(
            "scripts.model_provider_daily.income_output_path_for_week",
            lambda ws, output_dir=None: tmp_path / "out_income.xlsx",
        )

        assert run_daily_pipeline(target_date) == 0
        assert calls == ["core_usage", "price", "usage", "provider", "income", "dashboard"]


class TestModelProviderUsage:
    def test_week_dates_generates_seven_days(self):
        assert week_dates(date(2026, 5, 18)) == [
            date(2026, 5, 18),
            date(2026, 5, 19),
            date(2026, 5, 20),
            date(2026, 5, 21),
            date(2026, 5, 22),
            date(2026, 5, 23),
            date(2026, 5, 24),
        ]

    def test_read_price_uptime_providers(self, tmp_path):
        model = MonitoredModel(
            model_id="deepseek/deepseek-v4-flash",
            model_slug="deepseek-v4-flash",
            model_url="https://openrouter.ai/deepseek/deepseek-v4-flash",
        )
        write_weekly_workbook(
            model,
            date(2026, 5, 18),
            date(2026, 5, 18),
            [{"Provider": "GMICloud"}, {"Provider": "Baidu Qianfan"}, {"Provider": "GMICloud"}],
            output_dir=tmp_path / "Price&Uptime&Usage",
        )
        assert read_price_uptime_providers(
            model,
            date(2026, 5, 18),
            date(2026, 5, 18),
            price_uptime_output_dir=tmp_path / "Price&Uptime&Usage",
        ) == ["GMICloud", "Baidu Qianfan"]

    def test_provider_name_maps_to_slug(self):
        endpoints = [
            {"provider_name": "GMICloud", "tag": "gmicloud/fp8"},
            {"provider_name": "Baidu", "tag": "baidu/fp8"},
            {"provider_name": "AtlasCloud", "tag": "atlas-cloud/fp8"},
        ]
        slug_map = build_provider_slug_map(endpoints)
        assert resolve_provider_slug("GMICloud", slug_map) == "gmicloud"
        assert resolve_provider_slug("Baidu Qianfan", slug_map) == "baidu"
        assert resolve_provider_slug("AtlasCloud", slug_map) == "atlas-cloud"

    def test_build_usage_row_when_model_displayed(self):
        payload = {
            "total_tokens_text": "77.9B",
            "models": [
                {"model_name": "DeepSeek V4 Flash", "tokens_text": "72.7B", "tokens": 72_700_000_000},
                {"model_name": "DeepSeek V4 Pro", "tokens_text": "3.24B", "tokens": 3_240_000_000},
            ],
        }
        row = build_usage_row(
            provider_name="GMICloud",
            provider_slug="gmicloud",
            chart_payload=payload,
            model_display_name="DeepSeek V4 Flash",
            model_daily_total=393_665_947_498,
        )
        assert row["Provider"] == "GMICloud"
        assert row["Provider URL"] == "https://openrouter.ai/provider/gmicloud"
        assert row["模型每日总量"] == 393_665_947_498
        assert row["模型用量"] == 72_700_000_000
        assert row["Provider 承接用量"] == "72.7B"
        assert row["承接占比"] == 72_700_000_000 / 393_665_947_498
        assert row["展示状态"] == "已展示"
        assert row["Provider当日总量"] == "77.9B"

    def test_chart_model_matches_vendor_prefixed_display_name(self):
        assert _chart_model_matches("DeepSeek V4 Flash", "DeepSeek: DeepSeek V4 Flash")

    def test_build_usage_row_when_model_not_displayed(self):
        row = build_usage_row(
            provider_name="SmallProvider",
            provider_slug="small",
            chart_payload={"total_tokens_text": "10B", "models": []},
            model_display_name="DeepSeek V4 Flash",
            model_daily_total=393_665_947_498,
        )
        assert row["模型用量"] is None
        assert row["承接占比"] is None
        assert row["展示状态"] == MISSING_DISPLAY_TEXT
        assert row["Provider当日总量"] == "10B"

    def test_load_core_providers(self):
        providers = load_core_providers()
        slugs = {entry["provider_slug"] for entry in providers}
        assert "novita" in slugs
        assert "deepseek" in slugs
        assert "streamlake" in slugs

    def test_provider_in_core_list_by_slug_and_name(self):
        endpoints = [
            {"provider_name": "GMICloud", "tag": "gmicloud/fp8"},
            {"provider_name": "Parasail", "tag": "parasail/fp8"},
        ]
        slug_map = build_provider_slug_map(endpoints)
        core = load_core_providers()
        assert provider_in_core_list("GMICloud", slug_map, core)
        assert not provider_in_core_list("Parasail", slug_map, core)

    def test_partition_providers_by_core(self):
        endpoints = [
            {"provider_name": "GMICloud", "tag": "gmicloud/fp8"},
            {"provider_name": "Parasail", "tag": "parasail/fp8"},
            {"provider_name": "DeepSeek", "tag": "deepseek/fp8"},
        ]
        slug_map = build_provider_slug_map(endpoints)
        core, non_core = partition_providers_by_core(
            ["Parasail", "GMICloud", "DeepSeek"],
            slug_map,
            load_core_providers(),
        )
        assert core == ["GMICloud", "DeepSeek"]
        assert non_core == ["Parasail"]

    def test_core_coverage_sufficient_at_threshold(self):
        model_total = 1000
        rows = [
            {"模型用量": int(model_total * CORE_COVERAGE_THRESHOLD)},
            {"模型用量": None},
        ]
        assert core_usage_coverage_ratio(rows, model_total) == CORE_COVERAGE_THRESHOLD
        assert core_coverage_sufficient(rows, model_total)

    def test_core_coverage_insufficient_triggers_more_scraping(self):
        rows = [{"模型用量": 100}]
        assert not core_coverage_sufficient(rows, 1000)

    def test_build_skipped_low_usage_row(self):
        row = build_skipped_low_usage_row(
            "Parasail",
            "parasail",
            1_000_000,
            date(2026, 5, 20),
        )
        assert row["展示状态"] == SKIPPED_LOW_USAGE_TEXT
        assert row["Provider URL"] == "https://openrouter.ai/provider/parasail"
        assert row["模型用量"] is None
        assert row["Provider 承接用量"] == ""

    def test_updates_price_uptime_workbook_with_usage_columns(self, tmp_path):
        model = MonitoredModel(
            model_id="deepseek/deepseek-v4-flash",
            model_slug="deepseek-v4-flash",
            model_url="https://openrouter.ai/deepseek/deepseek-v4-flash",
        )
        week_start = date(2026, 5, 18)
        snapshot = date(2026, 5, 18)
        write_weekly_workbook(
            model,
            week_start,
            snapshot,
            [{
                "Provider": "GMICloud",
                "Region": "US",
                "Quantization": "fp8",
                "Latency": "2.38s",
                "Throughput": "56tps",
                "Uptime": 99.0,
                "Total Context": "1.05M",
                "Max Output": "1.05M",
                "Input Price": 0.112,
                "Output Price": 0.224,
                "Cache Read": 0.022,
            }],
            output_dir=tmp_path / "Price&Uptime&Usage",
        )

        path = update_price_uptime_workbook_with_usage(
            model,
            week_start,
            {
                snapshot: [{
                    "Provider": "GMICloud",
                    "Provider URL": "https://openrouter.ai/provider/gmicloud",
                    "展示状态": "已展示",
                    "Provider 承接用量": "72.7B",
                    "Provider当日总量": "77.9B",
                    "承接占比": 0.1847,
                }]
            },
            price_uptime_output_dir=tmp_path / "Price&Uptime&Usage",
        )

        wb = load_workbook(path, data_only=True)
        ws = wb["2026-05-18"]
        assert [ws.cell(1, col).value for col in range(1, ws.max_column + 1)] == [
            "Provider",
            "Provider URL",
            "Region",
            "Quantization",
            "Latency",
            "Throughput",
            "Uptime",
            "Total Context",
            "Max Output",
            "Input Price",
            "Output Price",
            "Cache Read",
            "展示状态",
            "Provider 承接用量",
            "Provider当日总量",
            "承接占比",
        ]
        assert ws.cell(2, 2).value == "https://openrouter.ai/provider/gmicloud"
        assert ws.cell(2, 13).value == "已展示"
        assert ws.cell(2, 14).value == "72.7B"
        assert ws.cell(2, 15).value == "77.9B"
        assert ws.cell(2, 16).value == 0.1847
        assert METADATA_SHEET in wb.sheetnames
        meta = {wb[METADATA_SHEET].cell(row, 1).value: wb[METADATA_SHEET].cell(row, 2).value for row in range(2, 20) if wb[METADATA_SHEET].cell(row, 1).value}
        assert meta["数据更新时间"] == "2026/5/18"


class TestWorkbookMetadata:
    def test_resolve_updated_at_prefers_explicit(self):
        assert resolve_updated_at(date(2026, 5, 20), snapshot_dates=[date(2026, 5, 18)]) == date(2026, 5, 20)

    def test_resolve_updated_at_uses_max_snapshot(self):
        assert resolve_updated_at(None, snapshot_dates=[date(2026, 5, 18), date(2026, 5, 20)]) == date(2026, 5, 20)

    def test_build_metadata_uses_week_start_not_plus_seven(self):
        week_start = date(2026, 5, 18)
        rows = build_metadata(week_start)
        updated = next(value for key, value in rows if key == "数据更新时间")
        assert updated == "2026/5/18"
        assert updated != "2026/5/25"

    def test_price_uptime_workbook_metadata_after_write(self, tmp_path):
        from pipeline.model_provider_price_uptime import MonitoredModel, write_weekly_workbook

        model = MonitoredModel(
            model_id="deepseek/deepseek-v4-flash",
            model_slug="deepseek-v4-flash",
            model_url="https://openrouter.ai/deepseek/deepseek-v4-flash",
        )
        week_start = date(2026, 5, 18)
        write_weekly_workbook(
            model,
            week_start,
            date(2026, 5, 18),
            [{"Provider": "GMICloud", "Region": "US", "Quantization": "fp8", "Latency": "2.38s", "Throughput": "56tps", "Uptime": 99.0, "Total Context": "1.05M", "Max Output": "1.05M", "Input Price": 0.112, "Output Price": 0.224, "Cache Read": 0.022}],
            output_dir=tmp_path / "Price&Uptime&Usage",
        )
        path = write_weekly_workbook(
            model,
            week_start,
            date(2026, 5, 20),
            [{"Provider": "GMICloud", "Region": "US", "Quantization": "fp8", "Latency": "2.38s", "Throughput": "56tps", "Uptime": 99.0, "Total Context": "1.05M", "Max Output": "1.05M", "Input Price": 0.112, "Output Price": 0.224, "Cache Read": 0.022}],
            output_dir=tmp_path / "Price&Uptime&Usage",
        )
        wb = load_workbook(path, data_only=True)
        assert METADATA_SHEET in wb.sheetnames
        meta = {wb[METADATA_SHEET].cell(row, 1).value: wb[METADATA_SHEET].cell(row, 2).value for row in range(2, 20) if wb[METADATA_SHEET].cell(row, 1).value}
        assert meta["数据更新时间"] == "2026/5/20"
        assert "2026-05-18" in meta["已写入日期"]
        assert "2026-05-20" in meta["已写入日期"]


class TestParseCompactNumber:
    def test_trillion(self):
        assert parse_compact_number("1.5T") == 1_500_000_000_000

    def test_billion(self):
        assert parse_compact_number("850B") == 850_000_000_000

    def test_million(self):
        assert parse_compact_number("12.3M") == 12_300_000

    def test_thousand(self):
        assert parse_compact_number("500K") == 500_000

    def test_plain_number(self):
        assert parse_compact_number("42") == 42

    def test_comma_separated(self):
        assert parse_compact_number("1,234.5") == 1234.5

    def test_none(self):
        assert parse_compact_number(None) is None

    def test_empty(self):
        assert parse_compact_number("") is None

    def test_with_suffix_text(self):
        assert parse_compact_number("1.2T tokens") == 1_200_000_000_000

    def test_negative(self):
        assert parse_compact_number("-5.2B") == -5_200_000_000


class TestNormalizeForMatch:
    def test_basic(self):
        assert normalize_for_match("DeepSeek V3.2") == "deepseekv32"

    def test_with_slash(self):
        assert normalize_for_match("deepseek/deepseek-v3.2") == "deepseekdeepseekv32"

    def test_uppercase(self):
        assert normalize_for_match("Kimi K2.6") == "kimik26"

    def test_empty(self):
        assert normalize_for_match("") == ""

class TestCoreModelsUsageExcel:
    def test_output_path(self, tmp_path):
        week_start = date(2026, 5, 18)
        from pipeline.core_models_usage import output_path_for_week
        path = output_path_for_week(week_start, output_dir=tmp_path / "Core_Models")
        assert path == tmp_path / "Core_Models" / "Core Model Usage 2026-W21.xlsx"

    def test_build_usage_row_week_sum_and_share(self):
        from pipeline.core_models_usage import build_usage_row
        from pipeline.model_provider_price_uptime import MonitoredModel

        week_start = date(2026, 5, 18)
        days = [week_start + __import__("datetime").timedelta(days=i) for i in range(7)]
        model = MonitoredModel(
            model_id="z-ai/glm-5.1",
            model_slug="glm-5.1",
            model_url="https://openrouter.ai/z-ai/glm-5.1",
        )
        totals = {
            "2026-05-18": 10_000_000_000,
            "2026-05-19": 20_000_000_000,
        }
        row = build_usage_row(model, "GLM 5.1", days, totals, rankings_total=1_000_000_000_000)
        assert row["模型ID"] == "z-ai/glm-5.1"
        assert row["2026-05-18"] == "10B"
        assert row["周合计"] == "30B"
        assert row["换算后用量（T）"] == 0.03
        assert row["占比（%）"] == 0.03

    def test_save_workbook_sheets(self, tmp_path):
        from pipeline.core_models_usage import save_core_models_usage_workbook
        from pipeline.model_provider_price_uptime import MonitoredModel
        from openpyxl import load_workbook

        week_start = date(2026, 5, 18)
        days = [week_start + __import__("datetime").timedelta(days=i) for i in range(7)]
        model = MonitoredModel(
            model_id="deepseek/deepseek-v4-flash",
            model_slug="deepseek-v4-flash",
            model_url="https://openrouter.ai/deepseek/deepseek-v4-flash",
        )
        row = {
            "模型ID": model.model_id,
            "模型名称": "DeepSeek V4 Flash",
            **{day.isoformat(): "1B" for day in days},
            "周合计": "7B",
            "换算后用量（T）": 0.007,
            "占比（%）": 0.001,
        }
        path = save_core_models_usage_workbook(
            [row],
            week_start,
            days,
            rankings_total=7_000_000_000_000,
            rankings_total_text="7T",
            output_dir=tmp_path / "Core_Models",
            updated_at=date(2026, 5, 25),
        )
        wb = load_workbook(path, data_only=True)
        assert wb.sheetnames == ["Usage", "Metadata"]
        assert wb["Usage"].cell(1, 1).value == "模型ID"
        assert wb["Usage"].cell(2, 3).value == "1B"
        assert wb["Metadata"].cell(2, 1).value == "数据范围"


class TestCoreModelsIncome:
    def test_load_income_daily_sums_platform_total(self, tmp_path):
        from datetime import datetime
        from openpyxl import Workbook
        from pipeline.core_models_income import build_income_rows, load_income_daily
        from pipeline.model_provider_price_uptime import MonitoredModel

        income_path = tmp_path / "model_income_W20.xlsx"
        wb = Workbook()
        ws = wb.active
        ws.title = "查询结果"
        ws.append(["emitted_day", "model_name", "paid_usd"])
        ws.append([datetime(2026, 5, 11), "DeepSeek-V4-Pro", 100.0])
        ws.append([datetime(2026, 5, 11), "Other-Model", 50.0])
        ws.append([datetime(2026, 5, 12), "DeepSeek-V4-Pro", 200.0])
        wb.save(income_path)

        week_start = date(2026, 5, 11)
        daily_index, platform_total = load_income_daily(income_path, week_start)
        assert platform_total == 350.0
        assert daily_index[("2026-05-11", "DeepSeek-V4-Pro")] == 100.0

        model = MonitoredModel(
            model_id="deepseek/deepseek-v4-pro",
            model_slug="deepseek-v4-pro",
            model_url="https://openrouter.ai/deepseek/deepseek-v4-pro",
        )
        days = [week_start, date(2026, 5, 12), date(2026, 5, 13), date(2026, 5, 14), date(2026, 5, 15), date(2026, 5, 16), date(2026, 5, 17)]
        rows = build_income_rows([model], days, daily_index, platform_total)
        assert rows[0]["周合计"] == 300.0
        assert rows[0]["占比（%）"] == round(300.0 / 350.0, 4)

    def test_output_path_for_week(self, tmp_path):
        from pathlib import Path
        from pipeline.core_models_income import output_path_for_week

        week_start = date(2026, 5, 18)
        path = output_path_for_week(week_start, output_dir=tmp_path / "out")
        assert path == tmp_path / "out/Core Model Income 2026-W21.xlsx"

    def test_save_income_workbook(self, tmp_path):
        from pipeline.core_models_income import save_core_models_income_workbook
        from pipeline.model_provider_price_uptime import MonitoredModel

        week_start = date(2026, 5, 11)
        from datetime import timedelta

        days = [week_start + timedelta(i) for i in range(7)]
        model = MonitoredModel(
            model_id="deepseek/deepseek-v4-pro",
            model_slug="deepseek-v4-pro",
            model_url="https://openrouter.ai/deepseek/deepseek-v4-pro",
        )
        income_source = tmp_path / "income.xlsx"
        row = {
            "模型ID": model.model_id,
            "模型名称": "DeepSeek-V4-Pro",
            **{d.isoformat(): 10.0 for d in days},
            "周合计": 70.0,
            "占比（%）": 1.0,
        }
        out = save_core_models_income_workbook(
            [row],
            week_start,
            days,
            income_path=income_source,
            platform_week_total=70.0,
            output_dir=tmp_path / "Core_Models",
        )
        wb = load_workbook(out, data_only=True)
        assert out.name == "Core Model Income 2026-W20.xlsx"
        assert wb.sheetnames == ["Income", "Metadata"]
        assert wb["Income"].cell(1, 3).value == "2026-05-11"
        assert wb["Income"].cell(2, 1).value == "deepseek/deepseek-v4-pro"
        assert wb["Income"].cell(2, 10).value == 70.0


class TestCoreModelsProvider:
    def test_output_path_for_week(self, tmp_path):
        from pipeline.core_models_provider import output_path_for_week

        week_start = date(2026, 5, 18)
        path = output_path_for_week(week_start, output_dir=tmp_path / "Core_Models")
        assert path == tmp_path / "Core_Models/Core Model Provider 2026-W21.xlsx"

    def test_aggregate_provider_workbook(self, tmp_path):
        from openpyxl import Workbook
        from pipeline.core_models_provider import generate_core_models_provider_excel
        from pipeline.model_provider_price_uptime import MonitoredModel, write_weekly_workbook

        week_start = date(2026, 5, 18)
        model = MonitoredModel(
            model_id="deepseek/deepseek-v4-pro",
            model_slug="deepseek-v4-pro",
            model_url="https://openrouter.ai/deepseek/deepseek-v4-pro",
        )
        input_root = tmp_path / "Price&Uptime&Usage"
        write_weekly_workbook(
            model,
            week_start,
            date(2026, 5, 18),
            [{"Provider": "Acme", "Region": "us", "Quantization": "fp8",
              "Latency": "1s", "Throughput": "2/s", "Uptime": 99,
              "Total Context": "128K", "Max Output": "8K",
              "Input Price": 0.1, "Output Price": 0.2, "Cache Read": 0.01}],
            output_dir=input_root,
        )

        out = generate_core_models_provider_excel(
            week_start,
            models=[model],
            price_uptime_dir=input_root,
            output_dir=tmp_path / "Core_Models",
        )
        wb = load_workbook(out, data_only=True)
        assert out.name == "Core Model Provider 2026-W21.xlsx"
        assert wb.sheetnames == ["deepseek-v4-pro", "Metadata"]
        ws = wb["deepseek-v4-pro"]
        assert ws.cell(1, 1).value == "日期"
        assert ws.cell(2, 1).value == "2026-05-18"
        assert ws.cell(2, 2).value == "Acme"
        assert wb["Metadata"].cell(2, 1).value == "数据范围"


class TestCoreModelsDashboard:
    def test_generate_dashboard_html(self, tmp_path):
        from datetime import timedelta
        from pathlib import Path
        from pipeline.core_models_dashboard import generate_core_models_dashboard
        from pipeline.core_models_income import save_core_models_income_workbook
        from pipeline.core_models_provider import save_core_models_provider_workbook
        from pipeline.core_models_usage import save_core_models_usage_workbook
        from pipeline.model_provider_price_uptime import MonitoredModel

        week_start = date(2026, 5, 18)
        days = [week_start + timedelta(i) for i in range(7)]
        model = MonitoredModel(
            model_id="deepseek/deepseek-v4-pro",
            model_slug="deepseek-v4-pro",
            model_url="https://openrouter.ai/deepseek/deepseek-v4-pro",
        )
        out_dir = tmp_path / "Core_Models"
        usage_path = save_core_models_usage_workbook(
            [{
                "模型ID": model.model_id,
                "模型名称": "Pro",
                **{d.isoformat(): "1B" for d in days},
                "周合计": "7B",
                "换算后用量（T）": 0.007,
                "占比（%）": 0.1,
            }],
            week_start,
            days,
            rankings_total=70_000_000_000_000,
            output_dir=out_dir,
        )
        income_path = save_core_models_income_workbook(
            [{
                "模型ID": model.model_id,
                "模型名称": "DeepSeek-V4-Pro",
                **{d.isoformat(): 10.0 for d in days},
                "周合计": 70.0,
                "占比（%）": 1.0,
            }],
            week_start,
            days,
            tmp_path / "income.xlsx",
            70.0,
            output_dir=out_dir,
        )
        provider_path = save_core_models_provider_workbook(
            week_start,
            [(model, ["日期", "Provider", "展示状态", "Provider 承接用量"], [["2026-05-18", "Acme", "已展示", "2B"]])],
            [("数据周", "2026-W21")],
            output_dir=out_dir,
        )

        html_path = generate_core_models_dashboard(
            week_start,
            output_dir=out_dir,
            usage_path=usage_path,
            provider_path=provider_path,
            income_path=income_path,
        )
        text = html_path.read_text(encoding="utf-8")
        assert html_path.name == "Core Models Dashboard 2026-W21.html"
        assert 'src="chart.umd.min.js"' in text
        assert (tmp_path / "Core_Models" / "chart.umd.min.js").exists()
        assert 'id="usage"' in text
        assert 'id="provider"' in text
        assert 'id="income"' in text
        assert "cdn.tailwindcss.com" in text
        assert "核心结论" in text
        assert "关键指标" in text or "监控模型数" in text
        assert "洞察：" in text
        assert "doughnutShareChart" in text
        assert "sharePercentLabels" in text
        assert "generateLabels" in text
        assert "DeepSeek-V4-Pro" in text or "Pro" in text
        assert "Acme" in text
        assert "Top3 承接总结：" in text
        assert "价格洞察：" in text
        assert "gstatic.com/faviconV2" in text
        assert '"color":' in text
        assert "datasetBarColors" in text

    def test_provider_color_is_stable_per_name(self):
        from pipeline.core_models_dashboard import provider_color

        assert provider_color("DeepSeek") == provider_color("DeepSeek")
        assert provider_color("DeepSeek") != provider_color("SiliconFlow")

    def test_model_author_icon_url_uses_author_homepage(self):
        from pipeline.core_models_dashboard import model_author_icon_url

        url = model_author_icon_url("deepseek")
        assert "gstatic.com/faviconV2" in url
        assert "deepseek.com" in url

    def test_build_provider_top3_summary_lists_top_three(self):
        from pipeline.core_models_dashboard import (
            PROVIDER_NO_DATA_TEXT,
            _build_provider_top3_summary,
        )

        model = {
            "dict_rows": [
                {"展示状态": "已展示", "Provider": "Alpha", "Provider 承接用量": "50B"},
                {"展示状态": "已展示", "Provider": "Beta", "Provider 承接用量": "30B"},
                {"展示状态": "已展示", "Provider": "Gamma", "Provider 承接用量": "15B"},
                {"展示状态": "已展示", "Provider": "Delta", "Provider 承接用量": "5B"},
            ],
        }
        summary = _build_provider_top3_summary(model)
        assert "Alpha" in summary and "50.0%" in summary
        assert "Beta" in summary and "30.0%" in summary
        assert "Gamma" in summary and "15.0%" in summary
        assert "Top3 合计占已展示承接量 95.0%。" in summary
        assert "Delta" not in summary

    def test_build_provider_top3_summary_no_shown_rows(self):
        from pipeline.core_models_dashboard import (
            PROVIDER_NO_DATA_TEXT,
            _build_provider_top3_summary,
        )

        model = {
            "dict_rows": [
                {"展示状态": "用量少未查询", "Provider": "Alpha", "Provider 承接用量": "50B"},
            ],
        }
        assert _build_provider_top3_summary(model) == PROVIDER_NO_DATA_TEXT

    def test_build_provider_top3_price_insight_compares_monday_to_prior_sunday(self):
        from datetime import date
        from pipeline.core_models_dashboard import _build_provider_top3_price_insight

        week_start = date(2026, 5, 18)
        prior_sunday = "2026-05-17"
        model = {
            "days": ["2026-05-18"],
            "dict_rows": [
                {
                    "日期": "2026-05-18",
                    "Provider": "Alpha",
                    "展示状态": "已展示",
                    "Provider 承接用量": "50B",
                    "Input Price": 0.9,
                    "Output Price": 2.0,
                    "Cache Read": 0.1,
                },
                {
                    "日期": "2026-05-18",
                    "Provider": "Beta",
                    "展示状态": "已展示",
                    "Provider 承接用量": "30B",
                    "Input Price": 1.5,
                    "Output Price": 3.0,
                    "Cache Read": 0.2,
                },
            ],
        }
        prior_model = {
            "days": [prior_sunday],
            "dict_rows": [
                {
                    "日期": prior_sunday,
                    "Provider": "Alpha",
                    "展示状态": "已展示",
                    "Provider 承接用量": "40B",
                    "Input Price": 1.0,
                    "Output Price": 2.0,
                    "Cache Read": 0.1,
                },
                {
                    "日期": prior_sunday,
                    "Provider": "Beta",
                    "展示状态": "已展示",
                    "Provider 承接用量": "20B",
                    "Input Price": 1.5,
                    "Output Price": 3.0,
                    "Cache Read": 0.2,
                },
            ],
        }
        insight = _build_provider_top3_price_insight(
            model,
            week_start=week_start,
            prior_model=prior_model,
        )
        assert "Alpha 本周一 05-18 较上周日（Input $1.0000→$0.9000）" in insight
        assert "Beta 本周一较上周日价格未变" in insight

    def test_build_provider_top3_price_insight_includes_non_top3_price_movers(self):
        from datetime import date
        from pipeline.core_models_dashboard import _build_provider_top3_price_insight

        week_start = date(2026, 5, 25)
        prior_sunday = "2026-05-24"
        model = {
            "days": ["2026-05-25"],
            "dict_rows": [
                {
                    "日期": "2026-05-25",
                    "Provider": "Alpha",
                    "展示状态": "已展示",
                    "Provider 承接用量": "100B",
                    "Input Price": 1.0,
                    "Output Price": 2.0,
                    "Cache Read": 0.1,
                },
                {
                    "日期": "2026-05-25",
                    "Provider": "Beta",
                    "展示状态": "已展示",
                    "Provider 承接用量": "90B",
                    "Input Price": 1.0,
                    "Output Price": 2.0,
                    "Cache Read": 0.1,
                },
                {
                    "日期": "2026-05-25",
                    "Provider": "Gamma",
                    "展示状态": "已展示",
                    "Provider 承接用量": "80B",
                    "Input Price": 1.0,
                    "Output Price": 2.0,
                    "Cache Read": 0.1,
                },
                {
                    "日期": "2026-05-25",
                    "Provider": "NovitaAI",
                    "展示状态": "已展示",
                    "Provider 承接用量": "8B",
                    "Input Price": 1.6,
                    "Output Price": 3.38,
                    "Cache Read": 0.135,
                },
            ],
        }
        prior_model = {
            "days": [prior_sunday],
            "dict_rows": [
                {
                    "日期": prior_sunday,
                    "Provider": "NovitaAI",
                    "展示状态": "未展示",
                    "Provider 承接用量": "",
                    "Input Price": 1.64,
                    "Output Price": 3.38,
                    "Cache Read": 0.135,
                },
            ],
        }
        insight = _build_provider_top3_price_insight(
            model,
            week_start=week_start,
            prior_model=prior_model,
        )
        assert "NovitaAI 本周一 05-25 较上周日（Input $1.6400→$1.6000）" in insight


class TestModelActivityRetry:
    def test_should_retry_transport_and_rate_limit_errors(self):
        from pipeline.model_activity import _should_retry
        import httpx

        request = httpx.Request("GET", "https://openrouter.ai/api/frontend/models")
        response = httpx.Response(503, request=request)
        assert _should_retry(httpx.HTTPStatusError("service unavailable", request=request, response=response))
        assert _should_retry(httpx.ConnectError("EOF occurred in violation of protocol"))
        assert not _should_retry(httpx.HTTPStatusError(
            "not found",
            request=request,
            response=httpx.Response(404, request=request),
        ))

    def test_get_json_with_retry_recovers_after_transient_failure(self, monkeypatch):
        from pipeline import model_activity as activity

        calls = {"count": 0}

        class FakeResponse:
            def raise_for_status(self):
                return None

            def json(self):
                return {"data": {"analytics": []}}

        def fake_get(*_args, **_kwargs):
            calls["count"] += 1
            if calls["count"] == 1:
                raise activity.httpx.ConnectError("EOF occurred in violation of protocol")
            return FakeResponse()

        monkeypatch.setattr(activity.httpx, "get", fake_get)
        monkeypatch.setattr(activity.time, "sleep", lambda _seconds: None)

        payload = activity._get_json_with_retry("https://example.com/test", context="Test")
        assert payload == {"data": {"analytics": []}}
        assert calls["count"] == 2
